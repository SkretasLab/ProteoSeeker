import re
import os
import sys
import csv
import copy
import time
import math
import shutil
import string
import random
import subprocess
import pdfiltering
import sam_analysis
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook


def help_par(text, step):
    words = text.split(" ")
    lines = [words[0]]
    for word in words[1:]:
        if len(lines[-1]) + len(word) < step:
            lines[-1] += (" " + word)
        else:
            lines.append(word)
    return lines


def help_message():
    print("ProteoSeeker Version 1.0.0")
    print()
    print("python enzann.py [options] -i <Input_Folder_Path>")
    print()
    print("Option description:")
    print("1. Parameter type")
    print("2. Req: Required, Opt: Optional")
    print("3. Default value (shown if not empty or none)")
    print("4. Description")
    print("Terminology:")
    print("spd: seek profile database")
    print("tpd: taxonomy profile database")
    print("sfpd: seek filtered protein database")
    print("tfpd: taxonomy filtered protein database")
    print()
    print("Options:")
    print("---------Input and output options---------")
    help_notes_dict = {
        "-i/--input": "Str -Req- The path of the folder with the input files. The input files can either be single-end or paired-end FASTQ files but not both.",
        "-sc/--sra-code": "Str -Opt- A RUN accession code from the SRA database of NCBI.",
        "-c/--contigs": "True/False -Opt: False- Indicates whether the files in the input folder are (non-compressed) files with contigs or genomes in FASTA format.",
        "-pi/--protein-input": "True/False -Opt: False- Indicates whether the input folder contains a file with protein sequences in FASTA format.",
        "-a/--adapters": "Str -Req: adapters.fa- Path to the file with the adapters.",
        "-pdp/--protein-database-path": "Str -Opt- Path to the protein database file.",
        "-kdp/--kraken-database-path": "Str -Opt- Path to the kraken database folder.",
        "-psp/--profiles-seek-path": "Str -Opt- Path to the seek profile database with the profiles associated with one or more protein families.",
        "-pyp/--profiles-phylo-path": "Str -Opt- Path to the phylo profile database with the profiles associated with one or more protein families.",
        "-pbp/--profiles-broad-path": "Str -Opt- Path to the profile database with the profiles to be searched in the each protein identified with at least one profile from the seek profile database.",
        "-sp/--swissprot-path": "Str -Opt- Path to the Swiss-Prot protein database.",
        "-mp/--motifs-path": "Str -Opt- Path to the file with the motifs.",
        "-pfp/--parameters-file-path": "Str -Opt- The path to the file with the parameters and their values.",
        "-o/--output": "Str -Opt- Path to the output folder.",
        "-fc/--family-code": "Str -Opt- The seek protein family codes.",
        "-fct/--family-code-taxonomy": "Str -Opt- The phylo protein family codes.",
        "-dn/--database-name": "Str -Opt- The seek profile database (spd) and seek filtered protein database name (sfdp).",
        "-dnt/--database-name-taxonomy": "Str -Opt- The taxonomy profile database (tpd) and taxonomy filtered protein database name (tfdp).",
        "-sns/--seek-names-status": "True/False -Opt: False- Determines whether the protein names used to filter the protein database and create the sfpd will be determined solely based on the protein names provided by the user (True) or solely based on protein names automatically identified with or without the addition protein names provided by the user (False).",
        "-spn/--seek-protein-names": "Str -Opt- Protein names, divided by commas, given as input from the user and used to filter the seek protein database. If such protein names are indeed provided and -sns is False, then these protein names are added to the automatically identified ones. If such protein names are not provided and -sns is False, then only the aumatically identified ones are used for the filtering.",
        "-tns/--taxonomy-names-status": "True/False -Opt: False- Determines whether the protein names used to filter the protein database and create the tfpd will be determined solely based on the protein names provided by the user (True) or solely based on protein names automatically identified with or without the addition protein names provided by the user (False).",
        "-tpn/--taxonomy-protein-names": "Str -Opt- Protein names, divided by commas, given as input from the user and used to filter the taxonomy protein database. If such protein names are indeed provided and -tns is False, then these protein names are added to the automatically identified ones. If such protein names are not provided and -tns is False, then only the aumatically identified ones are used for the filtering.",
        "-nt/--name-threshold": "Float -Opt: 0.5- The threshold used to filter the protein names associated with each protein family. Any protein name with a frequency below this threshold is omitted.",
        "-sr/--seek-route": "Int -Opt: 3- There are three analysis modes. The analysis mode determines the type of analysis by the seek functionality. '1': The seek functionality will only search for proteins to be annotated which include at least one of the profiles of the spd. '2': The seek functionality will only search for proteins to be annotated that have at least one hit against the sfpd with a low enough e-value. '3': The seek functionality includes both types of analysis '1' and '2'.",
        "-p/--paired-end": "True/False -Opt: True- Indicates whether the files in the input folder are paired-end (True) or single-end (False) files.",
        "-k/--compressed": "True/False -Opt: True- Indicates whether the files in the input folder are compressed (True) or not (False).",
        "-fpd/--filter-protein-database": "True/False -Opt: False- Determines whether the protein database will be filtered based on protein names to create the sfpd and tfpd.",
        "-ps/--preftech-size": "Int -Opt: 20- The maximum file size to download in KB (K for kilobytes, M for megabytes, G gigabytes).",
        "-as/--adapters-status": "Str -Opt: 'pre'- The following options are available: ide: Adds the overrepresented sequences identified by FastQC in the file with the adapters. 'fas': The file with the adapters will include only the overrepresented sequences identified by FastQC. 'pre': The  file with the adapters is used without any modification.",
        "-asi/--add-seek-info": "True/False -Opt: True- Determines whether the results in the TXT and the EXCEL file will only contain information for the proteins identified through the seek mode (True) or not (False). In case only the taxonomy mode is applied, this option has no effect on the results.",
        "-ati/--add-taxonomy-info": "True/False -Opt: True- Determines whether the results in the TXT and the EXCEL file will only contain information for the proteins characterized through the taxonomy mode (True) or not (False). The latter proteins are the ones encoded by genes which are part of contigs that are grouped in bins, which bins have also been associated with at least one species. In case only the seek mode is applied, this option has no effect on the results.",
        "-h/--help": "None -- Displays the help message.",
        "-sf/--skip-fastqc": "True/False -Opt: False- Determines whether the second time the FastQC analysis is applied, at the preprocessed reads, will be omitted (True) or not (False).",
        "-umr/--bbduk-max-ram": "Int -Opt 4: False- The maximum number of GBs of RAM that may be utilized by BBDuk.",
        "-cs/--clear-space": "True/False -Opt: False- Determines whether the compressed preprocessed reads (if any) will be deleted (True) before the assembly or not (False).",
        "-kl/--k-list": "Str -Opt- A list of k-mers (e.g., 15,17,21,29,39,59,79,99,119,141) to be used by Megahit.",
        "-km/--kraken-mode": "True/False -Opt: True- Determines whether the taxonomy functionality will be based on the taxonomy analysis applied by kraken2 (True) or not (False).",
        "-kt/--kraken-threshold": "Float/Int -Opt: 0.1- A list with read-filtering threshold for the species reported by kraken. The list should include integers of floats seperated by commads. An integer is used as an absolute read threshold and a float is used as a percentage threshold applied to the percentagies reported by kraken for each species (e.g., 100 to represent a threshold of 100 reads, 1 to represent a threshold of 1 read, 1.0 to represent a threshold of 1%, 12.5 to represent a threshold of 12.5%). In addition, the values of -1 or -2 can be provided, to automatically set the threshold. For the value of -1 the threshold is set specifically for non-gut metagenomes and for the value of -2 the threshold is set specifically for gut metagenomes. When kraken is selected a binning process takes place based on the filtered species from the results of kraken. The latter binning process is based on the filtering performed based on the first threshold value of the list (if not only one).",
        "-kmm/--kraken-memory-mapping": "True/False -Opt: True- Determines whether kraken2 will use memory mapping (True) or not (False). With memory mapping the analysis performed by kraken2 is slower but is not limited by the size of the RAM available at the time of the analysis, rather than by the free memory space of the disk. Without memory mapping the analysis performed by kraken2 is faster but is limited by the size of the RAM available at the time of the analysis.",
        "-bt/--binning-tool": "Int -Opt: 1- Determines the binning tool to be used by the functionality of taxonomy, when kraken2 is set not to be used (-km False). '1': MetaBinner. '2': COMEBin.",
        "-bmr/--binning-max-ram": "Int -Opt: 4- The maximum number of GBs of RAM that may be utilized by binning.",
        "-bc/--bin-contig-len": "Int -Opt: 500- The threshold for filtering the contigs based on their lengths before binning. Any contig with length below the threshold is omitted from the binning process.",
        "-bk/--bin-kmer": "Int -Opt: 4- The number of kmers to be used by the binning tool (MetaBinner or COMEBin).",
        "-cbs/--comebin-batch-size": "Int -Opt: 256- The batch size to be used for the analysis of COMEBin.",
        "-ct/--cdhit-threshold": "Float -Opt: 0.99- The threshold used by CD-HIT. The value must be a float number between 0 and 1.",
        "-cmr/--cd-hit-max-ram": "Int -Opt: 4000- The maximum number of MBs of RAM that may be utilized by CD-HIT.",
        "-ge/--gene-encoding": "Int -Opt: 1- Determines whether the proteins will be provided directly from the output of FragGeneScanRs (1) or will be the output (2) from applying the genetic code indicated by option -gc to encode the predicted genes from FragGeneScanRs.",
        "-gc/--genetic-code": "Int -Opt: 11- The genetic code to be used to encode the genes predicted by FragGeneScanRs to proteins, if such an action has been selected (-ge 2).",
        "-st/--score-type": "Str -Opt cut_ga- The scoring method used by HMMER. 'cut_ga': HMMER will use the GA gathering cutoffs of the profile to set all thresholding. 'default': HMMER will use its default scoring method.",
        "-sds/--second-domain-search": "True/False -Opt: True- Determines whether the screening of the proteins by hmmscan of HMMER against the Pfam database will be applied (True) or not (False) during the seek functionality.",
        "-ndt/--no-domains-thr": "Int -Opt: 70- The negative of this number becomes the power of 10 and the result is the e-value threshold used to retain proteins during analysis mode 2 in the seek functionality for further annotation.",
        "-at/--add-type": "Str -Opt- A comma-seperated list which includes kinds of information related to the analysis. The items of the list are added for each protein in each of the annotation files.",
        "-ai/--add-info": "Str -Opt- A comma-seperated list which includes values for the corresponding kinds of information set by -at. The items of the list are added for each protein in each of the annotation files.",
        "-t/--threads": "Int -Opt: 4- The maximum number of threads to be used by any of the processes used by ProteoSeeker.",
        "-ft/--filtering-threads": "Int -Opt: -t- The maximum number of threads to be used by the filtering process of the protein database by ProteoSeeker. If not modified, the default value is equal to the value given at -t. Otherwise, it overwrites the value of -t.",
        "-afp/--after-peprocessing": "True/False -Opt: False- The pipeline start after the preprocessing of the reads.",
        "-afa/--after-assembly": "True/False -Opt: False- The pipeline starts after the assembly of the reads. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-afg/--after-gene-pred": "True/False -Opt: False- The pipeline starts after gene prediction. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-afb/--after-binning": "True/False -Opt: False- The pipeline starts after binning. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-adb/--after-after-db": "True/False -Opt: False- The pipeline starts after screening against the filtered protein database in both cases of the seek and taxonomy functionalities. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-atp/--after-topology-prediction": "True/False -Opt: False- The pipeline starts after the topology prediction in the case of the seek functionality. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-afr/--after-analysis-processes": "True/False -Opt: False- The pipeline starts after all the analysis processes and only writes the annotation files for the proteins, both cases of the seek and taxonomy functionalities. The files generated by the previous steps should be present in the output folder provided by the user.",
        "-uts/--up-to-sra": "True/False -Opt: False- The pipeline ends after downloading and processing the sample corresponding to the SRA code.",
        "-utd/--up-to-databases": "True/False -Opt: False- The pipeline ends after creating the seek and taxonomy profile databases and the seek and taxonomy filtered protein databases are created.",
        "-utpc/--up-to-preprocessing-com": "True/False -Opt: False- The pipeline ends after the preprocessing of the reads.",
        "-utpu/--up-to-preprocessing-uncom": "True/False -Opt: False- The pipeline ends after decompressing the compressed preprocessed reads.",
        "-uta/--up-to-assembly": "True/False -Opt: False- The pipeline ends after the assembly of the reads.",
        "-sen/--sra-env": "Str -Opt: ps_sra_tools- The conda enviroment for sra tools. 'None/none': To not use an enviroment at all.",
        "-fen/--fastqc-env": "Str -Opt: ps_fastqc- The conda enviroment for FastQC. 'None/none': To not use an enviroment at all.",
        "-uen/--bbtools-env": "Str -Opt: ps_bbtools- The conda enviroment for bbtools. 'None/none': To not use an enviroment at all.",
        "-men/--megahit-env": "Str -Opt: ps_megahit- The conda enviroment for megahit. 'None/none': To not use an enviroment at all.",
        "-ken/--kraken-env": "Str -Opt: ps_kraken- The conda enviroment for kraken2. 'None/none': To not use an enviroment at all.",
        "-nen/--metabinner-env": "Str -Opt: ps_metabinner- The conda enviroment for MetaBinner. 'None/none': To not use an enviroment at all.",
        "-sen/--comebin-env": "Str -Opt: ps_comebin- The conda enviroment for sra tools. 'None/none': To not use an enviroment at all.",
        "-ien/--cdhit-env": "Str -Opt: ps_cd_hit- The conda enviroment for CD-HIT. 'None/none': To not use an enviroment at all.",
        "-gen/--genepred-env": "Str -Opt- The conda enviroment for FragGeneScanRs. 'None/none': To not use an enviroment at all.",
        "-hen/--hmmer-env": "Str -Opt: ps_hmmer- The conda enviroment for HMMER. 'None/none': To not use an enviroment at all.",
        "-den/--dimaond-env": "Str -Opt: ps_diamond- The conda enviroment for DIMAOND BLASTP. 'None/none': To not use an enviroment at all.",
        "-ten/--taxonkit-env": "Str -Opt: ps_taxonkit- The conda enviroment for taxonkit. 'None/none': To not use an enviroment at all.",
        "-pen/--phobius-env": "Str -Opt: ps_phobius- The conda enviroment for Phobius. 'None/none': To not use an enviroment at all.",
        "-ben/--bowtie-env": "Str -Opt: ps_bowtie- The conda enviroment for Bowtie2. 'None/none': To not use an enviroment at all.",
        "-adp/--anaconda-dir-path": "Str -Opt- The path to the anaconda installation directory. This directory includes directories like \"bin\" and \"etc\".",
        "-asp/--anaconda-sh-path": "Str -Opt- The path to conda.sh. If provided the path to conda.sh will not be automatically determined by the path to the conda installation directory (-adp).",
        "-rfp/--prefetch-path": "Str -Opt- The path to the prefetch executable.",
        "-vvp/--vdb_validate-path": "Str -Opt- The path to the vdb-validate executable.",
        "-fdp/--fastq-dump-path": "Str -Opt- The path to the fastq-dump executable.",
        "-fp/--fastqc-path": "Str -Opt- The path to the fastqc executable.",
        "-gzp/--gzip-path": "Str -Opt- The path to the gzip executable.",
        "-ctp/--cat-path": "Str -Opt- The path to the cat executable.",
        "-bdp/--bbduk-path": "Str -Opt- The path to the bbduk executable.",
        "-mp/--megahit-path": "Str -Opt- The path to the megahit executable.",
        "-kp/--kraken-path": "Str -Opt- The path to the kraken executable.",
        "-bfp/--binner-folder-path": "Str -Opt- The path to the bin folder of MetaBiner.",
        "-cfp/--comebin-folder-path": "Str -Opt- The path to the bin folder of COMEBin.",
        "-chp/--cd-hit-path": "Str -Opt- The path to the CD-HIT executable.",
        "-fgp/--fraggenescars-path": "Str -Opt- The path to the FragGeneScanRs executable.",
        "-hp/--hmmscan-path": "Str -Opt- The path to the hmmscan executable.",
        "-hp/--hmmscan-path": "Str -Opt- The path to the hmmscan executable.",
        "-hpp/--hmmpress-path": "Str -Opt- The path to the hmmpress executable.",
        "-hfp/--hmmfetch-path": "Str -Opt- The path to the hmmfetch executable.",
        "-dp/--diamond-path": "Str -Opt- The path to the diamond executable.",
        "-tkp/--taxonkit-path": "Str -Opt- The path to the taxonkit executable.",
        "-php/--phobius-folder-path": "Str -Opt- The path to the folder of phobius.",
        "-bbp/--bowtie-build-path": "Str -Opt- The path to the bowtie build executable.",
        "-hfp/--bowtie-path": "Str -Opt- The path to the bowtie executable."
    }
    split_step = 60
    for key_hn in help_notes_dict.keys():
        description = help_notes_dict[key_hn]
        des_length = len(description)
        if split_step >= des_length:
            print('{:2} {:30} {}\n'.format("", key_hn, description))
        else:
            pieces = help_par(description, split_step)
            for pmi in range(0, len(pieces)):
                piece_mod = pieces[pmi]
                if pmi == 0:
                    print('{:2} {:30} {}'.format("", key_hn, piece_mod))
                elif pmi + 1 == len(pieces):
                    print('{:33} {}\n'.format("", piece_mod))
                else:
                    print('{:33} {}'.format("", piece_mod))
        if key_hn  == "-o/--output":
            print("---------Protein family options---------")
        elif key_hn == "-nt/--name-threshold":
            print("---------General options: Pipeline---------")
        elif key_hn == "-h/--help":
            print("---------General options: FastQC---------")
        elif key_hn == "-sf/--skip-fastqc":
            print("---------General options: BBDuk---------")
        elif key_hn == "-cs/--clear-space":
            print("---------General options: Megahit---------")
        elif key_hn == "-kl/--k-list":
            print("---------General options: Kraken---------")
        elif key_hn == "-kmm/--kraken-memory-mapping":
            print("---------General options: Binning---------")
        elif key_hn == "-cbs/--comebin-batch-size":
            print("---------General options: CD-HIT---------")
        elif key_hn == "-cmr/--cd-hit-max-ram":
            print("---------General options: Gene prediction---------")
        elif key_hn == "-gc/--genetic-code":
            print("---------General options: HMMER---------")
        elif key_hn == "-ndt/--no-domains-thr":
            print("---------General options: Annotation---------")
        elif key_hn == "-ai/--add-info":
            print("---------Threads---------")
        elif key_hn == "-ft/--filtering-threads":
            print("---------Processes performed after---------")
        elif key_hn == "-afr/--after-analysis-processes":
            print("---------Processes performed up to---------")
        elif key_hn == "-uta/--up-to-assembly":
            print("---------Tool enviroments---------")
        elif key_hn == "-ben/--bowtie-env":
            print("---------Tool paths---------")


def check_i_value(i_value, option_str):
    if i_value in ["False", "false", "f", "0"]:
        i_value = False
    elif i_value in ["True", "true", "t", "1"]:
        i_value = True
    else:
        print("Wrong value given for the '{}' option. Exiting.".format(option_str))
        exit()
    return i_value


def read_file(file_path):
    file_handle = open(file_path, "r")
    pre_lines = file_handle.readlines()
    lines = []
    for i in pre_lines:
        i = i.rstrip("\n")
        lines.append(i)
    file_handle.close()
    return lines


def end_time_analysis(label, start_time, output_log_file):
    # End time
    end_time = time.time()
    elpased_time = end_time - start_time
    elpased_time_min = elpased_time / 60
    elpased_time_hour = elpased_time_min / 60
    elpased_time = round(elpased_time, 3)
    elpased_time_min = round(elpased_time_min, 3)
    elpased_time_hour = round(elpased_time_hour, 3)
    print("\n{}".format(label))
    print("Time elapsed: {} seconds / {} minutes / {} hours".format(elpased_time, elpased_time_min, elpased_time_hour))
    output_log_file.write("{}\n".format(label))
    output_log_file.write("Time elapsed: {} seconds / {} minutes / {} hours\n".format(elpased_time, elpased_time_min, elpased_time_hour))
    output_log_file.write("\n{}\n\n\n".format(100*"-"))
    return elpased_time


def write_time_dict(time_dict, time_analyis_path):
    time_analysis_file = open(time_analyis_path, "w")
    for key in time_dict.keys():
        temp_line = "{}\t{}".format(key, time_dict[key])
        time_analysis_file.write("{}\n".format(temp_line))
    time_analysis_file.close()


def translation(sequence, genetic_code):
    translation_table_1 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_2 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "W", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "M", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "*", "AGG": "*", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_3 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "W", "TGG": "W", "CTT": "T", "CTC": "T",
        "CTA": "T", "CTG": "T", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "M", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_4 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "W", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_5 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "W", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "M", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "S", "AGG": "S", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_11 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "L", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_table_12 = {
        "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S", "TAT": "Y",
        "TAC": "Y", "TAA": "*", "TAG": "*", "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W", "CTT": "L", "CTC": "L",
        "CTA": "L", "CTG": "S", "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P", "CAT": "H", "CAC": "H", "CAA": "Q",
        "CAG": "Q", "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R", "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
        "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T", "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K", "AGT": "S",
        "AGC": "S", "AGA": "R", "AGG": "R", "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "GCT": "A", "GCC": "A",
        "GCA": "A", "GCG": "A", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E", "GGT": "G", "GGC": "G", "GGA": "G",
        "GGG": "G"
    }
    translation_tables = {
        1: translation_table_1,
        2: translation_table_2,
        3: translation_table_3,
        4: translation_table_4,
        5: translation_table_5,
        11: translation_table_11,
        12: translation_table_12
    }
    protein = ""
    for i in range(0, len(sequence), 3):
        # If the whole sequence is less than 3 nucleotides, no tranlation occurs.
        if len(sequence) - 3 < i:
            continue
        else:
            # If the last nucleotides do not add up to three, then no tranlation occurs for the last 1 or 2 nucleotides.
            if i+1 >= len(sequence):
                continue
            translation_table = translation_tables[genetic_code]
            tripeptide = sequence[i] + sequence[i+1] + sequence[i+2]
            # Capitalize all nucleotide letters
            tripeptide = tripeptide.upper()
            # If the tripeptide contains at least one N then directly transalte to the "X" amino acid
            if "N" in tripeptide:
                amino_acid = "X"
            else:
                amino_acid = translation_table[tripeptide]
            protein += amino_acid
    return protein


def command_run(phrase_1, phrase_2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file):
    # Process
    input_log_file.write("{}\n".format(title_1))
    if pr_status:
        print("\n{}".format(phrase_1))
    input_log_file.write("{}\n\n".format(phrase_1))
    if shell_status:
        command = phrase_1
    else:
        command = phrase_1.split(" ")
    proc = subprocess.run(command, capture_output=capture_status, shell=shell_status)
    stdout_raw = proc.stdout
    stdout_str = stdout_raw.decode("utf-8")
    stdout_lines = stdout_str.split("\n")
    for line in stdout_lines:
        output_log_file.write("{}\n".format(line))
    # Version, if available through the command-line tool.
    input_log_file.write("{}\n".format(title_2))
    if phrase_2 and phrase_2 != "":
        input_log_file.write("{}\n\n".format(phrase_2))
        if shell_status:
            command = phrase_2
        else:
            command = phrase_2.split(" ")
        proc = subprocess.run(command, capture_output=capture_status, shell=shell_status)
        stdout_raw = proc.stdout
        stdout_str = stdout_raw.decode("utf-8")
        stdout_lines = stdout_str.split("\n")
        output_log_file.write("Version information:\n")
        for line in stdout_lines:
            output_log_file.write("{}\n".format(line))
    output_log_file.write("\n{}\n\n\n".format(100*"-"))


def paired_end_file_paths(folder, ignore_comp=True):
    file_paths_p = {}
    passed = []
    r = 0
    passed_counter = 0
    pre_file_paths = os.listdir(folder)
    for i in range(0, len(pre_file_paths)):
        if "Summary" in pre_file_paths[i] or "Error" in pre_file_paths[i]:
            continue
        if ignore_comp:
            if (pre_file_paths[i][-3:] == ".gz") or (pre_file_paths[i][-3:] == ".sh") or (pre_file_paths[i][-4:] == ".txt"):
                continue
        passed_counter += 1
        if i in passed:
            continue
        # The last file checked should always have been found earlier as the pair of another file. Thus it should always be in the list passed. But in case a mistake has been made in the names of the files
        # then it may not be in the list passed and the counter will be increased thus in the end an error will occur based on the difference of the length of the list passed and of the counter.
        if i + 1 == len(pre_file_paths):
            break
        for j in range(i + 1, len(pre_file_paths)):
            if len(pre_file_paths[i]) != len(pre_file_paths[j]):
                continue
            dif = 0
            dif_chars = []
            for char_index in range(0, len(pre_file_paths[i])):
                if pre_file_paths[i][char_index] != pre_file_paths[j][char_index]:
                    dif += 1
                    dif_chars.append(pre_file_paths[i][char_index])
                    dif_chars.append(pre_file_paths[j][char_index])
            if dif == 1:
                if len(dif_chars) == 2:
                    if "1" in dif_chars and "2" in dif_chars:
                        passed.append(i)
                        passed.append(j)
                        splited_1_a = pre_file_paths[i].split("_")
                        splited_1_b = splited_1_a[-1].split(".")
                        temp_path_1 = "{}/{}".format(folder, pre_file_paths[i])
                        temp_path_2 = "{}/{}".format(folder, pre_file_paths[j])
                        if splited_1_b[0] == "1":
                            file_paths_p[r] = [temp_path_1, temp_path_2]
                        else:
                            file_paths_p[r] = [temp_path_2, temp_path_1]
                        r += 1
    if len(passed) != passed_counter:
        print("Error. Not all files are paired-end reads or not all files could be mapped as paired-end reads due to their names being different in more than one character, which is expected to be number 1 and number 2.")
        exit()
    return file_paths_p


def file_paths_creation(input_folder, paired_end):
    file_paths = []
    file_paths_p = {}
    if input_folder[-1] == "/":
        input_folder = input_folder[:-1]
    # If the files are paired-end reads then except of the list file_paths which has the names of the files there is also a dictionary
    # which has as values lists with two paired end reads each.
    pre_file_paths = os.listdir(input_folder)
    for i in pre_file_paths:
        local_file_path = "{}/{}".format(input_folder, i)
        file_paths.append(local_file_path)
    if paired_end:
        file_paths_p = paired_end_file_paths(input_folder, False)
    return file_paths, file_paths_p
                

def locate_nc_files(paired_end, output_path_trimmed):
    tr_ex_file_paths_p = {}
    tr_ex_file_paths = []
    if paired_end:
        tr_ex_file_paths_p = paired_end_file_paths(output_path_trimmed, True)
    else:
        pre_tr_ex_file_paths = os.listdir(output_path_trimmed)
        for i in pre_tr_ex_file_paths:
            if "fastq" in i and i[-2:] != "gz":
                local_file_path = "{}/{}".format(output_path_trimmed, i)
                tr_ex_file_paths.append(local_file_path)
    return tr_ex_file_paths_p, tr_ex_file_paths


def craete_phmm_db(family_code, family_group_name, enzyme_domains_folder, fam_nums_file_name, fam_pfam_file_name, pfam_domains_names, hmmer_env, hmmfetch_path, hmmpress_path, profiles_broad_path, input_log_file, output_log_file, conda_sh_path):
    # Initialization of variables.
    d_unique = None
    profiles_name = None
    analysis_fam_names = []
    family_to_profile_dict = {}
    # Folder and file paths. Some are initialized only if at least one protein family was selected.
    phmm_database_folder_name = "{}/profile_dbs/phmm_db_{}".format(enzyme_domains_folder, family_group_name)
    fam_prof_info_path = "{}/family_profile_info.tsv".format(phmm_database_folder_name)
    if family_group_name is not None:
        new_file_dom_codes_name = "{}/profile_codes.txt".format(phmm_database_folder_name)
        new_file_dom_names_name = "{}/profile_names.txt".format(phmm_database_folder_name)
        profiles_name = "{}/profiles".format(phmm_database_folder_name)
        hmmer_db_bash_script = "{}/hmmer_db.sh".format(phmm_database_folder_name)
        hmmfetch_stdoe_path = "{}/hmmfetch_stdoe.txt".format(phmm_database_folder_name)
        hmmfetch_version_path = "{}/hmmfetch_version.txt".format(phmm_database_folder_name)
        hmmpress_stdoe_path = "{}/hmmpress_stdoe.txt".format(phmm_database_folder_name)
        hmmpress_version_path = "{}/hmmpress_version.txt".format(phmm_database_folder_name)
    # For each protein family selected...
    if family_code is not None:
        d_unique = []
        # Create the correspondances between the families and their numbering.
        fam_nums_lines = read_file(fam_nums_file_name)
        fam_num_dict = {}
        for line in fam_nums_lines:
            line_splited = line.split("\t")
            fam_num = line_splited[0]
            fam_name = line_splited[1]
            fam_num_dict[fam_num] = fam_name
        for family_code_cur in family_code:
            # Find the name of the selected protein family.
            # This must be changed to a dictionary and then search the protein length against the all mean legnth of each selected protein family and give out the family with the closest length.
            sel_fam = fam_num_dict[family_code_cur]
            print("Selected faimly: {}".format(sel_fam))
            analysis_fam_names.append(sel_fam)
            if sel_fam not in family_to_profile_dict.keys():
                family_to_profile_dict[sel_fam] = {}
            # The domain sets for each protein family are already those of the highest requency for the protein family.
            # When selecting multiple protein families, all protein sets from each protein family are collected, because the point is to
            # search for the domains corresponding to each of the selected protein families as to later on determine proteins that may belong
            # to any of the selected protein families.
            # Collect the domain and their counts, corresponding to the name of the selected protein family.
            # Collect each set of domains related to the protein family with their counts.
            # Collect the unique domains related to the protein family.
            pfam_lines = read_file(fam_pfam_file_name)
            for line in pfam_lines:
                line_splited = line.split("\t")
                line_fam = line_splited[0]
                if line_fam == sel_fam:
                    domain_sets = line_splited[1]
                    # Remove the first and last parenthesis of the outer list.
                    domain_sets = domain_sets[1:-1]
                    # Remove the first parenthesis of the first list and the last parenthesis of the last list.
                    # These lists might be the same.
                    domain_sets = domain_sets[1:-1]
                    # Now split the domains sets.
                    # If the domains sets are more than one then split them.
                    if "], [" in domain_sets:
                        # 'PF00541_1;PF00608_2'], ['PF00541_1'], ['PF00608_7'
                        domain_sets_splited = domain_sets.split("], [")
                        for ds in domain_sets_splited:
                            # Remove the single quotes.
                            ds = ds[1:-1]
                            # If the set of domains contains multiple domains split them.
                            if ";" in ds:
                                # PF00541_1;PF00608_2
                                ds_splited = ds.split(";")
                                # Extract the domain codes and their counters.
                                # ["PF00541_1", "PF00608_2"]
                                for item in ds_splited:
                                    item_splited = item.split("_")
                                    dname = item_splited[0]
                                    if dname not in d_unique:
                                        d_unique.append(dname)
                                    family_to_profile_dict[sel_fam][dname] = 0
                            else:
                                # PF00541_1
                                ds_splited = ds.split("_")
                                dname = ds_splited[0]
                                if dname not in d_unique:
                                    d_unique.append(dname)       
                                family_to_profile_dict[sel_fam][dname] = 0                 
                    else:
                        # 'PF01757_1;PF00541_1'
                        # Remove the single quotes.
                        domain_sets = domain_sets[1:-1]
                        # PF01757_1;PF00541_1
                        if ";" in domain_sets:
                            ds_splited = domain_sets.split(";")
                            # ["PF01757_1", "PF00541_1"]
                            for item in ds_splited:
                                item_splited = item.split("_")
                                dname = item_splited[0]
                                if dname not in d_unique:
                                    d_unique.append(dname)
                                family_to_profile_dict[sel_fam][dname] = 0
                        else:
                            # PF00541_1
                            ds_splited = domain_sets.split("_")
                            dname = ds_splited[0]
                            if dname not in d_unique:
                                d_unique.append(dname)
                            family_to_profile_dict[sel_fam][dname] = 0
        print("\nSelected domain codes: {}".format(d_unique))
        # If the path to the folder of the pHMM library does not exist, then create the folder and then the library.
        if not os.path.exists(phmm_database_folder_name):
            os.mkdir(phmm_database_folder_name)
            # Creating a file to store the families and their profiles.
            fam_prof_info_file = open(fam_prof_info_path, "w")
            for key_fam in family_to_profile_dict.keys():
                for key_prof in family_to_profile_dict[key_fam].keys():
                    fam_prof_info_file.write("{}\t{}\n".format(key_fam, key_prof))
            fam_prof_info_file.close()
            # Create dictionary of Pfam codes to names.
            # The Pfam code is correponded to the latest Pfam version of the Pfam code.
            pfam_names_doms_dict = {}
            pfam_doms_names_lines = read_file(pfam_domains_names)
            for line in pfam_doms_names_lines:
                splited_line = line.split("\t")
                dom_code = splited_line[0]
                dom_name = splited_line[1]
                if "." in dom_code:
                    dom_code_splited = dom_code.split(".")
                    dom_code = dom_code_splited[0]
                pfam_names_doms_dict[dom_code] = dom_name
            # Find the current Pfam codes and create a file with these codes.
            new_file_dom_codes = open(new_file_dom_codes_name, "w")
            new_file_dom_names = open(new_file_dom_names_name, "w")
            domains_selected = []
            for dom_code in d_unique:
                dom_name = pfam_names_doms_dict[dom_code]
                domains_selected.append(dom_name)
                new_file_dom_codes.write("{}\n".format(dom_code))
                new_file_dom_names.write("{}\n".format(dom_name))
            new_file_dom_codes.close()
            new_file_dom_names.close()
            # Createa a pHMM file with the corresponding profile names.
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(hmmer_db_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if hmmer_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(hmmer_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            if hmmfetch_path:
                phrase_1 = "\"{}\" -o \"{}\" -f \"{}\" \"{}\" &> \"{}\"".format(hmmfetch_path, profiles_name, profiles_broad_path, new_file_dom_names_name, hmmfetch_stdoe_path)
                phrase_2 = "\"{}\" \"{}\" &> \"{}\"".format(hmmpress_path, profiles_name, hmmpress_stdoe_path)
                phrase_3 = "\"{}\" -h > \"{}\"".format(hmmfetch_path, hmmfetch_version_path)
                phrase_4 = "\"{}\" -h > \"{}\"".format(hmmpress_path, hmmpress_version_path)
            else:
                phrase_1 = "hmmfetch -o \"{}\" -f \"{}\" \"{}\" &> \"{}\"".format(profiles_name, profiles_broad_path, new_file_dom_names_name, hmmfetch_stdoe_path)
                phrase_2 = "hmmpress \"{}\" &> \"{}\"".format(profiles_name, hmmpress_stdoe_path)
                phrase_3 = "hmmfetch -h > \"{}\"".format(hmmfetch_version_path)
                phrase_4 = "hmmpress -h > \"{}\"".format(hmmpress_version_path)
            new_file_bash.write("{}\n".format(phrase_1))
            new_file_bash.write("{}\n".format(phrase_2))
            new_file_bash.write("{}\n".format(phrase_3))
            new_file_bash.write("{}\n".format(phrase_4))
            if hmmer_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(hmmer_db_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(hmmer_db_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        else:
            if os.path.exists(fam_prof_info_path):
                fam_prof_info_lines = read_file(fam_prof_info_path)
                for line in fam_prof_info_lines:
                    line_splited = line.split("\t")
                    fam_temp = line_splited[0]
                    prof_temp = line_splited[1]
                    if fam_temp not in family_to_profile_dict.keys():
                        family_to_profile_dict[fam_temp] = {}
                    family_to_profile_dict[fam_temp][prof_temp] = 0
            print("\nThe pHMM database for this protein family already exists. Skipping creation of the pHMM database.")
    else:
        if os.path.exists(fam_prof_info_path):
            fam_prof_info_lines = read_file(fam_prof_info_path)
            for line in fam_prof_info_lines:
                line_splited = line.split("\t")
                fam_temp = line_splited[0]
                prof_temp = line_splited[1]
                if fam_temp not in family_to_profile_dict.keys():
                    family_to_profile_dict[fam_temp] = {}
                family_to_profile_dict[fam_temp][prof_temp] = 0
    return profiles_name, analysis_fam_names, family_to_profile_dict


def find_fam_names(family_code, family_group_name, create_nr_db_status, protein_db_path, fpr_db_gen_folder, fam_nums_file_name, name_thr, input_protein_names_status, input_protein_names, enzyme_names_dict, conda_sh_path):
    # Creation of the pnr database if the user selected the process to be performed and also provided a path for the nr database.
    fpr_db_fasta = ""
    fpr_db_name = ""
    fpr_db_folder = None
    fpr_db_fasta_prefix = ""
    enzyme_names = []
    if create_nr_db_status:
        # The name of the pnr database.
        # To create the pnr database the path to the nr database is needed.
        if os.path.exists(protein_db_path):
            fpr_db_folder = "{}/fpr_{}".format(fpr_db_gen_folder, family_group_name)
            fpr_db_fasta = "{}/fpr_{}.fasta".format(fpr_db_folder, family_group_name)
            fpr_db_fasta_prefix = "{}/fpr_{}".format(fpr_db_folder, family_group_name)
        else:
            print("\nThe path for the nr database has not been set or is wrong. Exiting.")
            exit()
        fpr_db_name = "{}/fpr_{}_database".format(fpr_db_folder, family_group_name)
        # If the pnr database already exists based on the selected family, do not create it again. It is already checked that, if no family code was
        # selected the pnr database does not exist based on the random counter.
        if not os.path.exists(fpr_db_folder):
            # Dictionary for protein family codes and protein family names.
            enzyme_code_famname_dict = {}
            enzyme_code_prsname_dict = {}
            fam_nums_lines = read_file(fam_nums_file_name)
            for line in fam_nums_lines:
                line_splited = line.split("\t")
                fam_num = line_splited[0]
                fam_name = line_splited[1]
                pr_names = line_splited[2:]
                enzyme_code_famname_dict[fam_num] = fam_name
                enzyme_code_prsname_dict[fam_num] = pr_names
            # Create the list of initial enzyme names.
            if family_code is not None:
                pre_enzyme_names = []
                if enzyme_code_famname_dict.keys():
                    for family_code_cur in family_code:
                        pre_enzyme_names_ori = enzyme_code_famname_dict[family_code_cur]
                        # For each family or subfamily case keep a name same as the the original text and a name without the "family" or "subfamily" word.
                        # AAA ATPase family. Katanin p60 subunit T.A1 subfamily. A-like 1 sub-subfamily
                        # pre_enzyme_names_ori_splited = ["AAA ATPase ", " Katanin p60 subunit T.A1 ", " A-like 1 sub-subfamily"]
                        # pre_enzyme_names_fam_words = ["AAA ATPase", "Katanin p60 subunit T.A1", "A-like 1 sub-subfamily"]
                        # pre_enzyme_names = pre_enzyme_names_fam_words
                        if ("family." in pre_enzyme_names_ori) or "subfamily." in pre_enzyme_names_ori:
                            pre_enzyme_names_fam_words = []
                            pre_enzyme_names_ori_splited = re.split('family.|subfamily.', pre_enzyme_names_ori)
                            for item_ori in pre_enzyme_names_ori_splited:
                                item_ori = item_ori.strip()
                                pre_enzyme_names_fam_words.append(item_ori)
                            pre_enzyme_names = copy.deepcopy(pre_enzyme_names_fam_words)
                            for item in pre_enzyme_names_fam_words:
                                if item[-10:] == " subfamily":
                                    item = item[:-10]
                                if item[-7:] == " family":
                                    item = item[:-7]
                                item = item.strip()
                                pre_enzyme_names.append(item)
                        else:
                            # Golgi pH regulator (TC 1.A.38) family
                            # pre_enzyme_names_ori = "Golgi pH regulator (TC 1.A.38) family"
                            # pre_enzyme_names = ["Golgi pH regulator (TC 1.A.38) family"]
                            # if pre_enzyme_names_ori[-10:] == " subfamily": False
                            # pre_enzyme_names_ori[-7:] == " family": True
                            # item = "Golgi pH regulator (TC 1.A.38)"
                            # pre_enzyme_names = ["Golgi pH regulator (TC 1.A.38) family", "Golgi pH regulator (TC 1.A.38)"]
                            pre_enzyme_names = [pre_enzyme_names_ori]
                            if pre_enzyme_names_ori[-10:] == " subfamily":
                                item = pre_enzyme_names_ori[:-10]
                            if pre_enzyme_names_ori[-7:] == " family":
                                item = pre_enzyme_names_ori[:-7]
                            item = item.strip()
                            pre_enzyme_names.append(item)
                for family_code_cur in family_code:
                    # Add each protein name based on a threhsold of frequency.
                    # Find the highest occuring name.
                    if name_thr <= 0:
                        print("\nThe threshold for the naming selection of the protein family can not be 0 or negative. Exiting.")
                        exit()
                    if isinstance(name_thr, float):
                        name_max_splited =  enzyme_code_prsname_dict[family_code_cur][0].split("::")
                        name_max = int(name_max_splited[1])
                        sel_thr = name_max * name_thr
                    elif isinstance(name_thr, int):
                        sel_thr = float(name_thr)
                    else:
                        print("\nAn error occured while computing the protein naming threshold. Exiting.")
                        exit()
                    print("Protein family: {} - Protein name frequency threshold: {}".format(family_code_cur, sel_thr))
                    for prn_cur in  enzyme_code_prsname_dict[family_code_cur]:
                        prn_cur_splited = prn_cur.split("::")
                        prn_cur_name = prn_cur_splited[0]
                        prn_cur_num = float(prn_cur_splited[1])
                        if prn_cur_num >= sel_thr:
                            pre_enzyme_names.append(prn_cur_name)
                # If selected the input enzyme names overwritten the ones found automatically (if any were found).
                # If any enzyme name have been given as input, they are added in the list of names for the selected enzyme
                # code.
                if not input_protein_names_status:
                    if input_protein_names:
                        enzyme_names = pre_enzyme_names + input_protein_names
                    else:
                        enzyme_names = copy.deepcopy(pre_enzyme_names)
                else:
                    enzyme_names = copy.deepcopy(input_protein_names)
                # Deduplicate list of enzyme names.
                enzyme_names_dupl = copy.deepcopy(enzyme_names)
                enzyme_names = []
                for item in enzyme_names_dupl:
                    if item not in enzyme_names:
                        enzyme_names.append(item)
        else:
            print("\nThe path to the fpr database ({}) for the selected protein family(ies) already exists. Skipping creation of the fpr database.".format(fpr_db_folder))
    temp_list = [enzyme_names, fpr_db_fasta, fpr_db_folder, fpr_db_name, fpr_db_fasta_prefix]
    if enzyme_names_dict:
        enzyme_names_dict[1] = copy.deepcopy(temp_list)
    else:
        enzyme_names_dict[0] = copy.deepcopy(temp_list)
    return fpr_db_fasta, fpr_db_name, fpr_db_folder, enzyme_names_dict


def process_nr_db(fpr_db_folder, diamond_db_bash_name, diamond_env, diamond_path, fpr_db_fasta, fpr_db_name, diamond_makedb_version_path, diamond_makedb_stdoe_path, thread_num, input_log_file, output_log_file, conda_sh_path):
    # Create the database with "makeblstdb", else with "diamond".
    # Create the Bash script.
    # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
    diamond_db_bash_script = "{}/{}".format(fpr_db_folder, diamond_db_bash_name)
    new_file_bash = open(diamond_db_bash_script, "w")
    phrase = "#!/bin/bash"
    new_file_bash.write("{}\n".format(phrase))
    phrase_1 = "cat \"{}/\"*.fasta > \"{}\"".format(fpr_db_folder, fpr_db_fasta)
    if diamond_env:
        phrase_s = "source \"{}\"".format(conda_sh_path)
        phrase_a = "conda activate \"{}\"".format(diamond_env)
        new_file_bash.write("{}\n".format(phrase_s))
        new_file_bash.write("{}\n".format(phrase_a))
    if diamond_path:
        phrase_2 = "\"{}\" makedb --threads {} --in \"{}\" --db \"{}\" &> \"{}\"".format(diamond_path, thread_num, fpr_db_fasta, fpr_db_name, diamond_makedb_stdoe_path)
        phrase_3 = 'echo "No version available for diamond makedb." > {}'.format(diamond_makedb_version_path)
    else:
        phrase_2 = "diamond makedb --threads {} --in \"{}\" --db \"{}\" &> \"{}\"".format(thread_num, fpr_db_fasta, fpr_db_name, diamond_makedb_stdoe_path)
        phrase_3 = 'echo "No version available for diamond makedb." > \"{}\"'.format(diamond_makedb_version_path)
    new_file_bash.write("{}\n".format(phrase_1))
    new_file_bash.write("{}\n".format(phrase_2))
    new_file_bash.write("{}\n".format(phrase_3))
    if diamond_env:
        phrase = "conda deactivate"
        new_file_bash.write("{}\n".format(phrase))
    new_file_bash.close()
    # Making the Bash script executable.
    # Sending command to run.
    phrase_b1 = "chmod +x {}".format(diamond_db_bash_script)
    phrase_b2 = "chmod --version"
    title_1 = "Making a Bash script executable:"
    title_2 = "Version of chmod:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Sending command to run.
    phrase_b1 = "bash {}".format(diamond_db_bash_script)
    phrase_b2 = "bash --version"
    title_1 = "Running bash:"
    title_2 = "Version of bash:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Deleting part fasta files.
    pd_folder_files = os.listdir(fpr_db_folder)
    for pd_file in pd_folder_files:
        pd_file_path = "{}/{}".format(fpr_db_folder, pd_file)
        if "_part_" in pd_file:
            os.remove(pd_file_path)
    

def create_nr_db(enzyme_names_dict, protein_db_path, diamond_db_bash_name, diamond_env, diamond_path, thread_num, pdf_threads, input_log_file, output_log_file, conda_sh_path):
    enz_names_analysis = enzyme_names_dict[0][0]
    enz_names_phylo = enzyme_names_dict[1][0]
    # Duplicate names are removed.
    if (not enz_names_analysis) and (not enz_names_phylo):
        print("\nNo names for the enzyme of the enzyme code selected were found. The nr database will not be filtered.")
    else:
        fpr_db_fasta = enzyme_names_dict[0][1]
        fpr_db_fasta_phylo = enzyme_names_dict[1][1]
        fpr_db_folder = enzyme_names_dict[0][2]
        fpr_db_folder_phylo = enzyme_names_dict[1][2]
        fpr_db_name = enzyme_names_dict[0][3]
        fpr_db_name_phylo = enzyme_names_dict[1][3]
        fpr_db_fasta_prefix = enzyme_names_dict[0][4]
        fpr_db_fasta_prefix_phylo  = enzyme_names_dict[1][4]
        # Create the folder for the database.
        if enz_names_analysis:
            os.mkdir(fpr_db_folder)
        if enz_names_phylo:
            os.mkdir(fpr_db_folder_phylo)
        print("\nNames collected to create the seek pnr: {}".format(enz_names_analysis))
        print("Names collected to create the phylo pnr: {}".format(enz_names_phylo))
        if enz_names_analysis:
            fpr_seeknames_path = "{}/seek_names.txt".format(fpr_db_folder)
            fpr_seeknames_file = open(fpr_seeknames_path, "w")
            for sn_item in enz_names_analysis:
                fpr_seeknames_file.write("{}\n".format(sn_item))
            fpr_seeknames_file.close()
        if enz_names_phylo:
            fpr_phylonames_path = "{}/phylo_names.txt".format(fpr_db_folder_phylo)
            fpr_phylonames_file = open(fpr_phylonames_path, "w")
            for pn_item in enz_names_phylo:
                fpr_phylonames_file.write("{}\n".format(pn_item))
            fpr_phylonames_file.close()
        info_file_path=""
        # Lines in nr: 2.706.460.600
        # Proteins in nr: 439.976.609
        print("\nCreating the fpr databases...")
        db_info_dict = {}
        temp_0 = [fpr_db_fasta_prefix, enz_names_analysis, False, None]
        temp_1 = [fpr_db_fasta_prefix_phylo, enz_names_phylo, False, None]
        # The log file is created in the folder of the seek database if both seek and phylo databases are to be created. Otherwise, the log file
        # is created in the seek or the phylo folder depending on which of the two functionalities has been selected.
        if enz_names_analysis and enz_names_phylo:
            log_file_path = "{}/pd_filter_log.txt".format(fpr_db_folder)
        elif enz_names_analysis:
            log_file_path = "{}/pd_filter_log.txt".format(fpr_db_folder)
        elif enz_names_phylo:
            log_file_path = "{}/pd_filter_log.txt".format(fpr_db_folder_phylo)
        # Database information
        if enz_names_analysis:
            db_info_dict[0] = temp_0
        if enz_names_phylo:
            db_info_dict[1] = temp_1
        pdfiltering.prfilter(protein_db_path, info_file_path, db_info_dict, log_file_path, pdf_threads)
    # Processing and creating the databases.
    if enz_names_analysis:
        diamond_makedb_version_path = "{}/diamond_makedb_version.txt".format(fpr_db_folder)
        diamond_makedb_stdoe_path = "{}/diamond_makedb_stdoe.txt".format(fpr_db_folder)
        process_nr_db(fpr_db_folder, diamond_db_bash_name, diamond_env, diamond_path, fpr_db_fasta, fpr_db_name, diamond_makedb_version_path, diamond_makedb_stdoe_path, thread_num, input_log_file, output_log_file, conda_sh_path)
    if enz_names_phylo:
        diamond_makedb_version_path = "{}/diamond_makedb_version.txt".format(fpr_db_folder_phylo)
        diamond_makedb_stdoe_path = "{}/diamond_makedb_stdoe.txt".format(fpr_db_folder_phylo)
        process_nr_db(fpr_db_folder_phylo, diamond_db_bash_name, diamond_env, diamond_path, fpr_db_fasta_phylo, fpr_db_name_phylo, diamond_makedb_version_path, diamond_makedb_stdoe_path, thread_num, input_log_file, output_log_file, conda_sh_path)


def collect_sra(sra_env, sra_folder, prefetch_size, sra_run_folder, sra_file, fastq_folder, fastq_single_end, sra_code, input_log_file, output_log_file, prefetch_path, vdb_validate_path, fastq_dump_path, sra_bash_script, conda_sh_path):
    print("\nSRA code selected: {}".format(sra_code))
    if not os.path.exists(sra_folder):
        os.mkdir(sra_folder)
    # Create a folder for the specific run.
    if os.path.exists(fastq_folder):
        print("\nThe folder with the FASTQ files of the specified run already exists. Analysis is continued based on the FASTQ files present in that folder ('{}').".format(fastq_folder))
    else:
        os.mkdir(sra_run_folder)
        prefetch_version_path = "{}/prefetch_version.txt".format(sra_run_folder)
        vdb_vaildate_version_path = "{}/vdb_vaildate_version.txt".format(sra_run_folder)
        fastq_dump_version_path = "{}/fastq_dump_version.txt".format(sra_run_folder)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(sra_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if sra_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(sra_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        if prefetch_path:
            phrase_1 = "\"{}\" -X {}G -O \"{}\" \"{}\"".format(prefetch_path, prefetch_size, sra_run_folder, sra_code)
            phrase_2 = "\"{}\" --version > \"{}\"".format(prefetch_path, prefetch_version_path)
        else:
            phrase_1 = "prefetch -X {}G -O \"{}\" \"{}\"".format(prefetch_size, sra_run_folder, sra_code)
            phrase_2 = "prefetch --version > \"{}\"".format(prefetch_version_path)
        if vdb_validate_path:
            phrase_3 = "\"{}\" \"{}/\"".format(vdb_validate_path, sra_file)
            phrase_4 = "\"{}\" --version > \"{}\"".format(vdb_validate_path, vdb_vaildate_version_path)
        else:
            phrase_3 = "vdb-validate \"{}/\"".format(sra_file)
            phrase_4 = "vdb-validate --version > \"{}\"".format(vdb_vaildate_version_path)
        if fastq_dump_path:
            phrase_5 = "\"{}\" --skip-technical --gzip --split-3 -O \"{}\" \"{}\"".format(fastq_dump_path, fastq_folder, sra_file)
            phrase_6 = "\"{}\" --version > \"{}\"".format(fastq_dump_path, fastq_dump_version_path)
        else:
            phrase_5 = "fastq-dump --skip-technical --gzip --split-3 -O \"{}\" \"{}\"".format(fastq_folder, sra_file)
            phrase_6 = "fastq-dump --version > {}".format(fastq_dump_version_path)
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        new_file_bash.write("{}\n".format(phrase_3))
        new_file_bash.write("{}\n".format(phrase_4))
        new_file_bash.write("{}\n".format(phrase_5))
        new_file_bash.write("{}\n".format(phrase_6))
        if sra_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(sra_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(sra_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Determine whether the files are SINGLE or PAIRED-END reads.
    fastq_files = os.listdir(fastq_folder)
    paired_end = None
    paired_end_size = 0
    single_end_size = 0
    # If there are 3 files in the folder, then 2 of them are paired-end and one single-end.
    if len(fastq_files) == 3:
        if os.path.exists(fastq_single_end):
            shutil.rmtree(fastq_single_end)
        os.mkdir(fastq_single_end)
        for ffi in fastq_files:
            ffi_path = "{}/{}".format(fastq_folder, ffi)
            if (ffi[-11:] != "_1.fastq.gz") and (ffi[-11:] != "_2.fastq.gz"):
                new_ffi_path = "{}/{}".format(fastq_single_end, ffi)
                shutil.move(ffi_path, new_ffi_path)
                pes = os.path.getsize(new_ffi_path)
                single_end_size += pes
            else:
                pes = os.path.getsize(ffi_path)
                paired_end_size += pes
        paired_end = True
        if single_end_size >= (0.5 * paired_end_size):
            print("\nThe size of the single-end file (forward or reverse reads whose pair was not found) is equal or greater than 50% of the size of the paired-end file. Exiting.")
            exit()
    elif len(fastq_files) == 2:
        paired_end = True
    elif len(fastq_files) == 1:
        paired_end = False
    else:
        print("\nError. Could not determine whether the files are single-end or paired-end.")
        exit()
    if paired_end:
        print("\nThe files are paired-end.")
    else:
        print("\nThe file is single-end.")
    return paired_end, fastq_folder


def fastqc(fastqc_env, output_path_fastqc, file_paths, fastqc_path, clear, thread_num, input_log_file, output_log_file, fastqc_bash_script, fastqc_version_path, fastqc_stdoe_path, conda_sh_path):
    print("\nRunning FastQC...")
    if clear:
        # The results folder and the folder for the temporary files generated by FastQC is created if FastQC is run for the first time in the pipeline.
        if os.path.exists(output_path_fastqc):
            shutil.rmtree(output_path_fastqc)
        os.mkdir(output_path_fastqc)
    # Create the Bash script.
    # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
    new_file_bash = open(fastqc_bash_script, "w")
    phrase = "#!/bin/bash"
    new_file_bash.write("{}\n".format(phrase))
    if fastqc_env:
        phrase_s = "source \"{}\"".format(conda_sh_path)
        phrase_a = "conda activate \"{}\"".format(fastqc_env)
        new_file_bash.write("{}\n".format(phrase_s))
        new_file_bash.write("{}\n".format(phrase_a))
    if fastqc_path:
        phrase_1 = "\"{}\" -t {} -o \"{}\"".format(fastqc_path, thread_num, output_path_fastqc)
        phrase_2 = "\"{}\" --version > \"{}\"".format(fastqc_path, fastqc_version_path)
    else:
        phrase_1 = "fastqc -t {} -o \"{}\"".format(thread_num, output_path_fastqc)
        phrase_2 = "fastqc --version > \"{}\"".format(fastqc_version_path)
    # Add each file to be analyzed at the end of the phrase.
    for i in file_paths:
        phrase_1 = "{} \"{}\"".format(phrase_1, i)
    phrase_1 = "{} &> \"{}\"".format(phrase_1, fastqc_stdoe_path)
    new_file_bash.write("{}\n".format(phrase_1))
    new_file_bash.write("{}\n".format(phrase_2))
    if fastqc_env:
        phrase = "conda deactivate"
        new_file_bash.write("{}\n".format(phrase))
    new_file_bash.close()
    # Making the Bash script executable.
    # Sending command to run.
    phrase_b1 = "chmod +x {}".format(fastqc_bash_script)
    phrase_b2 = "chmod --version"
    title_1 = "Making a Bash script executable:"
    title_2 = "Version of chmod:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Sending command to run.
    phrase_b1 = "bash {}".format(fastqc_bash_script)
    phrase_b2 = "bash --version"
    title_1 = "Running bash:"
    title_2 = "Version of bash:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)


def file_reads(output_path, paired_end, file_paths, file_paths_p, file_read_info_path):
    print("\nAnalyzing FastQC results...")
    # Number of total sequences from all input files
    total_sequences_num = 0
    paired_sequences_num = 0
    single_sequences_num = 0
    # The paired sequence refer to the sum of the sequences of only 1 of the 2 files of each pair of files.
    paired_to_seqnum_dict = {}
    single_to_sequences_dict = {}
    fastqc_files = os.listdir(output_path)
    for fqc_file in fastqc_files:
        if fqc_file[-5:] == ".html":
            # Counting the total number of sequences based on all input files.
            if paired_end:
                for key_p_temp in file_paths_p.keys():
                    file_path_p_temp_1 = file_paths_p[key_p_temp][0]
                    file_path_p_temp_2 = file_paths_p[key_p_temp][1]
                    # Matching the files from the fastqc folder to the ones provided.
                    # Catch the name after the last slash, if any exists.
                    if "/" in file_path_p_temp_1:
                        file_path_p_temp_1_splited = file_path_p_temp_1.split("/")
                        file_path_p_temp_1 = file_path_p_temp_1_splited[-1]
                    if "/" in file_path_p_temp_2:
                        file_path_p_temp_2_splited = file_path_p_temp_2.split("/")
                        file_path_p_temp_2 = file_path_p_temp_2_splited[-1]
                    # Remove the suffix of ".fastq" or ".fast.gz", if any exists.
                    if file_path_p_temp_1[-6:] == ".fastq":
                        file_path_p_temp_1 = file_path_p_temp_1[:-6]
                    elif file_path_p_temp_1[-9:] == ".fastq.gz":
                        file_path_p_temp_1 = file_path_p_temp_1[:-9]
                    if file_path_p_temp_2[-6:] == ".fastq":
                        file_path_p_temp_2 = file_path_p_temp_2[:-6]
                    elif file_path_p_temp_2[-9:] == ".fastq.gz":
                        file_path_p_temp_2 = file_path_p_temp_2[:-9]
                    # The path of the first parsed from the provided files.
                    fqc_file_base_name = fqc_file.split("_fastqc.html")[0]
                    if (fqc_file_base_name == file_path_p_temp_1) or (fqc_file_base_name == file_path_p_temp_2):
                        fqc_temp_file_path = "{}/{}".format(output_path, fqc_file)
                        fqc_lines_temp = read_file(fqc_temp_file_path)
                        for line_fqc_temp in fqc_lines_temp:
                            if "Total Sequences</td><td>" in line_fqc_temp:
                                splited_fqc_temp_1 = line_fqc_temp.split("Total Sequences</td><td>")
                                splited_fqc_temp_2 = splited_fqc_temp_1[1].split("</td></tr>")
                                sequences_num = int(splited_fqc_temp_2[0])
                                total_sequences_num += sequences_num
                                if key_p_temp not in paired_to_seqnum_dict.keys():
                                    paired_to_seqnum_dict[key_p_temp] = sequences_num
            else:
                for temp_i in range(0, len(file_paths)):
                    if temp_i not in single_to_sequences_dict.keys():
                        file_path_temp = file_paths[temp_i]
                        fqc_file_base_name = fqc_file.split("_fastqc.html")[0]
                        if (fqc_file_base_name in file_path_temp):
                            fqc_temp_file_path = "{}/{}".format(output_path, fqc_file)
                            fqc_lines_temp = read_file(fqc_temp_file_path)
                            for line_fqc_temp in fqc_lines_temp:
                                if "Total Sequences</td><td>" in line_fqc_temp:
                                    splited_fqc_temp_1 = line_fqc_temp.split("Total Sequences</td><td>")
                                    splited_fqc_temp_2 = splited_fqc_temp_1[1].split("</td></tr>")
                                    sequences_num = int(splited_fqc_temp_2[0])
                                    total_sequences_num += sequences_num
                                    single_to_sequences_dict[temp_i] = sequences_num
    if paired_end:
        paired_sequences_num = 0
        for key in paired_to_seqnum_dict.keys():
            paired_sequences_num += paired_to_seqnum_dict[key]
    else:
        single_sequences_num = total_sequences_num
    # Pritn information
    print("\nThe input number of reads is: {}".format(total_sequences_num))
    if paired_end:
        print("The input number of paired-end reads, of either direction (forward or reverse) is: {}".format(paired_sequences_num))
    else:
        print("The input number of single-end reads is: {}".format(single_sequences_num))
    # Write information
    if not os.path.exists(file_read_info_path):
        file_read_info_file = open(file_read_info_path, "w")
        file_read_info_file.write("Input reads\t{}\n".format(total_sequences_num))
        if paired_end:
            file_read_info_file.write("Input paired-end reads\t{}\n".format(paired_sequences_num))
        else:
            file_read_info_file.write("Input single-end reads\t{}\n".format(single_sequences_num))
        file_read_info_file.close()
    else:
        file_read_info_file = open(file_read_info_path, "a+")
        file_read_info_file.write("Trimmed reads\t{}\n".format(total_sequences_num))
        if paired_end:
            file_read_info_file.write("Trimmed paired-end reads\t{}\n".format(paired_sequences_num))
        else:
            file_read_info_file.write("Trimmed single-end reads\t{}\n".format(single_sequences_num))
        file_read_info_file.close()

    
def adapter_file(output_path_fastqc, adapters_status, adapters_path, adapters_ap_path):
    print("\nAnalyzing the file with the adapters...")
    # adapters_status
    # identify:The overrepresented sequences are added to the already existing adatpers to be searched.
    # fastqc: Only the overrepresented sequences are used as adatpers.
    # Multiple files can be given as inputs, so ultiple FastQC HMTL files may exist, thus multiple adapters can be found by each of these files. That's why the counter is initiazlied here.
    # When overrepresented sequences are not found then the sentence in the HTML file is a header.
    # When overrepresented sequences are found then these sequences are shown in a table.
    over_sequences = []
    if (adapters_status == "ide") or (adapters_status == "fas"):
        fastqc_files = os.listdir(output_path_fastqc)
        for fqc_file in fastqc_files:
            if fqc_file[-5:] == ".html":
                fqc_lines = read_file(output_path_fastqc + fqc_file)
                for fqc_line in fqc_lines:
                    if "Overrepresented sequences</h2><p>" in fqc_line:
                        splited_fq_1 = fqc_line.split("Overrepresented sequences</h2><p>")
                        splited_fq_2 = splited_fq_1[1].split("</p>")
                        over_info = splited_fq_2[0]
                        if over_info == "No overrepresented sequences":
                            print("No overrepresented sequences for the reads of the file: {}{}.\n".format(output_path_fastqc, fqc_file))
                    if "Overrepresented sequences</h2><table>" in fqc_line:
                        splited_fq_1 = fqc_line.split("Overrepresented sequences</h2><table>")
                        splited_fq_2 = splited_fq_1[1].split("<tr><td>")
                        for ov_seq_info in splited_fq_2[1:]:
                            over_seq_pre = ov_seq_info.split("</td><td>")[0]
                            over_sequences.append(over_seq_pre)
                            if "</table>" in ov_seq_info:
                                break
        # If no overrepresented sequences were found and "ide" was selected, then the adapters file will contain the already known adapters.
        # If no overrepresented sequences were found and "fas" was selected, then the adapters file will contain no sequences (empty file).
        if over_sequences:
            over_seq_counter = 1
            if adapters_status == "ide":
                shutil.copy(adapters_path, adapters_ap_path)
                adapters_ap_file = open(adapters_ap_path, "a")
                for over_seq in over_sequences:
                    adapters_ap_file.write(">Overrepresented_Sequence_{}\n".format(over_seq_counter))
                    adapters_ap_file.write("{}\n".format(over_seq))
                    over_seq_counter += 1
                adapters_ap_file.close()
            elif adapters_status == "fas":
                adapters_ap_file = open(adapters_ap_path, "w")
                for over_seq in over_sequences:
                    adapters_ap_file.write(">Overrepresented_Sequence_{}\n".format(over_seq_counter))
                    adapters_ap_file.write("{}\n".format(over_seq))
                    over_seq_counter += 1
                adapters_ap_file.close()
    elif adapters_status == "pre":
        shutil.copy(adapters_path, adapters_ap_path)
    else:
        shutil.copy(adapters_path, adapters_ap_path)


def bbduk(bbduk_env, output_path_trimmed, file_paths, adapters_ap_path, paired_end, file_paths_p, bbduk_path, thread_num, bbduk_max_ram, output_path_summaries_errors, input_log_file, output_log_file, bbduk_bash_script, bbduk_version_path, bbduk_stdoe_path, conda_sh_path):
    print("\nPreprocessing reads with BBDuk...")
    if os.path.exists(output_path_trimmed):
        shutil.rmtree(output_path_trimmed)
    os.mkdir(output_path_trimmed)
    # Run bbduk.
    stats_file_counter = 1
    phrase_1 = ""
    if paired_end:
        for key in file_paths_p.keys():
            seq_fastq_1 = file_paths_p[key][0]
            seq_fastq_2 = file_paths_p[key][1]
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(bbduk_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if bbduk_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(bbduk_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            if bbduk_path:
                phrase_1 = "\"{}\" \"in={}\" \"in2={}\" \"out={}/trimmed_results_{}\" \"out2={}/trimmed_results_{}\" interleaved=f \"stats={}/trimmed_stats_file_{}\" \"ref={}\" ktrim=r k=21 mink=11 hammingdistance=2 qtrim=rl trimq=20 minavgquality=20 minlength=25 tpe=t tbo=t threads={} -Xmx{}g &> \"{}\"".format(bbduk_path, seq_fastq_1, seq_fastq_2, output_path_trimmed, seq_fastq_1.split("/")[-1], output_path_trimmed, seq_fastq_2.split("/")[-1], output_path_summaries_errors, stats_file_counter, adapters_ap_path, thread_num, bbduk_max_ram, bbduk_stdoe_path)
                phrase_2 = "\"{}\" > {}".format(bbduk_path, bbduk_version_path)
            else:
                phrase_1 = "bbduk.sh \"in={}\" \"in2={}\" \"out={}/trimmed_results_{}\" \"out2={}/trimmed_results_{}\" interleaved=f \"stats={}/trimmed_stats_file_{}\" \"ref={}\" ktrim=r k=21 mink=11 hammingdistance=2 qtrim=rl trimq=20 minavgquality=20 minlength=25 tpe=t tbo=t threads={} -Xmx{}g &> \"{}\"".format(seq_fastq_1, seq_fastq_2, output_path_trimmed, seq_fastq_1.split("/")[-1], output_path_trimmed, seq_fastq_2.split("/")[-1], output_path_summaries_errors, stats_file_counter, adapters_ap_path, thread_num, bbduk_max_ram, bbduk_stdoe_path)
                phrase_2 = "bbduk.sh > \"{}\"".format(bbduk_version_path)
            stats_file_counter += 1
            new_file_bash.write("{}\n".format(phrase_1))
            new_file_bash.write("{}\n".format(phrase_2))
            if bbduk_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(bbduk_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(bbduk_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    else:
        for i in file_paths:
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(bbduk_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if bbduk_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(bbduk_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            if bbduk_path:
                phrase_1 = "\"{}\" \"in={}\" \"out={}/trimmed_results_{}\" \"interleaved=f stats={}/trimmed_stats_file_{}\" \"ref={}\" ktrim=r k=21 mink=11 hammingdistance=2 qtrim=rl trimq=20 minavgquality=20 minlength=25 tpe=t tbo=t threads={} -Xmx{}g &> \"{}\"".format(bbduk_path, i, output_path_trimmed, i.split("/")[-1], output_path_summaries_errors, stats_file_counter, adapters_ap_path, thread_num, bbduk_max_ram, bbduk_stdoe_path)
                phrase_2 = "\"{}\" --version > {}".format(bbduk_path, bbduk_version_path)
            else:
                phrase_1 = "bbduk.sh \"in={}\" \"out={}/trimmed_results_{}\" \"interleaved=f stats={}/trimmed_stats_file_{}\" \"ref={}\" ktrim=r k=21 mink=11 hammingdistance=2 qtrim=rl trimq=20 minavgquality=20 minlength=25 tpe=t tbo=t threads={} -Xmx{}g &> \"{}\"".format(i, output_path_trimmed, i.split("/")[-1], output_path_summaries_errors, stats_file_counter, adapters_ap_path, thread_num, bbduk_max_ram, bbduk_stdoe_path)
                phrase_2 = "bbduk.sh --version > {}".format(bbduk_version_path)
            stats_file_counter += 1
            new_file_bash.write("{}\n".format(phrase_1))
            new_file_bash.write("{}\n".format(phrase_2))
            if bbduk_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(bbduk_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(bbduk_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)


def reduce_volume(clear_space, input_folder, output_path_trimmed, input_log_file, del_pick):
    if clear_space:
        input_log_file.write("Deleting files:\n")
        if 1 in del_pick:
            input_file_paths = os.listdir(input_folder)
            for i in input_file_paths:
                input_file_rel_path = "{}/{}".format(input_folder, i)
                print("Deleting file: {}".format(input_file_rel_path))
                os.remove(input_file_rel_path)
                input_log_file.write("File deleted: {}\n".format(input_file_rel_path))
        if 2 in del_pick:
            # The file paths in "ca_file_paths" contain the file paths from the "trimmed_results" folder which are compressed, which are the files that
            # contain the "fastq" name in their names. For being sure we also check that the suffix is ".gz" before deleting the compressed files.
            trimmed_file_paths = os.listdir(output_path_trimmed)
            for i in trimmed_file_paths:
                if "fastq" in i and i[-3:] == ".gz":
                    local_file_path = "{}/{}".format(output_path_trimmed, i)
                    print("Deleting file: {}".format(local_file_path))
                    os.remove(local_file_path)
                    input_log_file.write("File deleted: {}\n".format(local_file_path))
        input_log_file.write("\n")


def unzip_files(file_paths, unzip_path, clear, input_log_file, output_log_file, gzip_path):
    if unzip_path:
        if clear:
            if os.path.exists(unzip_path):
                shutil.rmtree(unzip_path)
        os.mkdir(unzip_path)
    title_1 = "Uncompressing files - gzip:"
    title_2 = "Version of gzip:"
    capture_status = True
    shell_status = True
    pr_status = False
    for i in file_paths:
        file_name = i.split("/")[-1]
        file_name_parts = file_name.split(".")
        file_name = ".".join(file_name_parts[:-1])
        if unzip_path is not None:
            if gzip_path:
                phrase_1 = "\"{}\" -dk -c \"{}\"".format(gzip_path, i)
                phrase_1 = "{} > \"{}{}\"".format(phrase_1, unzip_path, file_name)
                phrase_2 = "{} --version".format(gzip_path)
            else:
                phrase_1 = "gzip -dk -c \"{}\"".format(i)
                phrase_1 = "{} > \"{}{}\"".format(phrase_1, unzip_path, file_name)
                phrase_2 = "gzip --version"
        else:
            if gzip_path:
                phrase_1 = "\"{}\" -dk \"{}\"".format(gzip_path, i)
                phrase_2 = "\"{}\" --version".format(gzip_path)
            else:
                phrase_1 = "gzip -dk \"{}\"".format(i)
                phrase_2 = "gzip --version"
        command_run(phrase_1, phrase_2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)


def megahit(megahit_env, output_path_megahit, tr_ex_file_paths, paired_end, tr_ex_file_paths_p, megahit_path, k_list, thread_num, input_log_file, output_log_file, megahit_bash_script, megahit_version_path, megahit_stdoe_path, conda_sh_path):
    print("\nRunning Megahit...")
    # Megahit (with the option "-o" as used in this case) demands that the output folder does not already exists. In this case the output folder its not "output_path_megahit", its "output_path_megahit/Megahit_Contigs/".
    if os.path.exists(output_path_megahit):
        shutil.rmtree(output_path_megahit)
    os.mkdir(output_path_megahit)
    # --k-list 15,17,21,29,39,59,79,99,119,141 --merge-level=20,0.7
    # --k-list 15,17,21,29,39,59,79,99,119,141
    # --presets meta-large: --min-count 1 --k-list 21,29,39,49,...,129,141
    # --presets meta-sensitive: --k-min 27 --k-max 127 --k-step 10
    if paired_end:
        if megahit_path:
            phrase_1 = "\"{}\" -t {}".format(megahit_path, thread_num)
            phrase_2 = "\"{}\" --version > \"{}\"".format(megahit_path, megahit_version_path)
        else:
            phrase_1 = "megahit -t {}".format(thread_num)
            phrase_2 = "megahit --version > \"{}\"".format(megahit_version_path)
        if k_list is not None:
            phrase_1 = "{} \"--k-list={}\"".format(phrase_1, k_list)
        phrase_1 = "{} -1 \"".format(phrase_1)
        tr_ex_file_paths_keys = list(tr_ex_file_paths_p.keys())
        for key_i in range(0, len(tr_ex_file_paths_keys)):
            key = tr_ex_file_paths_keys[key_i]
            if key_i + 1 == len(tr_ex_file_paths_keys):
                phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths_p[key][0])
            else:
                phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths_p[key][0])
        phrase_1 = "{} -2 \"".format(phrase_1)
        for key_i in range(0, len(tr_ex_file_paths_p.keys())):
            key = tr_ex_file_paths_keys[key_i]
            if key_i + 1 == len(tr_ex_file_paths_keys):
                phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths_p[key][1])
            else:
                phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths_p[key][1])
        phrase_1 = "{} --presets meta-large -o {}/megahit_contigs &> {}".format(phrase_1, output_path_megahit, megahit_stdoe_path)
    else:
        if megahit_path:
            phrase_1 = "\"{}\" -t {}".format(megahit_path, thread_num)
            phrase_2 = "\"{}\" --version > \"{}\"".format(megahit_path, megahit_version_path)
        else:
            phrase_1 = "megahit -t {}".format(thread_num)
            phrase_2 = "megahit --version > \"{}\"".format(megahit_version_path)
        if k_list is not None:
            phrase_1 = "{} \"--k-list={}\"".format(phrase_1, k_list)
        phrase_1 = "{} -r \"".format(phrase_1)
        for i in range(0, len(tr_ex_file_paths)):
            if i + 1 == len(tr_ex_file_paths):
                phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths[i])
            else:
                phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths[i])
        phrase_1 = "{} --presets meta-large -o \"{}/megahit_contigs\" &> \"{}\"".format(phrase_1, output_path_megahit, megahit_stdoe_path)
    # Create the Bash script.
    # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
    new_file_bash = open(megahit_bash_script, "w")
    phrase = "#!/bin/bash"
    new_file_bash.write("{}\n".format(phrase))
    if megahit_env:
        phrase_s = "source \"{}\"".format(conda_sh_path)
        phrase_a = "conda activate \"{}\"".format(megahit_env)
        new_file_bash.write("{}\n".format(phrase_s))
        new_file_bash.write("{}\n".format(phrase_a))
    new_file_bash.write("{}\n".format(phrase_1))
    new_file_bash.write("{}\n".format(phrase_2))
    if megahit_env:
        phrase = "conda deactivate"
        new_file_bash.write("{}\n".format(phrase))
    new_file_bash.close()
    # Making the Bash script executable.
    # Sending command to run.
    phrase_b1 = "chmod +x {}".format(megahit_bash_script)
    phrase_b2 = "chmod --version"
    title_1 = "Making a Bash script executable:"
    title_2 = "Version of chmod:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Sending command to run.
    phrase_b1 = "bash {}".format(megahit_bash_script)
    phrase_b2 = "bash --version"
    title_1 = "Running bash:"
    title_2 = "Version of bash:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)


def kraken_filtering(kraken_threshold, kraken_species_absab_dict, total_classified_reads_num, kraken_report_path, kraken_species_thr_path, kraken_filters_path):
    # Write the kraken thresholds applied in a file.
    kraken_filters_file = open(kraken_filters_path, "w")
    first_kt = True
    kraken_species_thr_dict = None
    for kt in kraken_threshold:
        kraken_species_thr_dict_temp = {}
        if "." in kt:
            kt = float(kt)
        else:
            kt = int(kt)
        kraken_prop = None
        kraken_abs = None
        shannon_index = None
        kraken_filters_file.write("Filter: {}\n".format(kt))
        # Compute the threshold, if needed.
        if kt in [-1, -2]:
            term_sum = 0
            for key in kraken_species_absab_dict.keys():
                if total_classified_reads_num != 0:
                    temp_absab = kraken_species_absab_dict[key]
                    temp_prop = temp_absab / total_classified_reads_num
                    if temp_prop != 0:
                        temp_term = temp_prop * math.log(temp_prop)
                        term_sum += temp_term
            shannon_index = -term_sum
            if kt == -1:
                if 0 <= shannon_index <= 2.5:
                    kraken_prop = 1
                elif 2.5 < shannon_index <= 4.5:
                    kraken_prop = 0.1
                elif 4.5 < shannon_index:
                    kraken_prop = 0
            elif kt == -2:
                if 0 <= shannon_index <= 2.5:
                    kraken_prop = 0.1
                elif 2.5 < shannon_index:
                    kraken_prop = 0
            if kraken_prop is not None:
                # Conerting kraken_prop to a percentage.
                kraken_prop = float(kraken_prop)
                kraken_abs = total_classified_reads_num * kraken_prop
            # Create a list with species above the threshold.
            if os.path.exists(kraken_report_path):
                with open(kraken_report_path) as report_lines:
                    for line in report_lines:
                        line = line.rstrip("\n")
                        line_splited = line.split("\t")
                        clade_type = line_splited[3]
                        if clade_type == "S":
                            abu_perc = float(line_splited[0])
                            taxon_count = int(line_splited[1])
                            taxonid = int(line_splited[4])
                            if abu_perc >= kraken_prop:
                                kraken_species_thr_dict_temp[taxonid] = [taxon_count, abu_perc]
            kraken_filters_file.write("Shannon Index: {}\n".format(shannon_index))
            kraken_filters_file.write("Kraken percentage filter: {}\n".format(kraken_prop))
            kraken_filters_file.write("Kraken absolute filter: {}\n".format(kraken_abs))
        elif isinstance(kt, int):
            if os.path.exists(kraken_report_path):
                with open(kraken_report_path) as report_lines:
                    for line in report_lines:
                        line = line.rstrip("\n")
                        line_splited = line.split("\t")
                        clade_type = line_splited[3]
                        if clade_type == "S":
                            abu_perc = float(line_splited[0])
                            taxon_count = int(line_splited[1])
                            taxonid = int(line_splited[4])
                            if taxon_count >= kt:
                                kraken_species_thr_dict_temp[taxonid] = [taxon_count, abu_perc]
        elif isinstance(kt, float):
            if os.path.exists(kraken_report_path):
                with open(kraken_report_path) as report_lines:
                    for line in report_lines:
                        line = line.rstrip("\n")
                        line_splited = line.split("\t")
                        clade_type = line_splited[3]
                        if clade_type == "S":
                            abu_perc = float(line_splited[0])
                            taxon_count = int(line_splited[1])
                            taxonid = int(line_splited[4])
                            if abu_perc >= kt:
                                kraken_species_thr_dict_temp[taxonid] = [taxon_count, abu_perc]
        else:
            print("\nAn error occured when filtering the species in the kraken report. Exiting.")
            exit()
        if kraken_prop is None:
            kraken_prop = "None"
            kraken_prop_temp = "None"
        else:
            kraken_prop_temp = round(kraken_prop, 5)
        if kraken_abs is None:
            kraken_abs = "None"
            kraken_abs_temp = "None"
        else:
            kraken_abs_temp = round(kraken_abs, 5)
        abs_sum = 0
        perc_sum = 0
        kraken_species_thr_path_temp = "{}_{}_{}_{}.tsv".format(kraken_species_thr_path, kt, kraken_prop_temp, kraken_abs_temp)
        new_file_species_thr = open(kraken_species_thr_path_temp, "w")
        for key in kraken_species_thr_dict_temp.keys():
            taxon_count = kraken_species_thr_dict_temp[key][0]
            abu_perc = kraken_species_thr_dict_temp[key][1]
            abs_sum += taxon_count
            perc_sum += abu_perc
            new_file_species_thr.write("{}\t{}\t{}\n".format(key, taxon_count, abu_perc))
        new_file_species_thr.write("Total\t{}\t{}\n".format(abs_sum, perc_sum))
        new_file_species_thr.close()
        # The kraken threshold used to create the bins based on the filtered species from kraken is the first threshold used in the kraken threshold list.
        if first_kt:
            kraken_species_thr_dict = copy.deepcopy(kraken_species_thr_dict_temp)
            first_kt = False
    kraken_filters_file.close()
    return kraken_species_thr_dict


def kraken(paired_end, tr_ex_file_paths_p, tr_ex_file_paths, output_path_kraken, kraken_db_path, kraken_results_path, kraken_report_path, kraken_threshold, kraken_species_path, kraken_species_thr_path, kraken_reads_path, kraken_taxname_path, conda_sh_path, kraken_env, kraken_path, kraken_bash_script, kraken_stde_path, kraken_version_path, kraken_status, kraken_memory_mapping, thread_num, time_dict, kraken_filters_path, input_log_file, output_log_file):
    print("\nRunning Kraken2...")
    read_to_species_dict = {}
    taxid_to_species_dict = {}
    if kraken_memory_mapping:
        kraken_memory_mapping = "--memory-mapping "
    else:
        kraken_memory_mapping = ""
    start_time_kraken_spec = time.time()
    if kraken_status:
        if os.path.exists(output_path_kraken):
            shutil.rmtree(output_path_kraken)
        os.mkdir(output_path_kraken)
        # Create the Bash script.
        new_file_bash = open(kraken_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if kraken_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(kraken_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        if kraken_path:
            phrase_1 = "\"{}\" --db \"{}\" --output \"{}\" --threads {} --report \"{}\" {}--paired --use-names".format(kraken_path, kraken_db_path, kraken_results_path, thread_num, kraken_report_path, kraken_memory_mapping)
            phrase_2 = "\"{}\" --version > \"{}\"".format(kraken_path, kraken_version_path)
        else:
            phrase_1 = "kraken2 --db \"{}\" --output \"{}\" --threads {} --report \"{}\" {}--paired --use-names".format(kraken_db_path, kraken_results_path, thread_num, kraken_report_path, kraken_memory_mapping)
            phrase_2 = "kraken2 --version > \"{}\"".format(kraken_version_path)
        if paired_end:
            phrase_1 = "{} \"".format(phrase_1)
            tr_ex_file_paths_keys = list(tr_ex_file_paths_p.keys())
            for key_i in range(0, len(tr_ex_file_paths_keys)):
                key = tr_ex_file_paths_keys[key_i]
                if key_i + 1 == len(tr_ex_file_paths_keys):
                    phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths_p[key][0])
                else:
                    phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths_p[key][0])
            phrase_1 = "{} \"".format(phrase_1)
            for key_i in range(0, len(tr_ex_file_paths_p.keys())):
                key = tr_ex_file_paths_keys[key_i]
                if key_i + 1 == len(tr_ex_file_paths_keys):
                    phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths_p[key][1])
                else:
                    phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths_p[key][1])
        else:
            phrase_1 = "{} \"".format(phrase_1)
            for i in range(0, len(tr_ex_file_paths)):
                if i + 1 == len(tr_ex_file_paths):
                    phrase_1 = "{}{}\"".format(phrase_1, tr_ex_file_paths[i])
                else:
                    phrase_1 = "{}{},".format(phrase_1, tr_ex_file_paths[i])
        phrase_1 = "{} 2> \"{}\"".format(phrase_1, kraken_stde_path)
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if kraken_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(kraken_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(kraken_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    label = "Process of kraken2 process (specific):"
    elpased_time = end_time_analysis(label, start_time_kraken_spec, output_log_file)
    time_dict["kraken_specific_time"] = elpased_time
    # Store the kraken species and their percs abundancies.
    kraken_species_dict = {}
    if os.path.exists(kraken_report_path):
        with open(kraken_report_path) as report_lines:
            for line in report_lines:
                line = line.rstrip("\n")
                line_splited = line.split("\t")
                clade_type = line_splited[3]
                if clade_type == "S":
                    abu_perc = float(line_splited[0])
                    taxon_count = int(line_splited[1])
                    taxonid = int(line_splited[4])
                    kraken_species_dict[taxonid] = [taxon_count, abu_perc]
    # Store the kraken species and their absolute abundancies.
    kraken_species_absab_dict = {}
    total_classified_reads_num = 0
    if os.path.exists(kraken_report_path):
        with open(kraken_report_path) as report_lines:
            for line in report_lines:
                line = line.rstrip("\n")
                line_splited = line.split("\t")
                clade_type = line_splited[3]
                taxon_name = line_splited[5]
                taxon_name = taxon_name.strip()
                # The total number of classified reads belong to the root taxonomy rank. The rest of the reads belong to the
                # unclassified reads.
                if taxon_name == "root":
                    taxon_count = int(line_splited[1])
                    total_classified_reads_num = taxon_count
                if clade_type == "S":
                    taxon_count = int(line_splited[1])
                    taxonid = int(line_splited[4])
                    taxon_name = line_splited[5]
                    kraken_species_absab_dict[taxonid] = taxon_count
    # Kraken filtering thresholds are applied.s
    kraken_species_thr_dict = kraken_filtering(kraken_threshold, kraken_species_absab_dict, total_classified_reads_num, kraken_report_path, kraken_species_thr_path, kraken_filters_path)
    # Check for the results
    if os.path.exists(kraken_results_path):
        with open(kraken_results_path) as results_lines:
            for line in results_lines:
                line = line.rstrip("\n")
                line_splited = line.split("\t")
                result_type = line_splited[0]
                if result_type == "C":
                    read_id = line_splited[1]
                    taxonid_phrase = line_splited[2]
                    if "taxid " in taxonid_phrase:
                        taxonid_phrase_splited = taxonid_phrase.split("taxid ")
                        taxonid_phrase_part = taxonid_phrase_splited[1]
                        if taxonid_phrase_part[-1] == ")":
                            taxonid_phrase_part = taxonid_phrase_part[:-1]
                            taxonid = int(taxonid_phrase_part)
                            # The read_to_species_dict is based on the dictionary (kraken_species_thr_dict) returned by the
                            # filtering which is the one generated by the first filter applied to the species of the kraken report.
                            # The binning based on the taxonomy assignment of kraken is based on the read_to_species_dict dictionary.
                            read_to_species_dict[read_id] = [taxonid, 0]
                            if taxonid in kraken_species_thr_dict.keys():
                                read_to_species_dict[read_id][1] = 1
                            # The taxid_to_species_dict is not depended on the kraken_species_thr_dict dictionary.
                            if taxonid not in taxid_to_species_dict.keys():
                                taxname = taxonid_phrase_splited[0]
                                if taxname[-1] == "(":
                                    taxname = taxname[:-1]
                                taxname = taxname.strip()
                                taxid_to_species_dict[taxonid] = taxname
    new_file_species = open(kraken_species_path, "w")
    for key in kraken_species_dict.keys():
        taxon_count = kraken_species_dict[key][0]
        abu_perc = kraken_species_dict[key][1]
        new_file_species.write("{}\t{}\t{}\n".format(key, taxon_count, abu_perc))
    new_file_species.close()
    new_file_reads = open(kraken_reads_path, "w")
    for key in read_to_species_dict.keys():
        taxonid = read_to_species_dict[key][0]
        thr_state = read_to_species_dict[key][1]
        new_file_reads.write("{}\t{}\t{}\n".format(key, taxonid, thr_state))
    new_file_reads.close()
    new_file_taxname = open(kraken_taxname_path, "w")
    for key_tax in taxid_to_species_dict.keys():
        tax_name = taxid_to_species_dict[key_tax]
        new_file_taxname.write("{}\t{}\n".format(key_tax, tax_name))
    new_file_taxname.close()
    return kraken_species_dict, kraken_species_thr_dict, read_to_species_dict, taxid_to_species_dict, time_dict
                

def binning(binning_tool, metabinner_env, comebin_env, contigs, protein_input, output_path_bin, metabinner_bin_path, comebin_bin_path, fullpath_contigs_formated, bin_bam_folder, enz_dir, output_path_conitgs, full_coverage_folder, fullpath_tr_fastq_files, full_bin_results_folder, full_coverage_profile_path, thread_num, input_log_file, output_log_file, metabinner_bash_script, comebin_bash_script, bin_ram_ammount, bin_num_contig_len, bin_num_kmer, bin_gen_coverage_stdoe_path, bin_gen_kmer_stdoe_path, bin_filter_tooshort_stdoe_path, binning_stdoe_path, comebin_batch_size, conda_sh_path):
    print("\nPerforming binning with MetaBinner or COMEBin...")
    # Delete other files from the contigs directory, related to previous binning processes.
    if os.path.exists(output_path_conitgs):
        fn_group = os.listdir(output_path_conitgs)
        for fn in fn_group:
            if fn not in ["contigs.fa", "contigs_formated.fna"]:
                fn_path = "{}/{}".format(output_path_conitgs, fn)
                os.remove(fn_path)
    if contigs or protein_input:
        return 0
    if os.path.exists(output_path_bin):
        shutil.rmtree(output_path_bin)
    os.mkdir(output_path_bin)
    if os.path.exists(full_bin_results_folder):
        shutil.rmtree(full_bin_results_folder)
    if binning_tool == 1:
        # Change directory to the scripts of the metabinner enviroment.
        gen_coverage_file_path = "{}/scripts/gen_coverage_file.sh".format(metabinner_bin_path)
        gen_kmer_path = "{}/scripts/gen_kmer.py".format(metabinner_bin_path)
        filter_tooshort_path = "{}/scripts/Filter_tooshort.py".format(metabinner_bin_path)
        relpath_filtered_contigs_path = "{}/contigs_formated_kmer_4_f{}.csv".format(output_path_conitgs, bin_num_contig_len)
        full_filtered_contigs_path = os.path.abspath(relpath_filtered_contigs_path)
        run_metabinner_path = "{}/run_metabinner.sh".format(metabinner_bin_path)
        contnigs_formated_splited = fullpath_contigs_formated.split(".")
        contnigs_formated_splited = contnigs_formated_splited[:-1]
        contnigs_filtered_path = ".".join(contnigs_formated_splited)
        contnigs_filtered_path = "{}_{}.fa".format(contnigs_filtered_path, bin_num_contig_len)
        # Find the first line of the script for filtering contigs.
        # This indicates whether metabinner was installed from source from github or through anaconda.
        # If place_python is True then metabinner was installed from github, python is needed in front of the command to use the python of the enviroment but the 
        # full_coverage_profile_path path should be based on the selected contig length. If the place_python is False then metabinner was installed through anaconda,
        # no python is needed in front the command but the full_coverage_profile_path path should not change.
        place_python = False
        fts_lines = read_file(filter_tooshort_path)
        first_line = fts_lines[0]
        if first_line == "#!/usr/bin/python":
            place_python = True
        else:
            full_coverage_profile_path = "{}/coverage_profile_f1k.tsv".format(full_coverage_folder)
        # Get the coverage profiles.
        phrase_1 = "\"{}\" -a \"{}\" -o \"{}\" -b \"{}\" -t {} -m {} -l {} {} &> \"{}\"".format(gen_coverage_file_path, fullpath_contigs_formated, full_coverage_folder, bin_bam_folder, thread_num, bin_ram_ammount, bin_num_contig_len, fullpath_tr_fastq_files, bin_gen_coverage_stdoe_path)
        # Create kmers.
        phrase_2 = "\"{}\" \"{}\" {} {} &> \"{}\"".format(gen_kmer_path, fullpath_contigs_formated, bin_num_contig_len, bin_num_kmer, bin_gen_kmer_stdoe_path)
        # Filter contigs based on their lengths.
        phrase_3 = "\"{}\" \"{}\" {} &> \"{}\"".format(filter_tooshort_path, fullpath_contigs_formated, bin_num_contig_len, bin_filter_tooshort_stdoe_path)
        if place_python:
            phrase_3 = "python {}".format(phrase_3)
        # Bin the contigs.
        phrase_4 = "\"{}\" -a \"{}\" -o \"{}\" -d \"{}\" -k \"{}\" -p \"{}\" -t {} &> \"{}\"".format(run_metabinner_path, contnigs_filtered_path, full_bin_results_folder, full_coverage_profile_path, full_filtered_contigs_path, metabinner_bin_path, thread_num, binning_stdoe_path)
        # Create the Bash script.
        new_file_bash = open(metabinner_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if metabinner_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(metabinner_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        new_file_bash.write("{}\n".format(phrase_3))
        new_file_bash.write("{}\n".format(phrase_4))
        if metabinner_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(metabinner_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(metabinner_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    else:
        gen_coverage_file_path = "{}/scripts/gen_cov_file.sh".format(comebin_bin_path)
        filter_tooshort_path = "{}/scripts/Filter_tooshort.py".format(comebin_bin_path)
        run_comebin_path = "{}/run_comebin.sh".format(comebin_bin_path)
        contnigs_formated_splited = fullpath_contigs_formated.split(".")
        contnigs_formated_splited = contnigs_formated_splited[:-1]
        contnigs_filtered_path = ".".join(contnigs_formated_splited)
        contnigs_filtered_path = "{}_{}.fa".format(contnigs_filtered_path, bin_num_contig_len)
        # Change path to the comebin folder.
        phrase_1 = "cd \"{}\"".format(comebin_bin_path)
        # Get the coverage profiles.
        phrase_2 = "\"{}\" -a \"{}\" -o \"{}\" -b \"{}\" -t {} -m {} -l {} {} &> \"{}\"".format(gen_coverage_file_path, fullpath_contigs_formated, full_coverage_folder, bin_bam_folder, thread_num, bin_ram_ammount, bin_num_contig_len, fullpath_tr_fastq_files, bin_gen_coverage_stdoe_path)
        # Filter contigs based on their lengths.
        phrase_3 = "\"{}\" \"{}\" {} &> \"{}\"".format(filter_tooshort_path, fullpath_contigs_formated, bin_num_contig_len, bin_filter_tooshort_stdoe_path)
        # Bin the contigs.
        phrase_4 = "\"{}\" -a \"{}\" -o \"{}\" -p \"{}\" -t {} -b {} &> \"{}\"".format(run_comebin_path, contnigs_filtered_path, full_bin_results_folder, bin_bam_folder, thread_num, comebin_batch_size, binning_stdoe_path)
        # Change path to the proteoseeker folder
        phrase_5 = "cd \"{}\"".format(enz_dir)
        # Create the Bash script.
        new_file_bash = open(comebin_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if comebin_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(comebin_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        new_file_bash.write("{}\n".format(phrase_3))
        new_file_bash.write("{}\n".format(phrase_4))
        new_file_bash.write("{}\n".format(phrase_5))
        if comebin_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(comebin_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(comebin_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)


def contig_formation(contigs, file_paths, output_final_contigs, output_megahit_contigs, output_final_contigs_formated, input_log_file, output_log_file, cat_path):
    print("\nLocating the file with the contigs...")
    # Copy contigs to contigs folder
    # If the input files contain files with contigs then their combination is the final contigs file.
    # If the input files contain only FASTQ files then the file with the contigs is produced by Megahit and it is found in the corresponding results folder (of Megahit).
    if contigs:
        if cat_path:
            phrase_1 = "\"{}\"".format(cat_path)
            phrase_2 = "\"{}\" --version".format(cat_path)
        else:
            phrase_1 = "cat"
            phrase_2 = "cat --version"
        # Making sure only Fasta files are copied to the folder with the contigs and not FastQ files.
        # We consider, based on the information of usage given for the program, that Fasta files include only contigs.
        for i in file_paths:
            if i[-5:] != "fastq":
                phrase_1 = "{} \"{}\"".format(phrase_1, i)
        phrase_1 = "\n{} > \"{}\"".format(phrase_1, output_final_contigs)
        capture_status = True
        shell_status = True
        pr_status = False
        title_1 = "Concatenate files - cat:"
        title_2 = "Version of cat:"
        command_run(phrase_1, phrase_2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    else:
        if os.path.exists(output_megahit_contigs):
            shutil.copy(output_megahit_contigs, output_final_contigs)
        else:
            print("\nNo files with contigs were generated from the assembly. Exiting.")
            exit()
    # Writting the final contigs in FASTA format.
    split_size = 70
    new_file_final_contigs_formated = open(output_final_contigs_formated, "w")
    with open(output_final_contigs, "r") as final_contigs_lines:
        for c_line in final_contigs_lines:
            c_line = c_line.rstrip("\n")
            if c_line:
                if c_line[0] == ">":
                    new_file_final_contigs_formated.write("{}\n".format(c_line))
                else:
                    c_fasta_splited = [c_line[ci:ci + split_size] for ci in range(0, len(c_line), split_size)]
                    for c_fasta_line in c_fasta_splited:
                        new_file_final_contigs_formated.write("{}\n".format(c_fasta_line))
    new_file_final_contigs_formated.close()
    

def gene_prediction(genepred_env, output_path_genepred, output_path_genepred_base, output_path_genepred_faa, output_fgs_protein_formatted_1, fullpath_contigs_formated, fraggenescanrs_path, thread_num, input_log_file, output_log_file, gene_bash_script, fraggenescanrs_version_path, fraggenescanrs_stdoe_path, conda_sh_path):
    print("\nPerofrming gene prediction...")
    if os.path.exists(output_path_genepred):
        shutil.rmtree(output_path_genepred)
    os.mkdir(output_path_genepred)
    # In the output of FragGeneScanRs in the headers, the first two last two numbers, which are divided by an underscore, consist the starting and ending points of the gene on the contig which it was found to be located in.
    # E.g. Header: >k141_180_2_133_- : This means that the name of the contig is "k141_180", the starting nucleotide on the contig is 2 and ending nucleotide on the contig is 133.
    # FragGeneScanRs will either be run to the contigs or the reads after they have been filtered by the trimming tool.
    input_log_file.write("Gene prediction:\n")
    capture_status = True
    shell_status = True
    # Run FragGeneScanRs
    if fraggenescanrs_path:
        phrase_1 = "\"{}\" -w 1 -o \"{}\" -s \"{}\" -p {} --training-file complete &> \"{}\"".format(fraggenescanrs_path, output_path_genepred_base, fullpath_contigs_formated, thread_num, fraggenescanrs_stdoe_path)
        phrase_2 = "\"{}\" --version > \"{}\"".format(fraggenescanrs_path, fraggenescanrs_version_path)
    else:
        phrase_1 = "FragGeneScanRs -w 1 -o \"{}\" -s \"{}\" -p {} --training-file complete &> \"{}\"".format(output_path_genepred_base, fullpath_contigs_formated, thread_num, fraggenescanrs_stdoe_path)
        phrase_2 = "FragGeneScanRs --version > \"{}\"".format(fraggenescanrs_version_path)
    # Create the Bash script.
    # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
    new_file_bash = open(gene_bash_script, "w")
    phrase = "#!/bin/bash"
    new_file_bash.write("{}\n".format(phrase))
    if genepred_env:
        phrase_s = "source \"{}\"".format(conda_sh_path)
        phrase_a = "conda activate \"{}\"".format(genepred_env)
        new_file_bash.write("{}\n".format(phrase_s))
        new_file_bash.write("{}\n".format(phrase_a))
    new_file_bash.write("{}\n".format(phrase_1))
    new_file_bash.write("{}\n".format(phrase_2))
    if genepred_env:
        phrase = "conda deactivate"
        new_file_bash.write("{}\n".format(phrase))
    new_file_bash.close()
    # Making the Bash script executable.
    # Sending command to run.
    phrase_b1 = "chmod +x {}".format(gene_bash_script)
    phrase_b2 = "chmod --version"
    title_1 = "Making a Bash script executable:"
    title_2 = "Version of chmod:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Sending command to run.
    phrase_b1 = "bash {}".format(gene_bash_script)
    phrase_b2 = "bash --version"
    title_1 = "Running bash:"
    title_2 = "Version of bash:"
    capture_status = True
    shell_status = True
    pr_status = False
    command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Check the output file.
    if os.path.getsize(output_path_genepred_faa) != 0:
        split_size = 70
        new_file_fgs_formatted = open(output_fgs_protein_formatted_1, "w")
        with open("{}".format(output_path_genepred_faa)) as fgs_output_lines:
            for line in fgs_output_lines:
                line = line.rstrip("\n")
                if line[0] == ">":
                    new_file_fgs_formatted.write("{}\n".format(line))
                else:
                    fasta_splited = [line[ssi:ssi + split_size] for ssi in range(0, len(line), split_size)]
                    for fasta_line in fasta_splited:
                        new_file_fgs_formatted.write("{}\n".format(fasta_line))
        new_file_fgs_formatted.close()


def gene_annotation(output_path_genepred, output_path_genepred_ffn, output_fgs_protein_formatted_2, output_final_contigs_formated, gene_contig_dist_path, gene_info_file, genetic_code):
    print("\nExtracting information from the predicted gene sequences...")
    # If there is a file with the gene sequence from FragGeneScanRs then:
    # For each protein ID of the putative proteins parse the file and find its gene sequence.
    # Then check for starting (including those of bacterial type) and end codons in the sequence and also check the
    # sequence for the number of rare codos and ideally run it across a program/software that scores the gene sequence
    # based on its expression capability.
    gene_sequences_dict = {}
    if os.path.exists(output_path_genepred) and os.path.exists(output_path_genepred_ffn):
        gene_lines = read_file(output_path_genepred_ffn)
        for i in range(0, len(gene_lines)):
            line = gene_lines[i]
            if line[0] == ">":
                protein_id = line[1:]
                gene_sequence = gene_lines[i+1]
                gene_sequences_dict[protein_id] = [gene_sequence]
    # Check for start and end codons for any gene sequence found.
    # For the next item after the gene sequence, if a start codon is present then 1 is placed, otherwise 0.
    # Similarly for the end codon.
    dict_genes = {}
    if gene_sequences_dict:
        for key in gene_sequences_dict.keys():
            if gene_sequences_dict[key][0][0:3] == "ATG" or gene_sequences_dict[key][0][0:3] == "GTG" or gene_sequences_dict[key][0][0:3] == "GTC":
                gene_sequences_dict[key].append("YES")
            else:
                gene_sequences_dict[key].append("NO")
            if gene_sequences_dict[key][0][-3:] == "TAA" or gene_sequences_dict[key][0][0:3] == "TAG" or gene_sequences_dict[key][0][0:3] == "TGA":
                gene_sequences_dict[key].append("YES")
            else:
                gene_sequences_dict[key].append("NO")
        new_file = open(gene_info_file, "w")
        for key in gene_sequences_dict.keys():
            new_file.write("{}\t{}\t{}\t{}\n".format(key, gene_sequences_dict[key][0], gene_sequences_dict[key][1], gene_sequences_dict[key][2]))
            dict_genes[key] = [gene_sequences_dict[key][0], gene_sequences_dict[key][1], gene_sequences_dict[key][2]]
        new_file.close()
    # Translate gene sequences
    pr_dict = {}
    if gene_sequences_dict:
        for key in gene_sequences_dict.keys():
            gene_seq = gene_sequences_dict[key][0]
            put_protein = translation(gene_seq, genetic_code)
            pr_dict[key] = put_protein
    # If proteines have been encoded by any genes, they are written in a file.
    if pr_dict:
        split_size = 70
        new_file_gc_formatted = open(output_fgs_protein_formatted_2, "w")
        for key in pr_dict.keys():
            pr_sqs = pr_dict[key]
            new_file_gc_formatted.write(">{}\n".format(key))
            fasta_splited = [pr_sqs[ssi:ssi + split_size] for ssi in range(0, len(pr_sqs), split_size)]
            for fasta_line in fasta_splited:
                new_file_gc_formatted.write("{}\n".format(fasta_line))
        new_file_gc_formatted.close()
    # Compute the distance of the gene's edges from the edges of its contig(s).
    # For megahit the headers of the contigs contain their lengths.
    contig_len_dict = None
    contig_gene_dist_dict = {}
    if os.path.exists(output_final_contigs_formated):
        contig_lines = read_file(output_final_contigs_formated)
        header = None
        header_contig = None
        contig_seq = ""
        first_header = True
        contig_len_dict = {}
        for line in contig_lines:
            if line[0] == ">":
                if not first_header:
                    contig_len_dict[header_contig] = len(contig_seq)
                    contig_seq = ""
                header = line[1:]
                first_header = False
                # Process the header: "k141_49635 flag=1 multi=3.0000 len=461"
                # Keep only the contig's code without k: "141_49635"
                if " " in header:
                    header_splited = header.split(" ")
                    header_contig = header_splited[0]
                else:
                    header_contig = header
            else:
                contig_seq = "{}{}".format(contig_seq, line)
        contig_len_dict[header_contig] = len(contig_seq)
    # Find the code of the contig in which the gene was found. If the code of the contig has been corresponded to its length, then compare the length of the gene with length of the
    # contig to find the ditance of the gene from the start and end of the contig.
    for key in pr_dict.keys():
        # Example: key = contig23830_303_It-6_meta_12_197_- or key = contig24087_1789_It-6_meta_1062_1211_+
        if "_" in key:
            key_splited = key.split("_")
            origin_contig = "_".join(key_splited[0:2])
            contig_length = None
            if origin_contig in contig_len_dict.keys():
                contig_length = contig_len_dict[origin_contig]
            gene_start = int(key_splited[-3])
            gene_end = int(key_splited[-2])
            start_dist = gene_start
            end_dist = contig_length - gene_end
            contig_gene_dist_dict[key] = [start_dist, end_dist]
        else:
            contig_gene_dist_dict[key] = ["-", "-"]
    # Write the information in a file.
    new_file = open(gene_contig_dist_path, "w")
    for key in contig_gene_dist_dict.keys():
        new_file.write("{}\t{}\t{}\n".format(key, contig_gene_dist_dict[key][0], contig_gene_dist_dict[key][1]))
    new_file.close()
    return gene_sequences_dict, dict_genes, contig_gene_dist_dict


def cd_hit(cdhit_env, output_path_cdhit, prs_source, output_fgs_protein_formatted_1, output_fgs_protein_formatted_2, cd_hit_results_path, cd_hit_results_fasta_path, cd_hit_t, cd_hit_path, cd_hit_mem, thread_num, input_log_file, output_log_file, cdhit_bash_script, cd_hit_version_path, cd_hit_stdoe_path, conda_sh_path):
    print("\nRunning CD-HIT...")
    if os.path.exists(output_path_cdhit):
        shutil.rmtree(output_path_cdhit)
    os.mkdir(output_path_cdhit)
    # Use the proteins generated by the gene prediction tool or the ones encoded by the selected genetic code.
    if prs_source == 1:
        if not os.path.exists(output_fgs_protein_formatted_1):
            print("Error. Source file for proteins from FragGeneScanRs does not exist.")
        output_proteins_formatted = output_fgs_protein_formatted_1
    elif prs_source == 2:
        if not os.path.exists(output_fgs_protein_formatted_2):
            print("Error. Source file for proteins based on the encoding of the genes found from FragGeneScanRs based on the selected genetic code does not exist.")
        output_proteins_formatted = output_fgs_protein_formatted_2
    else:
        output_proteins_formatted = None
    # Number of proteins before CD-HIT.
    cd_hit_bef_pr_num = 0
    if os.path.exists(output_proteins_formatted):
        output_prs_form_lines = read_file(output_proteins_formatted)
        for line in output_prs_form_lines:
            if line[0] == ">":
                cd_hit_bef_pr_num += 1
    print("\nNumber of proteins before CD-HIT: {}".format(cd_hit_bef_pr_num))
    # If there is an output file with proteins from the gene prediction algorithm, run CD-HIT.
    if os.path.exists(output_proteins_formatted):
        if cd_hit_path:
            phrase_1 = "\"{}\" -i \"{}\" -c {} -o \"{}\" -M {} -T {} -g 1 &> \"{}\"".format(cd_hit_path, output_proteins_formatted, cd_hit_t, cd_hit_results_path, cd_hit_mem, thread_num, cd_hit_stdoe_path)
            phrase_2 = "{} -h > {}".format(cd_hit_path, cd_hit_version_path)
        else:
            phrase_1 = "cd-hit -i \"{}\" -c {} -o \"{}\" -M {} -T {} -g 1 &> \"{}\"".format(output_proteins_formatted, cd_hit_t, cd_hit_results_path, cd_hit_mem, thread_num, cd_hit_stdoe_path)
            phrase_2 = "cd-hit -h > \"{}\"".format(cd_hit_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(cdhit_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if cdhit_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(cdhit_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if cdhit_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(cdhit_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(cdhit_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Process the output file.
        if os.path.getsize(cd_hit_results_path) != 0:
            new_file_cd_hit_formatted = open(cd_hit_results_fasta_path, "w")
            with open(cd_hit_results_path) as cdhit_results_file:
                for line in cdhit_results_file:
                    line = line.rstrip("\n")
                    if line[0] == ">":
                        new_file_cd_hit_formatted.write("{}\n".format(line))
                    else:
                        split_size = 70
                        fasta_splited = [line[ssi:ssi + split_size] for ssi in range(0, len(line), split_size)]
                        for fasta_line in fasta_splited:
                            new_file_cd_hit_formatted.write("{}\n".format(fasta_line))
            new_file_cd_hit_formatted.close()
    else:
        print("No genes found.")
        exit()
    # The number of proteins remaining after CD-HIT which are to be checked for the specified domains.
    protein_titles = []
    if os.path.exists(cd_hit_results_path):
        with open(cd_hit_results_path) as contig_file:
            for line in contig_file:
                line = line.rstrip("\n")
                if line[0] == ">":
                    protein_titles.append(line)
    prs_cd_hit_num = len(protein_titles)
    print("\nNumber of proteins after CD-HIT: {}".format(prs_cd_hit_num))


def hmmer_spec(hmmer_env, output_path_hmmer, hmmer_dmbl_results, hmmer_simple_results, hmmer_enz_domains_all_proteins, cd_hit_results_path, hmmscan_path, profiles_path, val_type, file_prs_seq_enz_domains_name, thread_num, input_log_file, output_log_file, hmmer_spec_bash_script, phylo_analysis, hmmscan_spec_version_path, hmmscan_spec_stdoe_path, conda_sh_path):
    print("\nRunning HMMER...")
    if not phylo_analysis:
        if os.path.exists(output_path_hmmer):
            shutil.rmtree(output_path_hmmer)
        os.mkdir(output_path_hmmer)
    else:
        if not os.path.exists(output_path_hmmer):
            os.mkdir(output_path_hmmer)
    # If the proteins used for the analysis, after CD-HIT are not found by FragGeneScanRs then it will be needed to add a step where the headers of the proteins will not have empty characters because HMMER only reports the headers
    # of the protein to which it finds domains based only on their part of the name until it meets the first empty character. FragGeneScanRs produces names without empty characters so there is no problem currently. If another gene prediction program is used there might be.
    if os.path.exists(cd_hit_results_path) and os.path.exists(profiles_path):
        if hmmscan_path:
            phrase_1 = "\"{}\" -o \"{}\" --domtblout \"{}\" --notextw {}--cpu {} \"{}\" \"{}\" &> \"{}\"".format(hmmscan_path, hmmer_simple_results, hmmer_dmbl_results, val_type, thread_num, profiles_path, cd_hit_results_path, hmmscan_spec_stdoe_path)
            phrase_2 = "\"{}\" -h > \"{}\"".format(hmmscan_path, hmmscan_spec_version_path)
        else:
            phrase_1 = "hmmscan -o \"{}\" --domtblout \"{}\" --notextw {}--cpu {} \"{}\" \"{}\" &> \"{}\"".format(hmmer_simple_results, hmmer_dmbl_results, val_type, thread_num, profiles_path, cd_hit_results_path, hmmscan_spec_stdoe_path)
            phrase_2 = "hmmscan -h > \"{}\"".format(hmmscan_spec_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(hmmer_spec_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if hmmer_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(hmmer_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if hmmer_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(hmmer_spec_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(hmmer_spec_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
    # Processing the output from hmmscan.
    pattern_domains = re.compile(r'^(\S+)\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+(.*)')
    # It is mandatory to write all the domains found for one protein in contineous lines, that is the reason for using a dictionary.
    domains_dict_spec = {}
    prs_with_enz_domains = []
    if os.path.exists(hmmer_dmbl_results):
        with open(hmmer_dmbl_results) as db_hmmer_file:
            for line in db_hmmer_file:
                line = line.rstrip("\n")
                if line[0] != "#":
                    result_domains = pattern_domains.search(line)
                    if result_domains:
                        pfam_name_short = result_domains.group(1)
                        pfam_accession = result_domains.group(2)
                        pfam_name_long = result_domains.group(12)
                        protein_name = result_domains.group(3)
                        e_value = result_domains.group(4)
                        score = result_domains.group(5)
                        c_e_value = result_domains.group(6)
                        i_e_value = result_domains.group(7)
                        hmm_from = result_domains.group(8)
                        hmm_to = result_domains.group(9)
                        env_from = result_domains.group(10)
                        env_to = result_domains.group(11)
                        line_w = ("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(protein_name, e_value, score, pfam_name_short, pfam_accession, pfam_name_long, c_e_value, i_e_value, hmm_from, hmm_to, env_from, env_to))
                        if protein_name not in domains_dict_spec.keys():
                            domains_dict_spec[protein_name] = [line_w]
                        else:
                            domains_dict_spec[protein_name].append(line_w)
                        if protein_name not in prs_with_enz_domains:
                            prs_with_enz_domains.append(protein_name)
    # The information are written in a file.
    if domains_dict_spec.keys():
        new_file_hmmer_output = open(hmmer_enz_domains_all_proteins, "w")
        new_file_hmmer_output.write("Protein_name\tE_value\tScore\tPfam_domain_short_name\tPfam_domain_accession_name\tPfam_domain_long_name\tc-E-value\ti-E-value\thmm_from\thmm_to\tenv_from\tenv_to\n")
        for key_pr in domains_dict_spec.keys():
            for domains_line in domains_dict_spec[key_pr]:
                new_file_hmmer_output.write("{}\n".format(domains_line))
        new_file_hmmer_output.close()
    else:
        print("\nNo proteins found with enzyme domains.\n")
    # Write the proteins found with the specified enzyme domains in a file.
    if prs_with_enz_domains:
        new_file = open(file_prs_seq_enz_domains_name, "w")
        with open(cd_hit_results_path) as cdhit_results_file:
            for line in cdhit_results_file:
                line = line.rstrip("\n")
                if line[0] == ">":
                    # The status showing a protein with any of the specified domains for a new protein sequence is
                    # initialized to False.
                    enz_pr_status = False
                    # If any header/name of the proteins found to have such domains is present in the current header,
                    # then this sequences is a protein with at least one of the specified domains.
                    for prd in prs_with_enz_domains:
                        if prd in line:
                            enz_pr_status = True
                if enz_pr_status:
                    new_file.write("{}\n".format(line))
        new_file.close()
    return prs_with_enz_domains


def analyze_no_enzs(diamond_env, output_path_hmmer, output_path_blastp, prs_with_enz_domains, blastp_results_no_doms_nr_file, blastp_info_no_doms_nr_file, blastp_no_doms_below_threshold, blastp_info_no_doms_below_threshold, cd_hit_results_path, file_prs_seq_no_enzs_name, fpr_db_fasta, fpr_db_name, e_value_nodom_thr, thread_num, input_log_file, output_log_file, diamond_path, blastp_nodoms_script, diamond_blastp_nr1_stdoe_path, conda_sh_path):
    print("\nRunning DIAMOND BLASTP for proteins without seek domains...")
    # This is the second time this folder is needed, so it has to be created only if it has not been already created,
    # but not overwrite the folder of this path, if it has already been created.
    if not os.path.exists(output_path_hmmer):
        os.mkdir(output_path_hmmer)
    # This is the first time this folder is needed, so it has to be created or overwrite this folder path.
    if os.path.exists(output_path_blastp):
        shutil.rmtree(output_path_blastp)
    os.mkdir(output_path_blastp)
    # Patterns
    pattern_blastp_info = re.compile(r'^(\S+)\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+(\S+)\s+(\S+)')
    # The list of proteins with the specified enzyme domains is either empty (if no analysi mode 1 occured) or it has
    # already been formed after analysis mode 1.
    # Write the proteins which did not have enzyme domains in a file.
    # If the analysis for pecified domains is not performed (or proteins with those domains have not been found) then
    # all of the proteins are written in a file and are further analyzed.
    prs_blast_thr = []
    new_file_prs_blast_thr = open(file_prs_seq_no_enzs_name, "w")
    with open(cd_hit_results_path) as cdhit_results_file:
        for line in cdhit_results_file:
            line = line.rstrip("\n")
            if line[0] == ">":
                enz_protein = False
                for i in prs_with_enz_domains:
                    if i in line:
                        enz_protein = True
            # The header of the protein sequence has already determined whether a enzyme domain has been found in the protein or not. Thus every line after the header will be written in the file
            # only if in the protein no enzyme domains had been found. Otherwise no line of the protein's sequence will be written in the file.
            if not enz_protein:
                if line[0] == ">":
                    prs_blast_thr.append(line)
                new_file_prs_blast_thr.write("{}\n".format(line))
    new_file_prs_blast_thr.close()
    # BLASTP for the proteins without enzyme domains.
    if os.path.exists(file_prs_seq_no_enzs_name):
        if os.path.exists(fpr_db_fasta):
            # Start time
            if diamond_path:
                phrase_1 = "\"{}\" blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(diamond_path, thread_num, fpr_db_name, blastp_results_no_doms_nr_file, file_prs_seq_no_enzs_name, diamond_blastp_nr1_stdoe_path)
                phrase_2 = ""
            else:
                phrase_1 = "diamond blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(thread_num, fpr_db_name, blastp_results_no_doms_nr_file, file_prs_seq_no_enzs_name, diamond_blastp_nr1_stdoe_path)
                phrase_2 = ""
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(blastp_nodoms_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if diamond_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(diamond_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            if phrase_2:
                new_file_bash.write("{}\n".format(phrase_2))
            if diamond_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(blastp_nodoms_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(blastp_nodoms_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        else:
            print("\nBLASTP of the proteins without any of the seek domains was not performed against the sfpd. The sfpd ({}) was not located.".format(fpr_db_fasta))
        # Collect results from BLASTP in the annotation folder.
        blastp_output_info = open(blastp_info_no_doms_nr_file, "w")
        blastp_output_info.write("Protein_name\tNR_ID\tPercentage_identity\tE_value\tBitscore\n")
        if os.path.exists(blastp_results_no_doms_nr_file):
            with open(blastp_results_no_doms_nr_file) as blastp_lines:
                for line in blastp_lines:
                    line = line.rstrip("\n")
                    result_blastp_info = pattern_blastp_info.search(line)
                    if result_blastp_info:
                        protein_name = result_blastp_info.group(1)
                        uniprot_ac = result_blastp_info.group(2)
                        # Insert check for name of entry with that Uniprot ID and possible GO annotations with it.
                        percentage_identity = result_blastp_info.group(3)
                        e_value = result_blastp_info.group(4)
                        bitscore = result_blastp_info.group(5)
                        blastp_output_info.write("{}\t{}\t{}\t{}\t{}\n".format(protein_name, uniprot_ac, percentage_identity, e_value, bitscore))
        blastp_output_info.close()
    else:
        print("\nBLASTP of the proteins without any of the seek domains was not performed against the sfpd. The file ({}) with the proteins without any of the seek domains, was not located.".format(file_prs_seq_no_enzs_name))
    # Information gathering, based on a given threshold, from the results of BLASTP of proteins without enzyme domains against the reductd nr.
    protein_ids_below_thr = []
    if os.path.exists(blastp_info_no_doms_nr_file):
        e_value_nodom_thr = float(e_value_nodom_thr)
        blastp_lines = read_file(blastp_info_no_doms_nr_file)
        header = True
        header_write = True
        for line in blastp_lines:
            if header:
                header = False
                continue
            line_splited = line.split("\t")
            protein_acc = line_splited[0]
            e_value = float(line_splited[-2])
            # 0.00001 = 1e-5
            if e_value <= e_value_nodom_thr:
                if protein_acc not in protein_ids_below_thr:
                    protein_ids_below_thr.append(protein_acc)
                if header_write:
                    new_file = open(blastp_info_no_doms_below_threshold, "w")
                    new_file.write("Protein_name\tNR_ID\tPercentage_identity\tE_value\tBitscore\n")
                    header_write = False
                new_file.write("{}\n".format(line))
        # If the file opened then close it. This can be checked based on whether the header line for the new file was written in it or not.
        if not header_write:
            new_file.close()
        # Write the sequences of the proteins with no domains and e-value euqal to or smaller than the set threshold.
        if protein_ids_below_thr:
            new_file = open(blastp_no_doms_below_threshold, "w")
            with open(cd_hit_results_path) as lines_seqs:
                for line_pr_seq in lines_seqs:
                    line_pr_seq = line_pr_seq.rstrip("\n")
                    if line_pr_seq[0] == ">":
                        protein_boolean = False
                        protein_acc = line_pr_seq[1:]
                        # The protein_ids_below_thr list contain the IDs of all the proteins which were found not to have enzyme domains and whose best hit in NR had an e-value lower than the specified threshold.
                        if protein_acc in protein_ids_below_thr:
                            protein_boolean = True
                    if protein_boolean:
                        new_file.write("{}\n".format(line_pr_seq))
            new_file.close()
    return prs_blast_thr, protein_ids_below_thr


def combine_predictions(hmmer_enz_domains_all_proteins, blastp_no_doms_below_threshold, proteins_combined_file):
    print("\nCombining predictions...")
    if os.path.exists(hmmer_enz_domains_all_proteins) and os.path.exists(blastp_no_doms_below_threshold):
        filenames = [hmmer_enz_domains_all_proteins, blastp_no_doms_below_threshold]
    elif os.path.exists(hmmer_enz_domains_all_proteins):
        filenames = [hmmer_enz_domains_all_proteins]
    elif os.path.exists(blastp_no_doms_below_threshold):
        filenames = [blastp_no_doms_below_threshold]
    else:
        filenames = []
    with open(proteins_combined_file, 'w') as outfile:
        for fname in filenames:
            with open(fname) as infile:
                for line in infile:
                    outfile.write(line)


def hmmer_broad(hmmer_env, prs_with_enz_domains, hmmer_dmbl_results, hmmer_simple_results, hmmer_enz_domains_all_proteins, proteins_combined_file, comb_dmbl_results, comb_simple_results, comb_all_domains_proteins, profiles_broad_path, hmmscan_path, val_type, second_dom_search, thread_num, input_log_file, output_log_file, hmmer_broad_bash_script, hmmscan_broad_version_path, hmmscan_broad_stdoe_path, conda_sh_path):
    print("\nRunning HMMER against Pfam...")
    # Example:
    #                                                                                    --- full sequence --- -------------- this domain -------------   hmm coord   ali coord   env coord
    # target name        accession   tlen query name                   accession   qlen   E-value  score  bias   #  of  c-Evalue  i-Evalue  score  bias  from    to  from    to  from    to  acc description
    # (Pro_CA)               (PF00484.21)   156 (contig270_439198_it3_meta_344131_344436_-) -            101   (3.5e-11)   (32.7)   0.1   1   1   4.7e-12   3.7e-11   32.7   0.1    38    85    12    55     1    86 0.74 (Carbonic anhydrase)
    pattern_domains = re.compile(r'^(\S+)\s+(\S+)\s+\S+\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+(\S+)\s+(\S+)\s+\S+\s+(.*)')
    # Rerun HMMER against all HMMs only for the proteins identified with any of the known enzyme domains.
    # File with protein sequences which have enzyme domains.
    # The file with all the representative sequences from CD-HIT is parsed and every header of the file is checked
    # on whether it contains any of the headers of the proteins found with enzyme domains. If such a protein is found
    # then it is removed from the list with the headers of proteins found to have enzyme domains as not to be checked
    # again in any of the lines of the (CD-HIT) file.
    dict_hm = {}
    if os.path.exists(proteins_combined_file):
        # Run against all HMMs of the second database (if any).
        # In case someone nees to use all the domains available from Pfam at the first search then there is no need to make
        # the second search. Therefore, in that case either the second database is left empty or the option second_dom_search
        # is to false.
        if (os.path.exists(profiles_broad_path)) and (second_dom_search is not False):
            if hmmscan_path:
                phrase_1 = "\"{}\" -o \"{}\" --domtblout \"{}\" --notextw {}--cpu {} \"{}\" \"{}\" &> \"{}\"".format(hmmscan_path, comb_simple_results, comb_dmbl_results, val_type, thread_num, profiles_broad_path, proteins_combined_file, hmmscan_broad_stdoe_path)
                phrase_2 = "\"{}\" -h > \"{}\"".format(hmmscan_path, hmmscan_broad_version_path)
            else:
                phrase_1 = "hmmscan -o \"{}\" --domtblout \"{}\" --notextw {}--cpu {} \"{}\" \"{}\" &> \"{}\"".format(comb_simple_results, comb_dmbl_results, val_type, thread_num, profiles_broad_path, proteins_combined_file, hmmscan_broad_stdoe_path)
                phrase_2 = "hmmscan -h > \"{}\"".format(hmmscan_broad_version_path)
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(hmmer_broad_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if hmmer_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(hmmer_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            new_file_bash.write("{}\n".format(phrase_2))
            if hmmer_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(hmmer_broad_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(hmmer_broad_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Recollect information
            # The information to collect from the hits of all domains in the proteins with enzyme domain are set here.
            # It is mandatory to write all the domains found for one protein in contineous lines, that is the reason for using a dictionary.
            if os.path.exists(comb_dmbl_results):
                with open(comb_dmbl_results) as db_hmmer_file:
                    for line in db_hmmer_file:
                        line = line.rstrip("\n")
                        if line[0] != "#":
                            result_domains = pattern_domains.search(line)
                            if result_domains:
                                pfam_name_short = result_domains.group(1)
                                pfam_accession = result_domains.group(2)
                                pfam_name_long = result_domains.group(12)
                                protein_name = result_domains.group(3)
                                e_value = result_domains.group(4)
                                score = result_domains.group(5)
                                c_e_value = result_domains.group(6)
                                i_e_value = result_domains.group(7)
                                hmm_from = result_domains.group(8)
                                hmm_to = result_domains.group(9)
                                env_from = result_domains.group(10)
                                env_to = result_domains.group(11)
                                line_w = ("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(protein_name, e_value, score, pfam_name_short, pfam_accession, pfam_name_long, c_e_value, i_e_value, hmm_from, hmm_to, env_from, env_to))
                                if protein_name not in dict_hm.keys():
                                    dict_hm[protein_name] = [line_w]
                                else:
                                    dict_hm[protein_name].append(line_w)
            # The information are written in a file.
            if dict_hm.keys():
                new_file = open(comb_all_domains_proteins, "w")
                new_file.write("Protein_name\tE_value\tScore\tPfam_domain_short_name\tPfam_domain_accession_name\tPfam_domain_long_name\tc-E-value\ti-E-value\thmm_from\thmm_to\tenv_from\tenv_to\n")
                for key_pr in dict_hm.keys():
                    for domains_line in dict_hm[key_pr]:
                        new_file.write("{}\n".format(domains_line))
                new_file.close()
        else:
            # If no such database exists, then just copy the results of the search based on the firt profile database.
            if os.path.exists(hmmer_dmbl_results):
                shutil.copyfile(hmmer_dmbl_results, comb_dmbl_results)
            if os.path.exists(hmmer_simple_results):
                shutil.copyfile(hmmer_simple_results, comb_simple_results)
            if os.path.exists(hmmer_enz_domains_all_proteins):
                shutil.copyfile(hmmer_enz_domains_all_proteins, comb_all_domains_proteins)
    return prs_with_enz_domains, dict_hm


def dom_percs(comb_all_domains_proteins, hmmer_profile_lengths_file, hmmer_common_lengths_file):
    print("\nComputing profile coverages...")
    if os.path.exists(comb_all_domains_proteins):
        profiles_lengths_lines = read_file(hmmer_profile_lengths_file)
        profiles_codes_lengths_dict = {}
        # Collect the length of each profile.
        for line in profiles_lengths_lines:
            line_splited = line.split("\t")
            profile_code = line_splited[1]
            profile_length = int(line_splited[2])
            if profile_code in profiles_codes_lengths_dict.keys():
                print("Profile code duplicate from lengths info found!")
                exit()
            else:
                profiles_codes_lengths_dict[profile_code] = profile_length
        new_file = open(hmmer_common_lengths_file, "w")
        # Parse each protein from the file of combined proteins.
        domains_info_lines = read_file(comb_all_domains_proteins)
        for i in range(1, len(domains_info_lines)):
            line = domains_info_lines[i]
            splited_line = line.split("\t")
            contig_name = splited_line[0]
            domain_code = splited_line[4]
            profile_common = int(splited_line[9]) - int(splited_line[8])
            # If for some reason the domain code is not found in the dictionary of profile lengths,
            # the profile coverage is undefined and it is written as -1.
            if domain_code in profiles_codes_lengths_dict.keys():
                profile_length_cur = profiles_codes_lengths_dict[domain_code]
                profile_coverage = profile_common / profile_length_cur
                profile_coverage = round(profile_coverage, 2)
                profile_coverage_perc = profile_coverage * 100
                profile_coverage_perc = int(profile_coverage_perc)
            else:
                profile_coverage_perc = -1
            new_file.write("{}\t{}%\n".format(contig_name, profile_coverage_perc))
        new_file.close()


def blastp(diamond_env, output_path_blastp, blastp_results_swissprot_file, blastp_doms_nr_file, blastp_info_swissprot_file, blastp_doms_info_nr_file, file_prs_seq_enz_domains_name, proteins_combined_file, swissprot_path, fpr_db_fasta, fpr_db_name, thread_num, input_log_file, output_log_file, diamond_path, blast_swissprot_bash_script, blast_nr_bash_script, phylo_analysis, conda_sh_path, diamond_blastp_swissprot_stdoe_path, diamond_blastp_nr2_stdoe_path):
    print("\nRunning DIAMOND BLASTP for the combined dataset...")
    # Second time this folder is needed, so if it already exits, then it is not overwritten. It is creted only if it
    # has not already.
    if not os.path.exists(output_path_blastp):
        os.mkdir(output_path_blastp)
    # Initialiazing variables
    dict_sp = None
    # Pattern
    pattern_blastp_info = re.compile(r'^(\S+)\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+\S+\s+(\S+)\s+(\S+)')
    # If the file with the combined predictions for proteins exists.
    # Proteins from analysis mode 1 need to be blasted against the databases: SwissProt, nr
    # Proteins from analysis mode 2 need to be blasted against the databases: SwissProt
    # Therefore the combined proteines are blasted against SwissProt and only the proteins from analysis
    # mdoe 1 are blasted against nr and their results are added to the result of Blastp against nr of the proteins
    # from analysi mode 2.
    # ---------BLASTP: combined proteins VS swissprot---------
    if os.path.exists(proteins_combined_file):
        # If the SwissProt database exists.
        if os.path.exists(swissprot_path):
            if diamond_path:
                phrase_1 = "\"{}\" blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(diamond_path, thread_num, swissprot_path, blastp_results_swissprot_file, proteins_combined_file, diamond_blastp_swissprot_stdoe_path)
                phrase_2 = ""
            else:
                phrase_1 = "diamond blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(thread_num, swissprot_path, blastp_results_swissprot_file, proteins_combined_file, diamond_blastp_swissprot_stdoe_path)
                phrase_2 = ""
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(blast_swissprot_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if diamond_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(diamond_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            if phrase_2:
                new_file_bash.write("{}\n".format(phrase_2))
            if diamond_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(blast_swissprot_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(blast_swissprot_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        else:
            print("\nBlastp was not performed against the Swiss-Prot database.")
        # Gathering information.
        # Collect results from BLASTP in the annotation folder for the SwissProt database. Based on this database the protein codes are checked for the GO annotations.
        dict_sp = {}
        min_evalue = None
        new_file = open(blastp_info_swissprot_file, "w")
        new_file.write("Protein_name\tUniprot_AC\tPercentage_identity\tE_value\tBitscore\n")
        if os.path.exists(blastp_results_swissprot_file):
            with open(blastp_results_swissprot_file) as blastp_lines:
                for line in blastp_lines:
                    line = line.rstrip("\n")
                    result_blastp_info = pattern_blastp_info.search(line)
                    if result_blastp_info:
                        protein_name = result_blastp_info.group(1)
                        uniprot_ac = result_blastp_info.group(2)
                        # Insert check for name of entry with that Uniprot ID and possible GO annotations with it.
                        percentage_identity = result_blastp_info.group(3)
                        e_value = float(result_blastp_info.group(4))
                        bitscore = result_blastp_info.group(5)
                        new_line = "{}\t{}\t{}\t{}\t{}".format(protein_name, uniprot_ac, percentage_identity, e_value, bitscore)
                        new_file.write("{}\n".format(new_line))
                        # Check if the current e-value of the Uniprot AC for the protein name (if any) is smaller or higher. If higher continue. If smaller change the value of the
                        # dictionary for the protein_name with the current line of the Uniprot AC.
                        if protein_name not in dict_sp.keys():
                            dict_sp[protein_name] = new_line
                            min_evalue = e_value
                        else:
                            if e_value < min_evalue:
                                dict_sp[protein_name] = new_line
                                min_evalue = e_value
        new_file.close()
    # ---------BLASTP: phylo domain proteins VS swissprot---------
    # Blastp of protein with phylo domains against the Swiss-Prot database.
    if phylo_analysis:
        # If the SwissProt database exists.
        if os.path.exists(swissprot_path) and os.path.exists(file_prs_seq_enz_domains_name):
            if diamond_path:
                phrase_1 = "\"{}\" blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(diamond_path, thread_num, swissprot_path, blastp_results_swissprot_file, file_prs_seq_enz_domains_name, diamond_blastp_swissprot_stdoe_path)
                phrase_2 = ""
            else:
                phrase_1 = "diamond blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(thread_num, swissprot_path, blastp_results_swissprot_file, file_prs_seq_enz_domains_name, diamond_blastp_swissprot_stdoe_path)
                phrase_2 = ""
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(blast_swissprot_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if diamond_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(diamond_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            if phrase_2:
                new_file_bash.write("{}\n".format(phrase_2))
            if diamond_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(blast_swissprot_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(blast_swissprot_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        else:
            print("\nBlastp was not performed against the Swiss-Prot database.")
        # Gathering information.
        # Collect results from BLASTP in the annotation folder for the SwissProt database. Based on this database the protein codes are checked for the GO annotations.
        dict_sp = {}
        min_evalue = None
        new_file = open(blastp_info_swissprot_file, "w")
        new_file.write("Protein_name\tUniprot_AC\tPercentage_identity\tE_value\tBitscore\n")
        if os.path.exists(blastp_results_swissprot_file):
            with open(blastp_results_swissprot_file) as blastp_lines:
                for line in blastp_lines:
                    line = line.rstrip("\n")
                    result_blastp_info = pattern_blastp_info.search(line)
                    if result_blastp_info:
                        protein_name = result_blastp_info.group(1)
                        uniprot_ac = result_blastp_info.group(2)
                        # Insert check for name of entry with that Uniprot ID and possible GO annotations with it.
                        percentage_identity = result_blastp_info.group(3)
                        e_value = float(result_blastp_info.group(4))
                        bitscore = result_blastp_info.group(5)
                        new_line = "{}\t{}\t{}\t{}\t{}".format(protein_name, uniprot_ac, percentage_identity, e_value, bitscore)
                        new_file.write("{}\n".format(new_line))
                        # Check if the current e-value of the Uniprot AC for the protein name (if any) is smaller or higher. If higher continue. If smaller change the value of the
                        # dictionary for the protein_name with the current line of the Uniprot AC.
                        if protein_name not in dict_sp.keys():
                            dict_sp[protein_name] = new_line
                            min_evalue = e_value
                        else:
                            if e_value < min_evalue:
                                dict_sp[protein_name] = new_line
                                min_evalue = e_value
        new_file.close()
    # ---------BLASTP: seek domain proteins VS swissprot---------
    # Specifically for the nr database, if the analysis mode if 2 or 3 only the proteins from mode analysis 1 are
    # uesd in BLAST.
    if os.path.exists(file_prs_seq_enz_domains_name):
        # If proteins from analysis mode 1 an the nr database exist, then run the proteins from analysis mode 1
        # against the nr database.
        if os.path.exists(fpr_db_fasta):
            if diamond_path:
                phrase_1 = "\"{}\" blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(diamond_path, thread_num, fpr_db_name, blastp_doms_nr_file, file_prs_seq_enz_domains_name, diamond_blastp_nr2_stdoe_path)
                phrase_2 = ""
            else:
                phrase_1 = "diamond blastp --threads {} --db \"{}\" --out \"{}\" --query \"{}\" --outfmt 6 &> \"{}\"".format(thread_num, fpr_db_name, blastp_doms_nr_file, file_prs_seq_enz_domains_name, diamond_blastp_nr2_stdoe_path)
                phrase_2 = ""
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(blast_nr_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if diamond_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(diamond_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            if phrase_2:
                new_file_bash.write("{}\n".format(phrase_2))
            if diamond_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(blast_nr_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(blast_nr_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Collect results from BLASTP in the annotation folder for the filtered nr database.
            new_file = open(blastp_doms_info_nr_file, "w")
            new_file.write("Protein_name\tNR_ID\tPercentage_identity\tE_value\tBitscore\n")
            if os.path.exists(blastp_doms_nr_file):
                with open(blastp_doms_nr_file) as blastp_lines:
                    for line in blastp_lines:
                        line = line.rstrip("\n")
                        result_blastp_info = pattern_blastp_info.search(line)
                        if result_blastp_info:
                            protein_name = result_blastp_info.group(1)
                            uniprot_ac = result_blastp_info.group(2)
                            # Insert check for name of entry with that Uniprot ID and possible GO annotations with it.
                            percentage_identity = result_blastp_info.group(3)
                            e_value = result_blastp_info.group(4)
                            bitscore = result_blastp_info.group(5)
                            new_file.write("{}\t{}\t{}\t{}\t{}\n".format(protein_name, uniprot_ac, percentage_identity, e_value, bitscore))
            new_file.close()
    return dict_sp


def combine_nr(blastp_info_no_doms_below_threshold, blastp_doms_info_nr_file, blastp_info_comb_nr_file):
    print("\nCombining predictions from the screen against the seek filtered protein database (sfpd)...")
    if os.path.exists(blastp_info_no_doms_below_threshold) and os.path.exists(blastp_doms_info_nr_file):
        filenames = [blastp_info_no_doms_below_threshold, blastp_doms_info_nr_file]
    elif os.path.exists(blastp_info_no_doms_below_threshold):
        filenames = [blastp_info_no_doms_below_threshold]
    elif os.path.exists(blastp_doms_info_nr_file):
        filenames = [blastp_doms_info_nr_file]
    else:
        filenames = []
    # The first line (header) of the first file is the only one written in file from the first lines
    # of any of the files.
    dict_nr = {}
    min_evalue = None
    with open(blastp_info_comb_nr_file, 'w') as outfile:
        outfile.write("Protein_name\tNR_ID\tPercentage_identity\tE_value\tBitscore\n")
        for i in range(0, len(filenames)):
            fname = filenames[i]
            header = True
            with open(fname) as infile:
                for line in infile:
                    if not header:
                        outfile.write(line)
                        line_splited = line.split("\t")
                        pr_name = line_splited[0]
                        e_value = float(line_splited[3])
                        if pr_name not in dict_nr.keys():
                            dict_nr[pr_name] = line
                            min_evalue = e_value
                        else:
                            if e_value < min_evalue:
                                dict_nr[pr_name] = line
                                min_evalue = e_value
                    header = False
    return dict_nr


def bin_phylo(fpr_db_fasta, blastp_doms_info_nr_file, dict_contigs_bins, binning_results_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prstax, binphylo_path_binstax, binphylo_path_binstax_max, binphylo_path_prsids, binphylo_path_binstax_names, binphylo_path_binstax_max_names, binphylo_freq_taxids_path, binphylo_maxfreq_taxids_path, cd_hit_results_fasta_path, taxonkit_env, taxonkit_path, tax_freq_bash_script, tax_maxfreq_bash_script, taxoknit_freq_version_path, taxoknit_freq_stdoe_path, taxoknit_maxfreq_version_path, taxoknit_maxfreq_stdoe_path, thread_num, conda_sh_path, freq_taxids_path, maxfreq_taxids_path, taxonkit_freq_line_bash_script, taxonkit_maxfreq_line_bash_script, taxonkit_freq_line_version_path, taxonkit_freq_line_stdoe_path, taxonkit_maxfreq_line_version_path, taxonkit_maxfreq_line_stdoe_path, freq_lineage_path, maxfreq_lineage_path, freq_lineage_form_path, maxfreq_lineage_form_path, csvtk_freq_version_path, csvtk_freq_stdoe_path, csvtk_maxfreq_version_path, csvtk_maxfreq_stdoe_path, family_to_profile_phylo_dict, hmmer_enz_domains_all_proteins_phylo, family_profile_path, input_log_file, output_log_file):
    print("\nPerforming a taxonomic analysis of the bins...")
    # The process is:
    # 1. Run HMMER spec for the proteins of CD-HIT for each database selected for phylogenetic analysis (these databases are already created).
    # 2. For each identified hit, run the corresponding nr database (these databases are already created).
    # 3. At this stage use the results of each of these proteins (the bins do not change) and sum them up to concluce the taxonomy(ies) of each bin and of each protein.
    # The pnr database must be parsed and collect the correspondance between nr ACs and their organisms.
    dict_pnr_tax = {}
    pattern_acc = re.compile(r'^>(\S+)')
    pattern_tax_1 = re.compile(r'\[(.*?)\]')
    pattern_tax_2 = re.compile(r'Tax=(.*?) TaxID=')
    if os.path.exists(fpr_db_fasta):
        with open(fpr_db_fasta) as pnr_lines:
            for line in pnr_lines:
                line = line.rstrip("\n")
                if line[0] == ">":
                    nr_acc = None
                    nr_tax = None
                    # Search for the accession number.
                    result_acc = pattern_acc.search(line)
                    if result_acc:
                        # The taxa name inside the the first brackets is retained.
                        nr_acc = result_acc.group(1)
                    # Pattern 1
                    result_tax_1 = re.findall(pattern_tax_1, line)
                    if result_tax_1:
                        nr_tax = []
                        # Count the species related to the protein sequence.
                        nr_tax_dict = {}
                        for taxum in result_tax_1:
                            if taxum not in nr_tax_dict.keys():
                                nr_tax_dict[taxum] = 1
                            else:
                                nr_tax_dict[taxum] += 1
                        # If only one species was found, keep that species, otherwise keep the one(s) with the highest count.
                        if len(list(nr_tax_dict.keys()))  == 1:
                            nr_tax = [result_tax_1[0]]
                        else:
                            max_count = 0
                            for cur_count in nr_tax_dict.values():
                                if cur_count > max_count:
                                    max_count = cur_count
                            for cut_tax in nr_tax_dict.keys():
                                if nr_tax_dict[cut_tax] == max_count:
                                    nr_tax.append(cut_tax)
                    # Pattern 2
                    result_tax_2 = re.search(pattern_tax_2, line)
                    if result_tax_2:
                        nr_tax = []
                        taxum = result_tax_2.group(1)
                        nr_tax.append(taxum)
                    # Result
                    if nr_acc is not None:
                        if nr_tax is not None:
                            if nr_tax:
                                dict_pnr_tax[nr_acc] = nr_tax
    else:
        return {}, {}, {}, {}, {}, {}

    # NR results
    dict_nr = {}
    if os.path.exists(blastp_doms_info_nr_file):
        protein_acc = None
        min_evalue = None
        header = True
        with open(blastp_doms_info_nr_file) as lines_nr:
            for line_nr in lines_nr:
                if header:
                    header = False
                    continue
                line_nr = line_nr.rstrip("\n")
                splited_nr = line_nr.split("\t")
                protein_acc = splited_nr[0]
                e_value = float(splited_nr[3])
                # Line: k141_39729_24847_26913_-    cbw76918.1      43.0    1.70e-159   473
                # Line: k141_39729_24847_26913_-    wp_014905841.1  44.2    1.20e-156   464
                # Line: k141_29822_4122_5963_-      cab1404101.1    58.3    4.40e-246   686
                if protein_acc not in dict_nr.keys():
                    dict_nr[protein_acc] = line_nr
                    min_evalue = e_value
                else:
                    if e_value < min_evalue:
                        dict_nr[protein_acc] = line_nr
                        min_evalue = e_value
    else:
        print("\nThe path to the file with the nr information, for the phylogenetic analysis of the bins, is wrong. The phylogenetic analysis based on the bins is not performed.")
    
    # Binning results
    dict_prs_bins = {}
    dict_bins_prs = {}
    if not dict_contigs_bins:
        if os.path.exists(binning_results_path):
            with open(binning_results_path) as und_lines:
                tsv_lines = csv.reader(und_lines, delimiter="\t", quotechar='"')
                for line in tsv_lines:
                    contig_id = line[0]
                    bin_id = line[1]
                    if bin_id[:5] == "group":
                        bin_id = bin_id[5:]
                    dict_contigs_bins[contig_id] = bin_id
        dict_prs_contigs = {}
        if os.path.exists(output_path_genepred) and os.path.exists(output_path_genepred_faa):
            protein_lines = read_file(output_path_genepred_faa)
            for line in protein_lines:
                if line[0] == ">":
                    protein_id = line[1:]
                    if "_" in protein_id:
                        protein_id_splited = protein_id.split("_")
                        origin_contig = "_".join(protein_id_splited[0:2])
                        dict_prs_contigs[protein_id] = origin_contig
        if dict_contigs_bins and dict_prs_contigs:
            for pr_id in dict_prs_contigs.keys():
                contig_id = dict_prs_contigs[pr_id]
                if contig_id in dict_contigs_bins.keys():
                    bin_id = dict_contigs_bins[contig_id]
                else:
                    bin_id = "-"
                dict_prs_bins[pr_id] = bin_id
                if bin_id not in dict_bins_prs.keys():
                    dict_bins_prs[bin_id] = [pr_id]
                else:
                    dict_bins_prs[bin_id].append(pr_id)
    if (not dict_prs_bins) or (not dict_bins_prs):
        return {}, {}, {}, {}, {}, {}

    # dict_pnr_tax:
    # wp_088159341.1: achromobacter xylosoxidans
    # dict_nr:
    # k141_10192_17880_18695_-: k141_10192_17880_18695_-\tmbh0196785.1\t59.0\t1.58e-55\t187
    # dict_prs_bins:
    # k141_49593_60106_60408_+: 2
    # k141_49596_2_409_+: -
    # The blastp_info_comb_nr_file contains the combined information from the proteins wihtout any domain of interest that scored e-values below the threshold against pnr and from the proteins
    # each which had at least one hit against the phmm database for the selected protein family.
    # Parse each bin:
    # Parse each of its proteins:
    # Check if the protein in part of the proteins in the file blastp_info_comb_nr_file.
    # If no, conitnue on.
    # If yes, collect the taxonomy of the protein.
    # If no such protein was found, then no taxonomy is assigned to the bin.
    # If at least one such protein was found for the bin then:
    # Assign to the bin, the the taxonomy of the highest frequency based on the proteins parse above.
    # If two or more proteins were found for the bin and two or more taxonomies had the highest frequency, then assign to the bin all taxonomies.
    # All genes and proteins of the bin are assigned the taxonomy(ies) of the bin, if any.
    # ---
    # Sort the bins in ascending order of the bin IDs.
    # Create a list of the bin IDs.
    dict_bins_prs_keys = list(dict_bins_prs.keys())
    # The bin IDs, as keys of a dictionary, are already unique.
    # Remove the "-" (undentified bin label) from the list.
    if "-" in dict_bins_prs_keys:
        dict_bins_prs_keys.remove("-")
    # Sort the keys in ascending order.
    dict_bins_prs_keys_sorted = sorted(dict_bins_prs_keys, key=int)
    # The sorted dictionary.
    dict_bins_prs_sorted = {}
    for bin_id in dict_bins_prs_keys_sorted:
        for key_bin in dict_bins_prs.keys():
            if bin_id == key_bin:
                dict_bins_prs_sorted[key_bin] = copy.deepcopy(dict_bins_prs[key_bin])
    # Write the bin and their protein IDs in a file.
    new_file_bin_proteins = open(binphylo_path_prsids, "w")
    for binid in dict_bins_prs_sorted.keys():
        for item in dict_bins_prs_sorted[binid]:
            new_file_bin_proteins.write("{}\t{}\n".format(binid, item))
    new_file_bin_proteins.close()

    # Dictionary with the information of bins - proteins - taxonomies.
    bins_prs_tax_dict = {}
    # Dictinoary that holds the frequencies fo the taxonomies for each bin.
    bin_tax_dict = {}
    # For a bin...
    for key_bin in dict_bins_prs_sorted.keys():
        if key_bin == "-":
            continue
        # Add the bin the dictionary of bins and taxonomies.
        bins_prs_tax_dict[key_bin] = {}
        bin_tax_dict[key_bin] = {}
        # For a protein...
        for pr_acc in dict_bins_prs_sorted[key_bin]:
            # Check if protein in the proteins with information from nr.
            if pr_acc in dict_nr.keys():
                # Analyze the information from the pnr database for the protein.
                nr_info = dict_nr[pr_acc]
                # Add the protien in the dictionary.
                bins_prs_tax_dict[key_bin][pr_acc] = {}
                # Collect the ID of the hit in the pnr/nr database.
                nr_info_splited = nr_info.split("\t")
                pr_nr_acc = nr_info_splited[1]
                # Collect the taxonomy of the accession number of the pnr/nr database.
                if pr_nr_acc in dict_pnr_tax.keys():
                    # pr_nr_tax is a list of highest occuring species from each protein
                    pr_nr_tax = dict_pnr_tax[pr_nr_acc]
                    for tax_item in pr_nr_tax:
                        # Bins - Proteins - Taxonmies Dicitonary
                        if tax_item not in bins_prs_tax_dict[key_bin][pr_acc].keys():
                            bins_prs_tax_dict[key_bin][pr_acc][tax_item] = 1
                        else:
                            bins_prs_tax_dict[key_bin][pr_acc][tax_item] += 1
                        # Bins - Taxonmies Dicitonary
                        if tax_item not in bin_tax_dict[key_bin].keys():
                            bin_tax_dict[key_bin][tax_item] = 1
                        else:
                            bin_tax_dict[key_bin][tax_item] += 1

    # Protein IDs to taxa with a flag for direct or indirect determination.
    # Select the highest occuring species from the bin_tax_dict.
    bin_tax_max_dict = {}
    for binid in bin_tax_dict.keys():
        bin_tax_max_dict[binid] = {}
        # Find the maximum taxonomy frequency.
        tax_freq_max = 0
        for species_id in bin_tax_dict[binid].keys():
            tax_freq_cur = bin_tax_dict[binid][species_id]
            # If the frequency of the current taxnomy is higher than the currently max one, replace the existing taxonomies for the current bin.
            # If the frequency of the current taxonomy is equal to the currently max one, add the existing taxonmy to the current bin.
            # Otherwise, do nothing.
            if tax_freq_cur > tax_freq_max:
                bin_tax_max_dict[binid] = {}
                bin_tax_max_dict[binid][species_id] = tax_freq_cur
                tax_freq_max = tax_freq_cur
            elif tax_freq_cur == tax_freq_max:
                bin_tax_max_dict[binid][species_id] = tax_freq_cur

    # Dictionary to write the information in the annotation files.
    bin_group_tax_dict = {}
    if not bin_group_tax_dict:
        for bin_id in bin_tax_max_dict.keys():
            bin_group_tax_dict[bin_id] = {
                "max_freq": None,
                "protein_ids_d": [],
                "protein_ids_i": [],
                "species": []
            }
            for species_id in bin_tax_max_dict[bin_id].keys():
                if bin_group_tax_dict[bin_id]["max_freq"] is None:
                    max_freq = bin_tax_max_dict[bin_id][species_id]
                    bin_group_tax_dict[bin_id]["max_freq"] = max_freq
                bin_group_tax_dict[bin_id]["species"].append(species_id)
                # For each protein of the bin, if the protein is present in bins_prs_tax_dict[bin_id] then it has immediate taxonomy inference otherwise intemediate.
                bin_protens_with_tax = list(bins_prs_tax_dict[bin_id].keys())
                for protein_id in dict_prs_bins.keys():
                    if dict_prs_bins[protein_id] == bin_id:
                        if protein_id in bin_protens_with_tax:
                            bin_group_tax_dict[bin_id]["protein_ids_d"].append(protein_id)
                        else:
                            bin_group_tax_dict[bin_id]["protein_ids_i"].append(protein_id)

    # Binned proteins FASTA file.
    bin_pr_accs = set()
    for key_bin in dict_bins_prs_sorted.keys():
        for pr_ac in dict_bins_prs_sorted[key_bin]:
            bin_pr_accs.add(pr_ac)

    # Protein - Bin
    new_file_bin_phylo_prs = open(binphylo_path_prstax, "w")
    for key_1 in bins_prs_tax_dict.keys():
        for key_2 in bins_prs_tax_dict[key_1].keys():
            for key_3 in bins_prs_tax_dict[key_1][key_2].keys():
                new_file_bin_phylo_prs.write("{}\t{}\t{}\t{}\n".format(key_1, key_2, key_3, bins_prs_tax_dict[key_1][key_2][key_3]))
    new_file_bin_phylo_prs.close()
    # Bin - Taxon - Frequency
    new_file_bin_phylo_bins = open(binphylo_path_binstax, "w")
    for key_1 in bin_tax_dict.keys():
        for key_2 in bin_tax_dict[key_1].keys():
            new_file_bin_phylo_bins.write("{}\t{}\t{}\n".format(key_1, key_2, bin_tax_dict[key_1][key_2]))
    new_file_bin_phylo_bins.close()
    # Taxon name - Frequency
    new_file_name_freq = open(binphylo_path_binstax_names, "w")
    for key_1 in bin_tax_dict.keys():
        for key_2 in bin_tax_dict[key_1].keys():
            new_file_name_freq.write("{}\n".format(key_2))
    new_file_name_freq.close()
    # Bin - Taxon - Max frequency
    new_file_bin_phylo_bins_max = open(binphylo_path_binstax_max, "w")
    for key_1 in bin_tax_max_dict.keys():
        for key_2 in bin_tax_max_dict[key_1].keys():
            new_file_bin_phylo_bins_max.write("{}\t{}\t{}\n".format(key_1, key_2, bin_tax_max_dict[key_1][key_2]))
    new_file_bin_phylo_bins_max.close()
    # Taxon name - Max frequency
    new_file_name_maxfreq = open(binphylo_path_binstax_max_names, "w")
    for key_1 in bin_tax_max_dict.keys():
        for key_2 in bin_tax_max_dict[key_1].keys():
            new_file_name_maxfreq.write("{}\n".format(key_2))
    new_file_name_maxfreq.close()

    # Finding taxonomy IDs
    if os.path.exists(binphylo_path_binstax_names):
        # Command
        if os.path.exists(taxonkit_path):
            phrase_1 = "\"{}\" name2taxid \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(taxonkit_path, binphylo_path_binstax_names, binphylo_freq_taxids_path, thread_num, taxoknit_freq_stdoe_path)
            phrase_2 = "\"{}\" -h > \"{}\"".format(taxonkit_path, taxoknit_freq_version_path)
        else:
            phrase_1 = "taxonkit name2taxid \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(binphylo_path_binstax_names, binphylo_freq_taxids_path, thread_num, taxoknit_freq_stdoe_path)
            phrase_2 = "taxonkit -h > \"{}\"".format(taxoknit_freq_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(tax_freq_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if taxonkit_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(taxonkit_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if taxonkit_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(tax_freq_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(tax_freq_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)

    if os.path.exists(binphylo_path_binstax_max_names):
        # Command
        if os.path.exists(taxonkit_path):
            phrase_1 = "\"{}\" name2taxid \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(taxonkit_path, binphylo_path_binstax_max_names, binphylo_maxfreq_taxids_path, thread_num, taxoknit_maxfreq_stdoe_path)
            phrase_2 = "\"{}\" -h > \"{}\"".format(taxonkit_path, taxoknit_maxfreq_version_path)
        else:
            phrase_1 = "taxonkit name2taxid \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(binphylo_path_binstax_max_names, binphylo_maxfreq_taxids_path, thread_num, taxoknit_maxfreq_stdoe_path)
            phrase_2 = "taxonkit -h > \"{}\"".format(taxoknit_maxfreq_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(tax_maxfreq_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if taxonkit_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(taxonkit_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if taxonkit_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(tax_maxfreq_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(tax_maxfreq_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)

    # Collected the unique taxonomy IDs
    if os.path.exists(binphylo_freq_taxids_path):
        unique_taxids = []
        freq_name_lines = read_file(binphylo_freq_taxids_path)
        for line in freq_name_lines:
            line_splited = line.split("\t")
            taxid = line_splited[1]
            if taxid not in unique_taxids:
                unique_taxids.append(taxid)
        freq_taxids_file = open(freq_taxids_path, "w")
        for item in unique_taxids:
            freq_taxids_file.write("{}\n".format(item))
        freq_taxids_file.close()
    if os.path.exists(binphylo_maxfreq_taxids_path):
        unique_taxids = []
        freq_name_lines = read_file(binphylo_maxfreq_taxids_path)
        for line in freq_name_lines:
            line_splited = line.split("\t")
            taxid = line_splited[1]
            if taxid not in unique_taxids:
                unique_taxids.append(taxid)
        freq_taxids_file = open(maxfreq_taxids_path, "w")
        for item in unique_taxids:
            freq_taxids_file.write("{}\n".format(item))
        freq_taxids_file.close()

    # Find the lineage
    if os.path.exists(freq_taxids_path):
        # Command
        if os.path.exists(taxonkit_path):
            phrase_1 = "\"{}\" lineage \"{}\" -o \"{}\" -j {} &> \"{}\"".format(taxonkit_path, freq_taxids_path, freq_lineage_path, thread_num, taxonkit_freq_line_stdoe_path)
            phrase_2 = "\"{}\" -h > \"{}\"".format(taxonkit_path, taxonkit_freq_line_version_path)
            phrase_3 = "\"{}\" lineage \"{}\" -j {} | csvtk pretty -Ht -x ';' -W 70 -S bold -o \"{}.txt\" &> \"{}\"".format(taxonkit_path, freq_taxids_path, thread_num, freq_lineage_form_path, csvtk_freq_stdoe_path)
            phrase_4 = "csvtk -h > {}".format(csvtk_freq_version_path)
        else:
            phrase_1 = "taxonkit lineage \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(freq_taxids_path, freq_lineage_path, thread_num, taxonkit_freq_line_stdoe_path)
            phrase_2 = "taxonkit -h > \"{}\"".format(taxonkit_freq_line_version_path)
            phrase_3 = "taxonkit lineage \"{}\" -j {} | csvtk pretty -Ht -x ';' -W 70 -S bold -o \"{}\" &> \"{}\"".format(freq_taxids_path, thread_num, freq_lineage_form_path, csvtk_freq_stdoe_path)
            phrase_4 = "csvtk -h > \"{}\"".format(csvtk_freq_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(taxonkit_freq_line_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if taxonkit_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(taxonkit_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        new_file_bash.write("{}\n".format(phrase_3))
        new_file_bash.write("{}\n".format(phrase_4))
        if taxonkit_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(taxonkit_freq_line_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(taxonkit_freq_line_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)

    if os.path.exists(maxfreq_taxids_path):
        # Command
        if os.path.exists(taxonkit_path):
            phrase_1 = "\"{}\" lineage \"{}\" -o \"{}\" -j {} &> \"{}\"".format(taxonkit_path, maxfreq_taxids_path, maxfreq_lineage_path, thread_num, taxonkit_maxfreq_line_stdoe_path)
            phrase_2 = "\"{}\" -h > \"{}\"".format(taxonkit_path, taxonkit_maxfreq_line_version_path)
            phrase_3 = "\"{}\" lineage \"{}\" -j {} | csvtk pretty -Ht -x ';' -W 70 -S bold -o \"{}.txt\" &> \"{}\"".format(taxonkit_path, maxfreq_taxids_path, thread_num, maxfreq_lineage_form_path, csvtk_maxfreq_stdoe_path)
            phrase_4 = "csvtk -h > \"{}\"".format(csvtk_maxfreq_version_path)
        else:
            phrase_1 = "taxonkit lineage \"{}\" -r -o \"{}\" -j {} &> \"{}\"".format(maxfreq_taxids_path, maxfreq_lineage_path, thread_num, taxonkit_maxfreq_line_stdoe_path)
            phrase_2 = "taxonkit -h > \"{}\"".format(taxonkit_maxfreq_line_version_path)
            phrase_3 = "taxonkit lineage \"{}\" -j {} | csvtk pretty -Ht -x ';' -W 70 -S bold -o \"{}\" &> \"{}\"".format(maxfreq_taxids_path, thread_num, maxfreq_lineage_form_path, csvtk_maxfreq_stdoe_path)
            phrase_4 = "csvtk -h > \"{}\"".format(csvtk_maxfreq_version_path)
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(taxonkit_maxfreq_line_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if taxonkit_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(taxonkit_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        new_file_bash.write("{}\n".format(phrase_3))
        new_file_bash.write("{}\n".format(phrase_4))
        if taxonkit_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(taxonkit_maxfreq_line_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(taxonkit_maxfreq_line_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)

    # Determine the frequency of profile usage for each family in locating the respective domains at the binned proteins.
    # The binned protein IDs hace been collected in bin_pr_accs.
    # The results of hmmscan of phylo profiles against the protein from cd-hit are in the hmmer_enz_domains_all_proteins_phylo file.
    family_to_profile_phylo_sum_dict = {}
    if bin_pr_accs and os.path.exists(hmmer_enz_domains_all_proteins_phylo):
        with open(hmmer_enz_domains_all_proteins_phylo) as hmmer_phylo_lines:
            for line in hmmer_phylo_lines:
                line_splited = line.split("\t")
                source_pr_acc = line_splited[0]
                if source_pr_acc in bin_pr_accs:
                    profile_id = line_splited[4]
                    # Remove any suffix after the dot, if a dot is present.
                    if "." in profile_id:
                        profile_id_splited = profile_id.split(".")
                        profile_id = profile_id_splited[0]
                    for key_fam in family_to_profile_phylo_dict.keys():
                        if key_fam not in family_to_profile_phylo_sum_dict.keys():
                            family_to_profile_phylo_sum_dict[key_fam] = 0
                        if profile_id in family_to_profile_phylo_dict[key_fam].keys():
                            family_to_profile_phylo_dict[key_fam][profile_id] += 1
                            family_to_profile_phylo_sum_dict[key_fam] += 1
    family_profile_file = open(family_profile_path, "w")
    family_profile_file.write("Protein family\tProfile ID\tFrequency\n")
    for key_fam in family_to_profile_phylo_dict.keys():
        for key_prof in family_to_profile_phylo_dict[key_fam].keys():
            fam_prof_freq = family_to_profile_phylo_dict[key_fam][key_prof]
            family_profile_file.write("{}\t{}\t{}\n".format(key_fam, key_prof, fam_prof_freq))
    if family_to_profile_phylo_sum_dict:
        family_profile_file.write("\n")
        family_profile_file.write("Protein family\tSum of profile hits\n")
        for key_fam in family_to_profile_phylo_sum_dict.keys():
            fam_prof_sum = family_to_profile_phylo_sum_dict[key_fam]
            family_profile_file.write("{}\t{}\n".format(key_fam, fam_prof_sum))
    family_profile_file.close()

    return dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict


def phobius(phobius_env, proteins_combined_file, output_path_topology, phobius_input_file_name, phobius_output_file_name, topology_info_path, phobius_path, input_log_file, output_log_file, phobius_bash_script, phobius_version_path, phobius_stde_path, conda_sh_path):
    print("\nRunning Phobius...")
    if os.path.exists(output_path_topology):
        shutil.rmtree(output_path_topology)
    os.mkdir(output_path_topology)
    # If there is a path to Phobius and a file with combined proteins, then use Phobius for transmembrane topology
    # predictions.
    dict_top = {}
    new_file = open(topology_info_path, "w")
    if os.path.exists(proteins_combined_file):
        shutil.copy(proteins_combined_file, phobius_input_file_name)
        # Command
        if os.path.exists(phobius_path):
            # In this case, at first export the path to the phobius folder.
            phrase_1 = "export PATH=\"$PATH:{}\"".format(phobius_path)
            phrase_2 = "phobius.pl \"{}\" > \"{}\" 2> \"{}\"".format(phobius_input_file_name, phobius_output_file_name, phobius_stde_path)
            phrase_3 = "phobius.pl -h > \"{}\"".format(phobius_version_path)
        else:
            phrase_1 = "perl phobius.pl \"{}\" > \"{}\" 2> \"{}\"".format(phobius_input_file_name, phobius_output_file_name, phobius_stde_path)
            phrase_2 = "perl phobius.pl -h > \"{}\"".format(phobius_version_path)
            phrase_3 = ""
        # Create the Bash script.
        # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
        new_file_bash = open(phobius_bash_script, "w")
        phrase = "#!/bin/bash"
        new_file_bash.write("{}\n".format(phrase))
        if phobius_env:
            phrase_s = "source \"{}\"".format(conda_sh_path)
            phrase_a = "conda activate \"{}\"".format(phobius_env)
            new_file_bash.write("{}\n".format(phrase_s))
            new_file_bash.write("{}\n".format(phrase_a))
        new_file_bash.write("{}\n".format(phrase_1))
        new_file_bash.write("{}\n".format(phrase_2))
        if phrase_3:
            new_file_bash.write("{}\n".format(phrase_3))
        if phobius_env:
            phrase = "conda deactivate"
            new_file_bash.write("{}\n".format(phrase))
        new_file_bash.close()
        # Making the Bash script executable.
        # Sending command to run.
        phrase_b1 = "chmod +x {}".format(phobius_bash_script)
        phrase_b2 = "chmod --version"
        title_1 = "Making a Bash script executable:"
        title_2 = "Version of chmod:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Sending command to run.
        phrase_b1 = "bash {}".format(phobius_bash_script)
        phrase_b2 = "bash --version"
        title_1 = "Running bash:"
        title_2 = "Version of bash:"
        capture_status = True
        shell_status = True
        pr_status = False
        command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Gathering information from the results of Phobius
        pattern_1 = re.compile(r'^ID\s+(.*)')
        pattern_2 = re.compile(r'^FT\s+(\S+)\s+(\S+)\s+(\S+)\s+(.*)')
        pattern_3 = re.compile(r'^FT\s+(\S+)\s+(\S+)\s+(\S+)')
        with open(phobius_output_file_name) as phobius_output_lines:
            for line in phobius_output_lines:
                line = line.rstrip("\n")
                result_dom = pattern_1.search(line)
                if result_dom:
                    phobius_protein_id = result_dom.group(1)
                else:
                    region_name = None
                    result_2 = pattern_2.search(line)
                    result_3 = pattern_3.search(line)
                    if result_2:
                        region_name_1 = result_2.group(1)
                        left_bounder = result_2.group(2)
                        right_bounder = result_2.group(3)
                        region_name_2 = result_2.group(4)
                        if region_name_1:
                            region_name_1 = region_name_1.strip()
                        if region_name_2:
                            region_name_2 = region_name_2.strip()
                            if region_name_2[-1] == ".":
                                region_name_2 = region_name_2[:-1]
                            region_name = "{}||{}".format(region_name_1, region_name_2)
                        else:
                            region_name = region_name_1
                    elif result_3:
                            region_name = result_3.group(1)
                            left_bounder = result_3.group(2)
                            right_bounder = result_3.group(3)
                    if region_name is not None:
                        # Writing the information in the file.
                        new_file.write("{}\t{}\t{}\t{}\n".format(phobius_protein_id, left_bounder, right_bounder, region_name))
                        temp_phrase = "{}\t{}\t{}".format(left_bounder, right_bounder, region_name)
                        # Adding the information in the dictionary.
                        if phobius_protein_id not in dict_top.keys():
                            dict_top[phobius_protein_id] = []
                        dict_top[phobius_protein_id].append(temp_phrase)
    if not dict_top:
        new_file.write("No sequences with transmembrane or signal peptides found.\n")
    new_file.close()
    return dict_top


def input_motif_search(motifs_path, output_path_motifs, input_motifs_results, proteins_combined_file, input_log_file):
    print("\nSearching for the input motifs...")
    if os.path.exists(output_path_motifs):
        shutil.rmtree(output_path_motifs)
    os.mkdir(output_path_motifs)
    protein_acc = None
    dict_input_motifs = {}
    if motifs_path is not None:
        # If the file with the CD-HIT filtered proteins has a non-zero size.
        if os.path.getsize(proteins_combined_file) != 0:
            motif_lines = read_file(motifs_path)
            motif_sqs = []
            for line in motif_lines:
                # If the line is not empty and contains a motif then add it to the list.
                if line:
                    motif_sqs.append(line)
            print("\nInput motifs to be searched:")
            print(motif_sqs)
            if motif_sqs:
                motif_patterns = []
                for mof in motif_sqs:
                    mof_pat = "({})".format(mof)
                    pattern = re.compile(mof_pat)
                    motif_patterns.append(pattern)
                first_line = True
                new_file = open(input_motifs_results, "w")
                with open(proteins_combined_file) as lines_cd_hit:
                    for line in lines_cd_hit:
                        line = line.rstrip("\n")
                        if line[0] == ">":
                            if not first_line:
                                for pi in range(0, len(motif_patterns)):
                                    pattern = motif_patterns[pi]
                                    mof = motif_sqs[pi]
                                    result = re.finditer(pattern, protein_seq)
                                    if result:
                                        for item in result:
                                            # Numbering starts from 0 and the ending of the pattern found is one position
                                            # higher. So, for a numbering starting from 1, the start of the pattern is
                                            # one higher than the one given while the end remains the same.
                                            result_start = item.start() + 1
                                            result_end = item.end()
                                            temp_mof_list = [mof, result_start, result_end]
                                            if protein_acc not in dict_input_motifs.keys():
                                                dict_input_motifs[protein_acc] = [temp_mof_list]
                                            else:
                                                dict_input_motifs[protein_acc].append(temp_mof_list)
                                            new_file.write("{}\t{}\t{}\t{}\n".format(protein_acc, mof, result_start, result_end))
                            protein_acc = line[1:]
                            protein_seq = ""
                            first_line = False
                        else:
                            protein_seq += line
                # This is the same motif analysi as in the loop above but it is also set here because the last protein sequence must also be examined.
                # The whole sequence will be colleceted at the last iteration so that sequence should also be examined.
                for pi in range(0, len(motif_patterns)):
                    pattern = motif_patterns[pi]
                    mof = motif_sqs[pi]
                    result = re.finditer(pattern, protein_seq)
                    if result:
                        for item in result:
                            result_start = item.start() + 1
                            result_end = item.end()
                            temp_mof_list = [mof, result_start, result_end]
                            if protein_acc not in dict_input_motifs.keys():
                                dict_input_motifs[protein_acc] = [temp_mof_list]
                            else:
                                dict_input_motifs[protein_acc].append(temp_mof_list)
                            new_file.write("{}\t{}\t{}\t{}\n".format(protein_acc, mof, result_start, result_end))
                new_file.close()
            else:
                print("\nFile for motifs is empty. No input motifs are searched in the proteins.")
        else:
            print("\nNo proteins to be searched for motifs.")
            input_log_file.write("No proteins to be searched for motifs.\n")
    else:
        print("\nNo input motifs will be searched.\n")
        input_log_file.write("No input motifs were searched.\n")
    return dict_input_motifs


def pr_fam_pred(analysis_fam_names, blastp_info_swissprot_file, proteins_combined_file, pr_fams_path, family_info_path):
    print("\nPredicting the protein family of each protein based on its hit of lowest e-value against the Swiss-Prot protein database...")
    # Keeping only the protein with the lowest e-value from the hits against the Swiss-Prot database.
    dict_sp_protac = {}
    protein_acc = None
    min_evalue = None
    header = True
    with open(blastp_info_swissprot_file) as lines_sp:
        for line_sp in lines_sp:
            if header:
                header = False
                continue
            line_sp = line_sp.rstrip("\n")
            splited_sp = line_sp.split("\t")
            protein_acc = splited_sp[0]
            swissprot_prot = splited_sp[1]
            e_value = float(splited_sp[3])
            if "." in swissprot_prot:
                swissprot_prot_splited = swissprot_prot.split(".")
                swissprot_prot = swissprot_prot_splited[0]
            if protein_acc not in dict_sp_protac.keys():
                dict_sp_protac[protein_acc] = swissprot_prot
                min_evalue = e_value
            else:
                if e_value < min_evalue:
                    dict_sp_protac[protein_acc] = swissprot_prot
                    min_evalue = e_value
    # Protein sequences
    dict_seqs = {}
    if os.path.exists(proteins_combined_file):
        with open(proteins_combined_file) as lines_seqs:
            for line_pr_seq in lines_seqs:
                line_pr_seq = line_pr_seq.rstrip("\n")
                if line_pr_seq[0] == ">":
                    protein_acc = line_pr_seq[1:]
                    if protein_acc in dict_seqs.keys():
                        print("Error. A duplicate header was found. Exiting.")
                        exit()
                    else:
                        dict_seqs[protein_acc] = ""
                else:
                    dict_seqs[protein_acc] += line_pr_seq
    # Protein families from the Swiss-Prot/UniprotKB database.
    fams_lenth_prs_dict = {}
    fam_lines = read_file(pr_fams_path)
    header = True
    for line in fam_lines:
        if header:
            header = False
            continue
        else:
            line_splited = line.split("\t")
            pr_fam_name = line_splited[0]
            swissprot_proteins = line_splited[-1]
            swissprot_proteins = swissprot_proteins[2:-2]
            swissprot_proteins_splited = swissprot_proteins.split("', '")
            temp_list = line_splited[1:-1] + [swissprot_proteins_splited]
            fams_lenth_prs_dict[pr_fam_name] = temp_list
    # For each putative protein being alazyed (key):
    # Based on its best hit (swissprot_pr) against the Swiss-Prot database:
    # For each protein family (key_fam) collected from the Swiss-Prot database, take its proteins (fam_prs) and:
    # If the best hit protein (swissprot_pr) is part of fam_prs:
    # And the length of the protein family (key_fam) is not None:
    # Add the putative protein (key) in a dictionary as a key and the family of its best hit as it value in a list
    # Add the length information of the faimly of the best hit (predicted protein family)
    # In the end, if no protein families had been initially selected to seek add "-", otherwise
    # After checking if any of the selected proteins family to seek matches the predicted one, if at least one matches add "1", otherwise if none matches
    # add "0". The same procedure is applied for any protein family the swissprot_pr is part of.
    swiss_fams_len_comp_dict = {}
    for key in dict_sp_protac.keys():
        swissprot_pr = dict_sp_protac[key]
        pr_len = None
        if key in dict_seqs.keys():
            pr_len = float(len(dict_seqs[key]))
        for key_fam in fams_lenth_prs_dict.keys():
            if fams_lenth_prs_dict[key_fam][2]:
                fam_prs = fams_lenth_prs_dict[key_fam][2]
                if swissprot_pr in fam_prs:
                    if pr_len is not None:
                        if key not in swiss_fams_len_comp_dict.keys():
                            swiss_fams_len_comp_dict[key] = [[key_fam]]
                        else:
                            swiss_fams_len_comp_dict[key].append([key_fam])
                        fam_mean_len = float(fams_lenth_prs_dict[key_fam][0])
                        mean_len_dist = abs(pr_len - fam_mean_len)
                        mean_len_dist = round(mean_len_dist, 2)
                        relative_change = (mean_len_dist / fam_mean_len) * 100
                        relative_change = round(relative_change, 2)
                        swiss_fams_len_comp_dict[key][-1].append(fam_mean_len)
                        swiss_fams_len_comp_dict[key][-1].append(mean_len_dist)
                        swiss_fams_len_comp_dict[key][-1].append(relative_change)
                        if not analysis_fam_names:
                            swiss_fams_len_comp_dict[key][-1].append("-")
                        else:
                            swiss_fams_len_comp_dict[key][-1].append(0)
                            for sel_fam in analysis_fam_names:
                                if sel_fam == key_fam:
                                    swiss_fams_len_comp_dict[key][-1].append(1)
                                    break
    # Write the information in a file.
    new_file = open(family_info_path, "w")
    for key_comp in swiss_fams_len_comp_dict.keys():
        for set_info in swiss_fams_len_comp_dict[key_comp]:
            new_file.write("{}\t{}\t{}\t{}\t{}\t{}\n".format(key_comp, set_info[0], set_info[1], set_info[2], set_info[3], set_info[4]))
    new_file.close()
    return swiss_fams_len_comp_dict, dict_seqs


def bowtie(output_path_bowtie, paired_end, tr_ex_file_paths, tr_ex_file_paths_p, conda_sh_path, bowtie_build_path, bowtie_build_version_path, bowtie_build_stdoe_path, bowtie_env, bowtie_bash_script, bowtie_path, bowtie_version_path, bowtie_stdoe_path, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, mapped_reads_path, thread_num, output_final_contigs_formated, bowtie_contigs_basename, bowtie_status, input_log_file, output_log_file):
    print("\nRunning bowtie2...")
    reads_to_contigs_dict = {}
    contigs_to_reads_dict = {}
    if bowtie_status:
        if os.path.exists(output_path_bowtie):
            shutil.rmtree(output_path_bowtie)
        os.mkdir(output_path_bowtie)
    if os.path.exists(output_final_contigs_formated):
        if bowtie_status:
            # bowtie_single_unaligned_path: Unpaired reads (meaning single-end reads) that did not align.
            # bowtie_paired_unaligned_con_path: Paired reads that did not align concordantly.
            # where are the paired reads that did not align at all?
            if bowtie_build_path:
                phrase_1 = "\"{}\" --threads {} \"{}\" \"{}\" &> \"{}\"".format(bowtie_build_path, thread_num, output_final_contigs_formated, bowtie_contigs_basename, bowtie_build_stdoe_path)
                phrase_2 = "\"{}\" --version > \"{}\"".format(bowtie_build_path, bowtie_build_version_path)
            else:
                phrase_1 = "bowtie2-build --threads {} \"{}\" \"{}\" &> \"{}\"".format(thread_num, output_final_contigs_formated, bowtie_contigs_basename, bowtie_build_stdoe_path)
                phrase_2 = "bowtie2-build --version > \"{}\"".format(bowtie_build_version_path)
            if paired_end:
                paired_one_phrase = ""
                tr_ex_file_paths_keys = list(tr_ex_file_paths_p.keys())
                for key_file in range(0, len(tr_ex_file_paths_keys)):
                    key = tr_ex_file_paths_keys[key_file]
                    file_path_1 = tr_ex_file_paths_p[key][0]
                    if key_file + 1 == len(tr_ex_file_paths_keys):
                        paired_one_phrase = "{}{}".format(paired_one_phrase, file_path_1)
                    else:
                        paired_one_phrase = "{}{},".format(paired_one_phrase, file_path_1)
                paired_two_phrase = ""
                tr_ex_file_paths_keys = list(tr_ex_file_paths_p.keys())
                for key_file in range(0, len(tr_ex_file_paths_keys)):
                    key = tr_ex_file_paths_keys[key_file]
                    file_path_2 = tr_ex_file_paths_p[key][1]
                    if key_file + 1 == len(tr_ex_file_paths_keys):
                        paired_two_phrase = "{}{}".format(paired_two_phrase, file_path_2)
                    else:
                        paired_two_phrase = "{}{},".format(paired_two_phrase, file_path_2)
                if bowtie_path:
                    phrase_3 = "\"{}\" -x \"{}\" -1 \"{}\" -2 \"{}\" --very-sensitive --un \"{}\" --un-conc \"{}\" --met-file \"{}\" --omit-sec-seq -p {} -S \"{}\" &> \"{}\"".format(bowtie_path, bowtie_contigs_basename, paired_one_phrase, paired_two_phrase, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, thread_num, mapped_reads_path, bowtie_stdoe_path)
                    phrase_4 = "\"{}\" --version > \"{}\"".format(bowtie_path, bowtie_version_path)
                else:
                    phrase_3 = "bowtie2 -x \"{}\" -1 \"{}\" -2 \"{}\" --very-sensitive --un \"{}\" --un-conc \"{}\" --met-file \"{}\" --omit-sec-seq -p {} -S \"{}\" &> \"{}\"".format(bowtie_contigs_basename, paired_one_phrase, paired_two_phrase, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, thread_num, mapped_reads_path, bowtie_stdoe_path)
                    phrase_4 = "bowtie2 --version > \"{}\"".format(bowtie_version_path)
            else:
                single_phrase = ""
                for i in range(0, len(tr_ex_file_paths)):
                    file_path = tr_ex_file_paths[i]
                    if i + 1 == len(tr_ex_file_paths):
                        single_phrase = "{}{}".format(single_phrase, file_path)
                    else:
                        single_phrase = "{}{},".format(single_phrase, file_path)
                if bowtie_path:
                    phrase_3 = "\"{}\" -x \"{}\" -U \"{}\" --very-sensitive --un \"{}\" --un-conc \"{}\" --met-file \"{}\" --omit-sec-seq -S \"{}\" -p {} &> \"{}\"".format(bowtie_path, bowtie_contigs_basename, single_phrase, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, mapped_reads_path, thread_num, bowtie_stdoe_path)
                    phrase_4 = "\"{}\" --version > \"{}\"".format(bowtie_path, bowtie_version_path)
                else:
                    phrase_3 = "bowtie2 -x \"{}\" -U \"{}\" --very-sensitive --un \"{}\" --un-conc \"{}\" --met-file \"{}\" --omit-sec-seq -S \"{}\" -p {} &> \"{}\"".format(bowtie_path, bowtie_contigs_basename, single_phrase, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, mapped_reads_path, thread_num, bowtie_stdoe_path)
                    phrase_4 = "bowtie2 --version > \"{}\"".format(bowtie_version_path)
            # Create the Bash script.
            # Four cases: 1: Conda enviroment and path needed for the script. 2: Conda enviroment and no path needed for the script. 3: No conda enviroment and path needed for the script. 4: No conda enviroment and no path needed for the script.
            new_file_bash = open(bowtie_bash_script, "w")
            phrase = "#!/bin/bash"
            new_file_bash.write("{}\n".format(phrase))
            if bowtie_env:
                phrase_s = "source \"{}\"".format(conda_sh_path)
                phrase_a = "conda activate \"{}\"".format(bowtie_env)
                new_file_bash.write("{}\n".format(phrase_s))
                new_file_bash.write("{}\n".format(phrase_a))
            new_file_bash.write("{}\n".format(phrase_1))
            new_file_bash.write("{}\n".format(phrase_2))
            new_file_bash.write("{}\n".format(phrase_3))
            new_file_bash.write("{}\n".format(phrase_4))
            if bowtie_env:
                phrase = "conda deactivate"
                new_file_bash.write("{}\n".format(phrase))
            new_file_bash.close()
            # Making the Bash script executable.
            # Sending command to run.
            phrase_b1 = "chmod +x {}".format(bowtie_bash_script)
            phrase_b2 = "chmod --version"
            title_1 = "Making a Bash script executable:"
            title_2 = "Version of chmod:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
            # Sending command to run.
            phrase_b1 = "bash {}".format(bowtie_bash_script)
            phrase_b2 = "bash --version"
            title_1 = "Running bash:"
            title_2 = "Version of bash:"
            capture_status = True
            shell_status = True
            pr_status = False
            command_run(phrase_b1, phrase_b2, title_1, title_2, capture_status, shell_status, pr_status, input_log_file, output_log_file)
        # Find the aligned reads.
        reads_to_contigs_dict, contigs_to_reads_dict = sam_analysis.readmapal(mapped_reads_path)
    else:
        print("\nOne of the files needed for the Bowtie analysis is missing.")
    return reads_to_contigs_dict, contigs_to_reads_dict


def binning_analysis(mapped_reads_path, binning_results_path, file_read_info_path, reads_to_contigs_dict, contigs_to_reads_dict, output_final_contigs_formated, dict_contigs_bins, contig_read_summary_path, binphylo_path_binstax_max, binphylo_maxfreq_taxids_path, maxfreq_lineage_path, bin_summary_info_path, bin_group_tax_dict, binner_bin_info_path):
    print("\nAnalyzing mapped reads and binned contigs...")
    if (not reads_to_contigs_dict) or (not contigs_to_reads_dict):
        # Find the aligned reads.
        reads_to_contigs_dict, contigs_to_reads_dict = sam_analysis.readmapal(mapped_reads_path)

    dict_bins_contigs = {}
    if os.path.exists(binning_results_path) and os.path.exists(file_read_info_path) and reads_to_contigs_dict and contigs_to_reads_dict:
        # Get the contig IDs of all the contigs.
        # Contig status: 1 for binned, 0 for not binned
        contig_status_dict = {}
        pattern_contig_id = re.compile(r'^>(\S+)')
        with open(output_final_contigs_formated) as conitg_lines:
            for line in conitg_lines:
                line = line.rstrip()
                result_contig_id = pattern_contig_id.search(line)
                if result_contig_id:
                    contig_id = result_contig_id.group(1)
                    contig_status_dict[contig_id] = 0
        # Contigs to bins
        if not dict_contigs_bins:
            with open(binning_results_path) as und_lines:
                tsv_lines = csv.reader(und_lines, delimiter="\t", quotechar='"')
                for line in tsv_lines:
                    contig_id = line[0]
                    bin_id = line[1]
                    if bin_id[:5] == "group":
                        bin_id = bin_id[5:]
                    contig_status_dict[contig_id] = 1
                    dict_contigs_bins[contig_id] = bin_id
                    bin_id_str = str(bin_id)
                    if bin_id_str not in dict_bins_contigs.keys():
                        dict_bins_contigs[bin_id_str] = []
                    dict_bins_contigs[bin_id_str].append(contig_id)
                    # The condition below exists because we have contigs that belong to bins and contigs to which reads were mapped. There is a chance, that a read was mapped to a contig
                    # that was not binned or that a contig was binned and to which contig no read was mapped.
                    # Therefore, there is a chance that contig_id from binning_results_path (results of binning) does not exist in the contigs mapped to reads.
                    if contig_id in contigs_to_reads_dict.keys():
                        contigs_to_reads_dict[contig_id][0] = 1
        else:
            for key_contig in dict_contigs_bins.keys():
                contig_status_dict[key_contig] = 1
                if key_contig in contigs_to_reads_dict.keys():
                    contigs_to_reads_dict[key_contig][0] = 1
                bin_id = dict_contigs_bins[key_contig]
                bin_id_str = str(bin_id)
                if bin_id_str not in dict_bins_contigs.keys():
                    dict_bins_contigs[bin_id_str] = []
                dict_bins_contigs[bin_id_str].append(key_contig)
        # Reads to bins
        # The read IDs in bins_to_reads are unique because they come from as keys from the reads_to_contigs_dict.
        bins_to_reads = {}
        for key_read in reads_to_contigs_dict.keys():
            contig_id = reads_to_contigs_dict[key_read][0]
            if contig_id in dict_contigs_bins.keys():
                bin_id = dict_contigs_bins[contig_id]
                if bin_id not in bins_to_reads.keys():
                    bins_to_reads[bin_id] = []
                bins_to_reads[bin_id].append(key_read)
        # Bins to read numbers
        bins_to_readnums = {}
        for key_bin in bins_to_reads.keys():
            read_number = len(bins_to_reads[key_bin])
            bins_to_readnums[key_bin] = read_number
        # Get the read numbers from the file.
        input_total_reads = None
        input_paired_reads = None
        trimmed_total_reads = None
        trimmed_paired_reads = None
        readinfo_lines = read_file(file_read_info_path)
        line_index = 0
        for line in readinfo_lines:
            line_splited = line.split("\t")
            if line_index == 0:
                input_total_reads = int(line_splited[1])
            elif line_index == 1:
                input_paired_reads = int(line_splited[1])
            elif line_index == 2:
                trimmed_total_reads = int(line_splited[1])
            elif line_index == 3:
                trimmed_paired_reads = int(line_splited[1])
            line_index += 1
        # Comparing the bin numbers to the initial number of reads
        bin_to_readfreq_dict = {}
        for key_bin in bins_to_readnums.keys():
            bin_to_readfreq_dict[key_bin] = [None, None, None, None]
            bin_readnum = bins_to_readnums[key_bin]
            if input_paired_reads is not None:
                input_paired_readfreq = bin_readnum / input_paired_reads
                input_paired_readfreq = 100 * input_paired_readfreq
                input_paired_readfreq = round(input_paired_readfreq, 2)
                bin_to_readfreq_dict[key_bin][1] = input_paired_readfreq
            if trimmed_paired_reads is not None:
                trimmed_paired_readfreq = bin_readnum / trimmed_paired_reads
                trimmed_paired_readfreq = 100 * trimmed_paired_readfreq
                trimmed_paired_readfreq = round(trimmed_paired_readfreq, 2)
                bin_to_readfreq_dict[key_bin][3] = trimmed_paired_readfreq
        # Get the contigs that were not binned.
        binned_contigs_num = 0
        nonbinned_contigs_num = 0
        for contig_id in contig_status_dict.keys():
            if contig_status_dict[contig_id] == 1:
                binned_contigs_num += 1
            else:
                nonbinned_contigs_num += 1
        # Count the reads that belong to binned contigs.
        binned_reads_num = 0
        nonbinned_reads_num = 0
        for key_contig in contigs_to_reads_dict.keys():
            if contigs_to_reads_dict[key_contig][0] == 1:
                binned_reads_num += len(contigs_to_reads_dict[key_contig]) - 1
            else:
                nonbinned_reads_num += len(contigs_to_reads_dict[key_contig]) - 1
        # Input total reads:
        # 1) Reads filtered out by preprocessing.
        # 2) Reads not assembled to contigs.
        # 3) Reads assmbled to contigs.
        # Reads from contigs:
        # 1) Reads of binned contigs.
        # 2) Reads of non-binned contigs.
        # Reads from binned contigs:
        # 1) Reads from binned contigs with organism(s).
        # 2) Reads from binned contigs without organism(s).
        processed_paired_reads = input_paired_reads - trimmed_paired_reads
        assembled_paired_reads = len(list(reads_to_contigs_dict.keys()))
        non_assembled_paired_reads = trimmed_paired_reads - assembled_paired_reads
        print("\nTotal contigs: {}".format(len(list(contig_status_dict.keys()))))
        print("Binned contigs: {}".format(binned_contigs_num))
        print("Non-binned contigs: {}".format(nonbinned_contigs_num))
        print("\nTotal reads: {}".format(input_paired_reads))
        print("Filtered reads: {}".format(processed_paired_reads))
        print("Contig reads: {}".format(assembled_paired_reads))
        print("Non-contig reads: {}".format(non_assembled_paired_reads))
        print("Binned reads: {}".format(binned_reads_num))
        print("Non-binned reads, from the contig reads: {}".format(nonbinned_reads_num))
        # Store the general information for the contigs and the reads.
        contig_read_summary_file = open(contig_read_summary_path, "w")
        contig_read_summary_file.write("Type\tCount\n")
        contig_read_summary_file.write("Total contigs\t{}\n".format(len(list(contig_status_dict.keys()))))
        contig_read_summary_file.write("Binned contigs\t{}\n".format(binned_contigs_num))
        contig_read_summary_file.write("Non-binned contigs\t{}\n".format(nonbinned_contigs_num))
        contig_read_summary_file.write("Total reads\t{}\n".format(input_paired_reads))
        contig_read_summary_file.write("Filtered reads\t{}\n".format(processed_paired_reads))
        contig_read_summary_file.write("Contig reads\t{}\n".format(assembled_paired_reads))
        contig_read_summary_file.write("Non-contig reads\t{}\n".format(non_assembled_paired_reads))
        contig_read_summary_file.write("Binned reads\t{}\n".format(binned_reads_num))
        contig_read_summary_file.write("Non-binned reads, from the contig reads\t{}\n".format(nonbinned_reads_num))
        contig_read_summary_file.close()
        bins_to_summary_dict = {}
        if os.path.exists(binphylo_path_binstax_max) and os.path.exists(binphylo_maxfreq_taxids_path) and os.path.exists(maxfreq_lineage_path):
            # 1    Intestinimonas butyriciproducens    3
            lines_1 = read_file(binphylo_path_binstax_max)
            for line in lines_1:
                line_splited = line.split("\t")
                bin = line_splited[0]
                taxname_1 = line_splited[1]
                if bin not in bins_to_summary_dict.keys():
                    bins_to_summary_dict[bin] = {}
                bins_to_summary_dict[bin][taxname_1] = []
            # Intestinimonas butyriciproducens    1297617    species
            max_taxids_lines = read_file(binphylo_maxfreq_taxids_path)
            for line in max_taxids_lines:
                line_splited = line.split("\t")
                taxname_2 = line_splited[0]
                taxid_1 = line_splited[1]
                taxrank = line_splited[2]
                for key_bin in bins_to_summary_dict.keys():
                    if taxname_2 in bins_to_summary_dict[key_bin].keys():
                        bins_to_summary_dict[key_bin][taxname_2] = [taxid_1, taxrank]
            # 1297617    cellular organisms;Bacteria;Terrabacteria group;Bacillota;Clostridia;Eubacteriales;Eubacteriales incertae sedis;Intestinimonas;Intestinimonas butyriciproducens    species
            max_lineages_lines = read_file(maxfreq_lineage_path)
            for line in max_lineages_lines:
                line_splited = line.split("\t")
                taxid_2 = line_splited[0]
                lineage = line_splited[1]
                for key_bin in bins_to_summary_dict.keys():
                    for name_bin in bins_to_summary_dict[key_bin].keys():
                        if bins_to_summary_dict[key_bin][name_bin][0] == taxid_2:
                            bins_to_summary_dict[key_bin][name_bin].append(lineage)
            for key_bin in bin_to_readfreq_dict.keys():
                # There is a chance that no species were found for a bin.
                if key_bin in bins_to_summary_dict.keys():
                    for taxname in bins_to_summary_dict[key_bin].keys():
                        readfreq_1 = bin_to_readfreq_dict[key_bin][1]
                        readfreq_3 = bin_to_readfreq_dict[key_bin][3]
                        bins_to_summary_dict[key_bin][taxname].append(readfreq_1)
                        bins_to_summary_dict[key_bin][taxname].append(readfreq_3)
        # Write information.
        bin_summary_info_file = open(bin_summary_info_path, "w")
        for key_bin in bins_to_summary_dict.keys():
            for key_name in bins_to_summary_dict[key_bin].keys():
                bin_summary_info_file.write("{}\t{}".format(key_bin, key_name))
                for item in bins_to_summary_dict[key_bin][key_name]:
                    bin_summary_info_file.write("\t{}".format(item))
                bin_summary_info_file.write("\n")
        bin_summary_info_file.close()
        # Write all the information realted to binning in one file.
        # Bin ID -> Protein IDs -> Contig IDs -> Read IDs
        # There is no need to check in the contig of contigs_to_reads_dict whether the read/contig belong to a bin or not. Each contig parsed from dict_bins_contigs is part of a bin, thus each
        # read of those contigs is also part of the bin.
        bin_sum_file = open(binner_bin_info_path, "w")
        for key_bin in bin_group_tax_dict.keys():
            max_freq = bin_group_tax_dict[key_bin]["max_freq"]
            if "protein_ids_d" in bin_group_tax_dict[key_bin].keys():
                protein_list = bin_group_tax_dict[key_bin]["protein_ids_d"]
            elif "protein_ids_i" in bin_group_tax_dict[key_bin].keys():
                protein_list = bin_group_tax_dict[key_bin]["protein_ids_i"]
            else:
                protein_list = []
            species_list = bin_group_tax_dict[key_bin]["species"]
            bin_sum_file.write("Bin ID:\n")
            bin_sum_file.write("{}\n".format(key_bin))
            bin_sum_file.write("Max read frequency (to any of the binned contigs):\n")
            bin_sum_file.write("{}\n".format(max_freq))
            bin_sum_file.write("Species:\n")
            for sp_id in species_list:
                bin_sum_file.write("{}\n".format(sp_id))
            bin_sum_file.write("Binned protein IDs:\n")
            for pr_id in protein_list:
                bin_sum_file.write("{}\n".format(pr_id))
            bin_sum_file.write("Binned contig IDs:\n")
            for cg_id in dict_bins_contigs[key_bin]:
                bin_sum_file.write("{}\n".format(cg_id))
            bin_sum_file.write("Binned read IDs:\n")
            for cg_id in dict_bins_contigs[key_bin]:
                if cg_id in contigs_to_reads_dict.keys():
                    for rd_id in contigs_to_reads_dict[cg_id][1:]:
                        bin_sum_file.write("{}\n".format(rd_id))
            bin_sum_file.write("//\n")
        bin_sum_file.close()
    else:
        print("\nOne of the files needed for the binning analysis is missing.")


def kraken_binning(kraken_mode, contigs_to_reads_dict, read_to_species_dict, taxid_to_species_dict, kraken_taxname_path, kraken_reads_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prsids, binned_ctb_path, binned_btc_path, binned_taxa_path, kraken_bin_info_path):
    print("\nPerforming binning based on the results of kraken2...")
    dict_contigs_bins = {}
    dict_bins_contigs = {}
    bin_group_tax_dict = {}
    if not taxid_to_species_dict:
        with open(kraken_taxname_path) as taxname_lines:
            for line in taxname_lines:
                line = line.rstrip("\n")
                line_splited = line.split("\t")
                taxid = line_splited[0]
                taxname = line_splited[1]
                taxid_to_species_dict[taxid] = taxname
    if kraken_mode:
        # For each contig:
        # Collect all its aligned reads. Get the species assigned to each read. Count each species
        # for the contig based on its reads. Keep the species of the highest count for each contig.
        # Based on the contigs that were assigned one species, group them based on their species of the highest counts.
        # Each group is a bin. Each bin has one species. The species of the bin is assigned to all genes of the bin.
        if not read_to_species_dict:
            with open(kraken_reads_path) as readtax_lines:
                for line in readtax_lines:
                    line = line.rstrip("\n")
                    line_splited = line.split("\t")
                    read_id = line_splited[0]
                    taxonid = line_splited[1]
                    thr_state = line_splited[2]
                    read_to_species_dict[read_id] = [taxonid, thr_state]
        # Contig -> Read -> Species -> Thr state
        contig_to_species = {}
        for key_contig in contigs_to_reads_dict.keys():
            if key_contig not in contig_to_species.keys():
                contig_to_species[key_contig] = {}
            if key_contig in contigs_to_reads_dict.keys():
                for read_id in contigs_to_reads_dict[key_contig][1:]:
                    if read_id in read_to_species_dict.keys():
                        read_species = read_to_species_dict[read_id][0]
                        thr_state = int(read_to_species_dict[read_id][1])
                        if thr_state == 1:
                            if read_species not in contig_to_species[key_contig].keys():
                                contig_to_species[key_contig][read_species] = 1
                            else:
                                contig_to_species[key_contig][read_species] += 1
        # Filter the species of maximum count for each contig.
        contig_to_species_max = {}
        for key_contig in contig_to_species.keys():
            max_count = 0
            for key_species in contig_to_species[key_contig].keys():
                species_count = contig_to_species[key_contig][key_species]
                if species_count > max_count:
                    max_count = species_count
            for key_species in contig_to_species[key_contig].keys():
                species_count = contig_to_species[key_contig][key_species]
                if max_count == species_count:
                    if key_contig not in contig_to_species_max:
                        contig_to_species_max[key_contig] = {}
                    contig_to_species_max[key_contig][key_species] = max_count
        # Filter the contigs with only one species of maximum count.
        contig_to_species_max_single = {}
        for key_contig in contig_to_species_max.keys():
            species_list = list(contig_to_species_max[key_contig].keys())
            species_num = len(species_list)
            if species_num == 1:
                contig_to_species_max_single[key_contig] = {}
                for key_species in contig_to_species_max[key_contig].keys():
                    count = contig_to_species_max[key_contig][key_species]
                    contig_to_species_max_single[key_contig][key_species] = count
        # Bin the contigs.
        bin_id_init = 0
        species_to_binids_dict = {}
        for key_contig in contig_to_species_max_single.keys():
            for key_species in contig_to_species_max_single[key_contig].keys():
                if key_species not in species_to_binids_dict.keys():
                    bin_id_init += 1
                    bin_id = bin_id_init
                    species_to_binids_dict[key_species] = bin_id
                    bin_id_str = str(bin_id)
                    dict_bins_contigs[bin_id_str] = [key_contig]
                else:
                    bin_id = species_to_binids_dict[key_species]
                    bin_id_str = str(bin_id)
                    dict_bins_contigs[bin_id_str].append(key_contig)
                dict_contigs_bins[key_contig] = bin_id_str
        # Binning results
        dict_prs_bins = {}
        dict_bins_prs = {}
        dict_prs_contigs = {}
        if os.path.exists(output_path_genepred) and os.path.exists(output_path_genepred_faa):
            protein_lines = read_file(output_path_genepred_faa)
            for line in protein_lines:
                if line[0] == ">":
                    protein_id = line[1:]
                    if "_" in protein_id:
                        protein_id_splited = protein_id.split("_")
                        origin_contig = "_".join(protein_id_splited[0:2])
                        dict_prs_contigs[protein_id] = origin_contig
        if dict_contigs_bins and dict_prs_contigs:
            for pr_id in dict_prs_contigs.keys():
                contig_id = dict_prs_contigs[pr_id]
                if contig_id in dict_contigs_bins.keys():
                    bin_id = dict_contigs_bins[contig_id]
                else:
                    bin_id = "-"
                dict_prs_bins[pr_id] = bin_id
                if bin_id not in dict_bins_prs.keys():
                    dict_bins_prs[bin_id] = [pr_id]
                else:
                    dict_bins_prs[bin_id].append(pr_id)
        # Create a list of the bin IDs.
        dict_bins_prs_keys = list(dict_bins_prs.keys())
        if "-" in dict_bins_prs_keys:
            dict_bins_prs_keys.remove("-")
        dict_bins_prs_keys_sorted = sorted(dict_bins_prs_keys, key=int)
        # The sorted dictionary.
        dict_bins_prs_sorted = {}
        for bin_id in dict_bins_prs_keys_sorted:
            for key_bin in dict_bins_prs.keys():
                if bin_id == key_bin:
                    dict_bins_prs_sorted[key_bin] = copy.deepcopy(dict_bins_prs[key_bin])
        # Store proteins based on their bin IDs.
        prid_to_bin_dict = {}
        for protein_id, bin_id in dict_prs_bins.items():
            if bin_id not in prid_to_bin_dict:
                prid_to_bin_dict[bin_id] = [protein_id]
            else:
                prid_to_bin_dict[bin_id].append(protein_id)
        # In contrast with MetaBinner and COMEBin the taxonomy of the proteins in the bins created based on kraken2 are inferred all by the same means.
        # Dictionary to write the information in the annotation files.
        # In this case max_freq is the highest number of reads towards the primary species (the species of the bin) of any of the contigs of the bin.
        bin_group_tax_dict = {}
        for key_contig in contig_to_species_max_single.keys():
            bin_id = dict_contigs_bins[key_contig]
            if bin_id not in bin_group_tax_dict.keys():
                bin_group_tax_dict[bin_id] = {
                    "max_freq": None,
                    "protein_ids_k": [],
                    "species": []
                }
            for species_id in contig_to_species_max_single[key_contig].keys():
                if bin_group_tax_dict[bin_id]["max_freq"] is None:
                    max_freq = contig_to_species_max_single[key_contig][species_id]
                    bin_group_tax_dict[bin_id]["max_freq"] = max_freq
                species_name = taxid_to_species_dict[species_id]
                species_comb = "{}||{}".format(species_name, species_id)
                if species_comb not in bin_group_tax_dict[bin_id]["species"]:
                    bin_group_tax_dict[bin_id]["species"].append(species_comb)
                if bin_id in prid_to_bin_dict.keys():
                    bin_group_tax_dict[bin_id]["protein_ids_k"] = prid_to_bin_dict[bin_id]
        print("\nWriting information related to kraken2 binning.")
        # Write the binned contigs according to conigs
        binned_ctb_file = open(binned_ctb_path, "w")
        for key_contig in dict_contigs_bins.keys():
            bin_id = dict_contigs_bins[key_contig]
            binned_ctb_file.write("{}\t{}\n".format(key_contig, bin_id))
        binned_ctb_file.close()
        # Write the binned contigs according to bins
        bin_ids_list_dupl = list(dict_contigs_bins.values())
        bin_ids_list = []
        for item in bin_ids_list_dupl:
            if item not in bin_ids_list:
                bin_ids_list.append(item)
        bin_ids_list = sorted(bin_ids_list, key=int)
        binned_btc_file = open(binned_btc_path, "w")
        for bin_id in bin_ids_list:
            for key_contig in dict_contigs_bins.keys():
                if bin_id == dict_contigs_bins[key_contig]:
                    binned_btc_file.write("{}\t{}\n".format(bin_id, key_contig))
        binned_btc_file.close()
        # Write the binned contigs
        # Write this file an use it in info collection properly.
        binned_taxa_file = open(binned_taxa_path, "w")
        for key_bin in bin_group_tax_dict.keys():
            max_freq = bin_group_tax_dict[key_bin]["max_freq"]
            protein_list = bin_group_tax_dict[key_bin]["protein_ids_k"]
            species_list = bin_group_tax_dict[key_bin]["species"]
            binned_taxa_file.write("bin:\n")
            binned_taxa_file.write("{}\n".format(key_bin))
            binned_taxa_file.write("max frequency:\n")
            binned_taxa_file.write("{}\n".format(max_freq))
            binned_taxa_file.write("binned proteins:\n")
            for item in protein_list:
                binned_taxa_file.write("{}\n".format(item))
            binned_taxa_file.write("species:\n")
            for item in species_list:
                binned_taxa_file.write("{}\n".format(item))
            binned_taxa_file.write("//\n")
        binned_taxa_file.close()
        # Write the bin and their protein IDs in a file.
        new_file_bin_proteins = open(binphylo_path_prsids, "w")
        for binid in dict_bins_prs_sorted.keys():
            for item in dict_bins_prs_sorted[binid]:
                new_file_bin_proteins.write("{}\t{}\n".format(binid, item))
        new_file_bin_proteins.close()
        # Write all the information realted to binning in one file.
        # Bin ID -> Protein IDs -> Contig IDs -> Read IDs
        # There is no need to check in the contig of contigs_to_reads_dict whether the read/contig belong to a bin or not. Each contig parsed from dict_bins_contigs is part of a bin, thus each
        # read of those contigs is also part of the bin.
        bin_sum_file = open(kraken_bin_info_path, "w")
        for key_bin in bin_group_tax_dict.keys():
            max_freq = bin_group_tax_dict[key_bin]["max_freq"]
            protein_list = bin_group_tax_dict[key_bin]["protein_ids_k"]
            species_list = bin_group_tax_dict[key_bin]["species"]
            bin_sum_file.write("Bin ID:\n")
            bin_sum_file.write("{}\n".format(key_bin))
            bin_sum_file.write("Max read frequency (to any of the binned contigs):\n")
            bin_sum_file.write("{}\n".format(max_freq))
            bin_sum_file.write("Species:\n")
            for sp_id in species_list:
                bin_sum_file.write("{}\n".format(sp_id))
            bin_sum_file.write("Binned protein IDs:\n")
            for pr_id in protein_list:
                bin_sum_file.write("{}\n".format(pr_id))
            bin_sum_file.write("Binned contig IDs:\n")
            for cg_id in dict_bins_contigs[key_bin]:
                bin_sum_file.write("{}\n".format(cg_id))
            bin_sum_file.write("Binned read IDs:\n")
            for cg_id in dict_bins_contigs[key_bin]:
                for rd_id in contigs_to_reads_dict[cg_id][1:]:
                    bin_sum_file.write("{}\n".format(rd_id))
            bin_sum_file.write("//\n")
        bin_sum_file.close()
    return bin_group_tax_dict


def info_collection(comb_all_domains_proteins, proteins_combined_file, blastp_info_swissprot_file, blastp_info_comb_nr_file, topology_info_path, gene_info_file, gene_contig_dist_path, input_motifs_results, family_info_path, binning_results_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prstax, binphylo_path_binstax, binphylo_path_binstax_max, binphylo_path_prsids, dict_hm, dict_sp, dict_nr, dict_top, swiss_fams_len_comp_dict, dict_input_motifs, dict_genes, contig_gene_dist_dict, dict_contigs_bins, dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict, kraken_mode, binned_taxa_path, dict_seqs, only_seek_mode, only_phylo_mode, cd_hit_results_fasta_path, add_seek_info, add_taxonomy_info):
    print("\nCollecting the information associated with the annotation of the proteins...")
    # Binned proteins
    if not dict_bins_prs_sorted:
        if os.path.exists(binphylo_path_prsids):
            bin_phylo_bins_lines = read_file(binphylo_path_prsids)
            for line in bin_phylo_bins_lines:
                line_splited = line.split("\t")
                bin_id = line_splited[0]
                protein_id = line_splited[1]
                if bin_id not in dict_bins_prs_sorted.keys():
                    dict_bins_prs_sorted[bin_id] = [protein_id]
                else:
                    dict_bins_prs_sorted[bin_id].append(protein_id)
    # Find the binned proteins.
    binned_protein_ids = set()
    if dict_bins_prs_sorted:
        for key_bin in dict_bins_prs_sorted.keys():
            binned_prs = dict_bins_prs_sorted[key_bin]
            for pr_item in binned_prs:
                binned_protein_ids.add(pr_item)
    # Sequences
    # 1. Seek mode: proteins_combined_file
    # 2. Taxonomy mode: Kraken2: Proteins from the bins.
    # 3. Taxonomy mode: MetaBinner / COMEBin: Proteins from the bins.
    dict_seqs = {}
    if not only_phylo_mode:
        if add_seek_info:
            if os.path.exists(proteins_combined_file):
                with open(proteins_combined_file) as lines_seqs_cf:
                    for line_pr_seq in lines_seqs_cf:
                        line_pr_seq = line_pr_seq.rstrip("\n")
                        if line_pr_seq[0] == ">":
                            protein_acc = line_pr_seq[1:]
                            dict_seqs[protein_acc] = ""
                        else:
                            dict_seqs[protein_acc] += line_pr_seq
    if not only_seek_mode:
        if add_taxonomy_info:
            found_binned_pr = False
            if os.path.exists(cd_hit_results_fasta_path):
                with open(cd_hit_results_fasta_path) as lines_seqs_ch:
                    for line_pr_seq in lines_seqs_ch:
                        line_pr_seq = line_pr_seq.rstrip("\n")
                        if line_pr_seq[0] == ">":
                            found_binned_pr = False
                            protein_acc = line_pr_seq[1:]
                            if protein_acc in binned_protein_ids:
                                found_binned_pr = True
                            if found_binned_pr:
                                dict_seqs[protein_acc] = ""
                        else:
                            if found_binned_pr:
                                dict_seqs[protein_acc] += line_pr_seq
    annotated_proteins_num = len(list(dict_seqs.keys()))
    print("\nNumber of proteins in the results: {}".format(annotated_proteins_num))
    # Domains
    if not dict_hm:
        if os.path.exists(comb_all_domains_proteins):
            dict_hm = {}
            header = True
            with open(comb_all_domains_proteins) as dom_lines:
                for line in dom_lines:
                    line = line.rstrip("\n")
                    line_splited = line.split("\t")
                    pr_name = line_splited[0]
                    if not header:
                        if pr_name not in dict_hm.keys():
                            dict_hm[pr_name] = [line]
                        else:
                            dict_hm[pr_name].append(line)
                    header = False
    # Phobius
    if not dict_top:
        if os.path.exists(topology_info_path):
            dict_top = {}
            with open(topology_info_path) as phobius_lines:
                for line in phobius_lines:
                    line = line.rstrip("\n")
                    # k141_12880_94839_96755_+\t1\t41\tSIGNAL
                    # k141_12880_94839_96755_+\t906\t926\tDOMAIN||C-REGION
                    # k141_12880_94839_96755_+\t906\t926\tDOMAIN||NON CYTOPLASMIC
                    # k127_2204_1502_2944_-	250	268	TRANSMEM
                    if "\t" in line:
                        line_splited = line.split("\t")
                        pr_id = line_splited[0]
                        temp_phrase = "\t".join(line_splited[1:])
                        if pr_id not in dict_top.keys():
                            dict_top[pr_id] = []
                        dict_top[pr_id].append(temp_phrase)
    # Swiss-Prot. Keeping only the protein with the lowest e-value from the hits against the Swiss-Prot database.
    if not dict_sp:
        if os.path.exists(blastp_info_swissprot_file):
            dict_sp = {}
            protein_acc = None
            min_evalue = None
            header = True
            with open(blastp_info_swissprot_file) as lines_sp:
                for line_sp in lines_sp:
                    if header:
                        header = False
                        continue
                    line_sp = line_sp.rstrip("\n")
                    splited_sp = line_sp.split("\t")
                    protein_acc = splited_sp[0]
                    e_value = float(splited_sp[3])
                    if protein_acc not in dict_sp.keys():
                        dict_sp[protein_acc] = line_sp
                        min_evalue = e_value
                    else:
                        if e_value < min_evalue:
                            dict_sp[protein_acc] = line_sp
                            min_evalue = e_value
    # NR. Keeping only the protein with the lowest e-calue from the hits against the nr database.
    if not dict_nr:
        if os.path.exists(blastp_info_comb_nr_file):
            dict_nr = {}
            protein_acc = None
            min_evalue = None
            header = True
            with open(blastp_info_comb_nr_file) as lines_nr:
                for line_nr in lines_nr:
                    if header:
                        header = False
                        continue
                    line_nr = line_nr.rstrip("\n")
                    splited_nr = line_nr.split("\t")
                    protein_acc = splited_nr[0]
                    e_value = float(splited_nr[3])
                    if protein_acc not in dict_nr.keys():
                        dict_nr[protein_acc] = line_nr
                        min_evalue = e_value
                    else:
                        if e_value < min_evalue:
                            dict_nr[protein_acc] = line_nr
                            min_evalue = e_value
    # Genes
    if not dict_genes:
        if os.path.exists(gene_info_file):
            dict_genes = {}
            gene_lines = read_file(gene_info_file)
            for line in gene_lines:
                line_splited = line.split("\t")
                dict_genes[line_splited[0]] = [line_splited[1], line_splited[2], line_splited[3]]
    # Gene distances from contig's edges.
    if not contig_gene_dist_dict:
        if os.path.exists(gene_contig_dist_path):
            contig_gene_dist_dict = {}
            gene_dist_lines = read_file(gene_contig_dist_path)
            for line in gene_dist_lines:
                line_splited = line.split("\t")
                protein_name = line_splited[0]
                start_dist = line_splited[1]
                end_dist = line_splited[2]
                contig_gene_dist_dict[protein_name] = [start_dist, end_dist]
    # Input motifs
    if not dict_input_motifs:
        dict_input_motifs = {}
        if os.path.exists(input_motifs_results):
            if os.path.getsize(input_motifs_results) != 0:
                motifs_lines = read_file(input_motifs_results)
                for line in motifs_lines:
                    line_splited = line.split("\t")
                    protein_acc = line_splited[0]
                    mof = line_splited[1]
                    result_start = line_splited[2]
                    result_end = line_splited[3]
                    temp_mof_list = [mof, result_start, result_end]
                    if protein_acc not in dict_input_motifs.keys():
                        dict_input_motifs[protein_acc] = [temp_mof_list]
                    else:
                        dict_input_motifs[protein_acc].append(temp_mof_list)
    # SwissProt predicted family and length difference.
    if not swiss_fams_len_comp_dict:
        if os.path.exists(family_info_path):
            swiss_fams_len_comp_dict = {}
            swiss_fam_lines = read_file(family_info_path)
            for line in swiss_fam_lines:
                line_splited = line.split("\t")
                protein_name = line_splited[0]
                pf_name = line_splited[1]
                pf_mean_length = line_splited[2]
                len_diff = line_splited[3]
                len_rel_change = line_splited[4]
                fam_dif = line_splited[5]
                temp_list = [pf_name, pf_mean_length, len_diff, len_rel_change, fam_dif]
                if pf_name not in swiss_fams_len_comp_dict.keys():
                    swiss_fams_len_comp_dict[protein_name] = [temp_list]
                else:
                    swiss_fams_len_comp_dict[protein_name].append(temp_list)
    # Binning results
    if (not dict_contigs_bins) or (not dict_prs_bins):
        if os.path.exists(binning_results_path):
            with open(binning_results_path) as und_lines:
                tsv_lines = csv.reader(und_lines, delimiter="\t", quotechar='"')
                for line in tsv_lines:
                    contig_id = line[0]
                    bin_id = line[1]
                    if bin_id[:5] == "group":
                        bin_id = bin_id[5:]
                    dict_contigs_bins[contig_id] = bin_id
        dict_prs_contigs = {}
        if os.path.exists(output_path_genepred) and os.path.exists(output_path_genepred_faa):
            protein_lines = read_file(output_path_genepred_faa)
            for line in protein_lines:
                if line[0] == ">":
                    protein_id = line[1:]
                    if "_" in protein_id:
                        protein_id_splited = protein_id.split("_")
                        origin_contig = "_".join(protein_id_splited[0:2])
                        dict_prs_contigs[protein_id] = origin_contig
        if dict_contigs_bins and dict_prs_contigs:
            for pr_id in dict_prs_contigs.keys():
                contig_id = dict_prs_contigs[pr_id]
                if contig_id in dict_contigs_bins.keys():
                    bin_id = dict_contigs_bins[contig_id]
                else:
                    bin_id = "-"
                dict_prs_bins[pr_id] = bin_id
    # Bin taxonomies
    if not bins_prs_tax_dict:
        if os.path.exists(binphylo_path_prstax):
            bin_phylo_prs_lines = read_file(binphylo_path_prstax)
            for line in bin_phylo_prs_lines:
                line_splited = line.split("\t")
                bin_id = line_splited[0]
                protein_id = line_splited[1]
                species_id = line_splited[2]
                species_num = int(line_splited[3])
                if bin_id not in bins_prs_tax_dict.keys():
                    bins_prs_tax_dict[bin_id] = {}
                if protein_id not in bins_prs_tax_dict[bin_id].keys():
                    bins_prs_tax_dict[bin_id][protein_id] = {}
                if species_id not in bins_prs_tax_dict[bin_id][protein_id].keys():
                    bins_prs_tax_dict[bin_id][protein_id][species_id] = species_num
    if not bin_tax_dict:
        if os.path.exists(binphylo_path_binstax):
            bin_phylo_bins_lines = read_file(binphylo_path_binstax)
            for line in bin_phylo_bins_lines:
                line_splited = line.split("\t")
                bin_id = line_splited[0]
                species_id = line_splited[1]
                species_num = int(line_splited[2])
                if bin_id not in bin_tax_dict.keys():
                    bin_tax_dict[bin_id] = {}
                if species_id not in bin_tax_dict[bin_id].keys():
                    bin_tax_dict[bin_id][species_id] = species_num
    if not bin_tax_max_dict:
        if os.path.exists(binphylo_path_binstax_max):
            bin_phylo_bins_max_lines = read_file(binphylo_path_binstax_max)
            for line in bin_phylo_bins_max_lines:
                line_splited = line.split("\t")
                bin_id = line_splited[0]
                species_id = line_splited[1]
                species_num = int(line_splited[2])
                if bin_id not in bin_tax_max_dict.keys():
                    bin_tax_max_dict[bin_id] = {}
                if species_id not in bin_tax_max_dict[bin_id].keys():
                    bin_tax_max_dict[bin_id][species_id] = species_num
    if not bin_group_tax_dict:
        # As for the kraken analysis a similar file "binned_taxa_path" should be created for binnin with MetaBinner and COMEBin and ue it below.
        if kraken_mode:
            if os.path.exists(binned_taxa_path):
                pr_line = ""
                protein_status = False
                species_status = False
                with open(binned_taxa_path) as bintaxa_lines:
                    for line in bintaxa_lines:
                        line = line.rstrip("\n")
                        if pr_line == "bin:":
                            bin_id = line
                            bin_group_tax_dict[bin_id] = {
                                "max_freq": None,
                                "protein_ids_k": [],
                                "species": []
                            }
                        elif pr_line == "max frequency:":
                            bin_group_tax_dict[bin_id]["max_freq"] = line
                        elif pr_line == "binned proteins:":
                            protein_status = True
                            species_status = False
                        elif pr_line == "species:":
                            protein_status = False
                            species_status = True
                        elif line == "//":
                            protein_status = False
                            species_status = False
                        if protein_status:
                            bin_group_tax_dict[bin_id]["protein_ids_k"].append(line)
                        if species_status:
                            bin_group_tax_dict[bin_id]["species"].append(line)
                        pr_line = line
        else:
            if os.path.exists(binning_results_path):
                bin_group_tax_dict = {}
                if not bin_group_tax_dict:
                    for bin_id in bin_tax_max_dict.keys():
                        bin_group_tax_dict[bin_id] = {
                            "max_freq": None,
                            "protein_ids_d": [],
                            "protein_ids_i": [],
                            "species": []
                        }
                        for species_id in bin_tax_max_dict[bin_id].keys():
                            if bin_group_tax_dict[bin_id]["max_freq"] is None:
                                max_freq = bin_tax_max_dict[bin_id][species_id]
                                bin_group_tax_dict[bin_id]["max_freq"] = max_freq
                            bin_group_tax_dict[bin_id]["species"].append(species_id)
                        # For each protein of the bin, if the protein is present in bins_prs_tax_dict[bin_id] then it has immediate taxonomy inference otherwise intemediate.
                        bin_protens_with_tax = list(bins_prs_tax_dict[bin_id].keys())
                        for protein_id in dict_prs_bins.keys():
                            if dict_prs_bins[protein_id] == bin_id:
                                if protein_id in bin_protens_with_tax:
                                    bin_group_tax_dict[bin_id]["protein_ids_d"].append(protein_id)
                                else:
                                    bin_group_tax_dict[bin_id]["protein_ids_i"].append(protein_id)
    return dict_seqs, dict_hm, dict_top, dict_sp, dict_nr, dict_genes, contig_gene_dist_dict, dict_input_motifs, swiss_fams_len_comp_dict, dict_contigs_bins, dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict


def tax_info_convert(bin_group_tax_dict):
    print("\nReforming the information related to the binned proteins...")
    # The process is faster if the key is the protein and the value is its bin, species and caterogization method.
    pr_tax_info_dcit = {}
    if bin_group_tax_dict:
        for bin_id in bin_group_tax_dict.keys():
            cat_type = "U"
            taxonomies = bin_group_tax_dict[bin_id]["species"]
            for key_info in bin_group_tax_dict[bin_id].keys():
                if key_info == "protein_ids_d":
                    cat_type = "D"
                if key_info == "protein_ids_i":
                    cat_type = "I"
                if key_info == "protein_ids_k":
                    cat_type = "K"
                if key_info in ["protein_ids_d", "protein_ids_i", "protein_ids_k"]:
                    for key_pr_ac in bin_group_tax_dict[bin_id][key_info]:
                        pr_tax_info_dcit[key_pr_ac] = [bin_id, taxonomies, cat_type]
    return pr_tax_info_dcit


def txt_results(annotation_file_txt_name, output_path_annotation, dict_hm, dict_seqs, dict_top, dict_sp, dict_nr, dict_input_motifs, dict_genes, swiss_fams_len_comp_dict, contig_gene_dist_dict, pr_tax_info_dcit, add_type, add_info):
    print("\nWritting the results in a TXT File...")
    # Final folder with annotation files is created.
    if os.path.exists(output_path_annotation):
        shutil.rmtree(output_path_annotation)
    os.mkdir(output_path_annotation)
    # Creating the TXT file for the results
    annotation_file = open(annotation_file_txt_name, "a")
    counter_seq = 0
    split_size = 70
    percentage = 10
    total_seqs = len(dict_seqs.keys())
    ten_percent_total_seqs = total_seqs * 0.1
    for key_pr_ac in dict_seqs:
        counter_seq += 1
        if counter_seq >= ten_percent_total_seqs:
            print("{}%".format(percentage))
            percentage += 10
            counter_seq = 0
        # Protein sequence and length
        annotation_file.write("\\\\{}\n".format(100*"-"))
        annotation_file.write("Putative protein ID:\n{}\n\n".format(key_pr_ac))
        seq = dict_seqs[key_pr_ac]
        seq_length = len(seq)
        annotation_file.write("Protein sequence:\n")
        ann_fasta_splited = [seq[anni:anni + split_size] for anni in range(0, len(seq), split_size)]
        for ann_fasta_line in ann_fasta_splited:
            annotation_file.write("{}\n".format(ann_fasta_line))
        annotation_file.write("\nSequence length:")
        annotation_file.write("\n{}\n\n".format(seq_length))
        annotation_file.write("Domains (Description\tName\tPfam ID\tSequence E-value\tBitscore\tDomain c-E-value\tDomain i-E-value\tDomain from\tDomain to\tSequence from\tSequence to):")
        if key_pr_ac in dict_hm.keys():
            for val_list in dict_hm[key_pr_ac]:
                domains_splited = val_list.split("\t")
                domains_line = "\t".join([domains_splited[5], domains_splited[3], domains_splited[4], domains_splited[1], domains_splited[2], domains_splited[6], domains_splited[7], domains_splited[8], domains_splited[9], domains_splited[10], domains_splited[11]])
                annotation_file.write("\n{}".format(domains_line))
        else:
            annotation_file.write("\nNo domains found.")
        # Bin taxonomies
        # There is a chance that the proteins, which are encoded from genes whose taxonomy was directly inferred, are not present in the annotation files because they do not contain
        # seek domains.
        if pr_tax_info_dcit:
            if key_pr_ac in pr_tax_info_dcit.keys():
                pr_bin_id = pr_tax_info_dcit[key_pr_ac][0]
                pr_taxonomies = pr_tax_info_dcit[key_pr_ac][1]
                pr_cat_type = pr_tax_info_dcit[key_pr_ac][2]
                annotation_file.write("\n\nBin ID:\n{}".format(pr_bin_id))
                annotation_file.write("\n\nProtein taxonomy - {}:\n".format(pr_cat_type))
                for ti in range(0, len(pr_taxonomies)):
                    item = pr_taxonomies[ti]
                    if (ti + 1) == len(pr_taxonomies):
                        annotation_file.write("{}".format(item))
                    else:
                        annotation_file.write("{}\t".format(item))
            else:
                annotation_file.write("\n\nBin ID:\n-")
                annotation_file.write("\n\nProtein taxonomy:\nCould not be inferred.")
        # Signal peptide and transmembrane regions
        if dict_top:
            annotation_file.write("\n\nSignal peptide and Transmembrane regions (Region\tRegion from\tRegion to):")
            if key_pr_ac not in dict_top.keys():
                annotation_file.write("\nNo signal peptide or transmembrane regions found.")
            else:
                for sp_tr_region in dict_top[key_pr_ac]:
                    sp_tr_region_splited = sp_tr_region.split("\t")
                    annotation_file.write("\n{}\t{}\t{}".format(sp_tr_region_splited[2], sp_tr_region_splited[0],  sp_tr_region_splited[1]))
        # Best hit in the Swissprot protein database.
        if key_pr_ac in dict_sp.keys():
            annotation_file.write("\n\nSwissProt database hit (SwissProt Accession Number\tE-value\tPercentage Identity\tBitscore):")
            sp_hit = dict_sp[key_pr_ac]
            sp_hit = sp_hit.split("\t")
            sp_line = "\t".join([sp_hit[1], sp_hit[3], sp_hit[2], sp_hit[4]])
            annotation_file.write("\n{}".format(sp_line))
        # Best hit in the NR protein database
        if key_pr_ac in dict_nr.keys():
            annotation_file.write("\n\nNon-redudant database hit (NR protein ID\tE-value\tPercentage Identity\tBitscore):")
            nr_hit = dict_nr[key_pr_ac]
            nr_hit = nr_hit.split("\t")
            nr_line = "\t".join([nr_hit[1], nr_hit[3], nr_hit[2], nr_hit[4]])
            annotation_file.write("\n{}".format(nr_line))
        # Any input motif found for a protein is written
        if key_pr_ac in dict_input_motifs.keys():
            annotation_file.write("\n\nInput motifs identified (Motif\tSequence Start\tSequence End):")
            motif_hit = dict_input_motifs[key_pr_ac]
            for motif_list in motif_hit:
                motif_seq = motif_list[0]
                motif_seq_splited = motif_seq.split("\\S")
                motif_seq = "X".join(motif_seq_splited)
                annotation_file.write("\n{}\t{}\t{}".format(motif_seq, motif_list[1], motif_list[2]))
        # Gene information
        if key_pr_ac in dict_genes.keys():
            annotation_file.write("\n\nGene Sequence:")
            gene_sequence = dict_genes[key_pr_ac][0]
            annotation_file.write("\n>{}".format(key_pr_ac))
            ann_fasta_splited = [gene_sequence[anni:anni + split_size] for anni in range(0, len(gene_sequence), split_size)]
            for ann_fasta_line in ann_fasta_splited:
                annotation_file.write("\n{}".format(ann_fasta_line))
            star_codon_status = dict_genes[key_pr_ac][1]
            end_codon_status = dict_genes[key_pr_ac][2]
            annotation_file.write("\nStart Codon:\n{}".format(star_codon_status))
            annotation_file.write("\nEnd Codon:\n{}".format(end_codon_status))
        # Gene distance from contig's edges.
        if contig_gene_dist_dict and (key_pr_ac in contig_gene_dist_dict.keys()):
            annotation_file.write("\n\nGene distance from contigs (Distance from contig start\tDistance from contig end):")
            start_dist = contig_gene_dist_dict[key_pr_ac][0]
            end_dist = contig_gene_dist_dict[key_pr_ac][1]
            annotation_file.write("\n{}\t{}".format(start_dist, end_dist))
        # Family prediction and family mean length comparison.
        if swiss_fams_len_comp_dict and (key_pr_ac in swiss_fams_len_comp_dict.keys()):
            annotation_file.write("\n\nPredicted protein family (PPF) and length comparison (PPF name\tPPF mean length\tPPF median length\tLength difference\tLength relative change\tFamily difference:")
            for item in swiss_fams_len_comp_dict[key_pr_ac]:
                pf_name = item[0]
                pf_mean_length = item[1]
                len_diff = item[2]
                len_rel_change = item[3]
                fam_dif = item[4]
                annotation_file.write("\n{}\t{}\t{}\t{}\t{}".format(pf_name, pf_mean_length, len_diff, len_rel_change, fam_dif))
        # Added types and information
        if add_type:
            annotation_file.write("\n\nCustom information:")
            for add_i in range(0, len(add_type)):
                type_item = add_type[add_i]
                info_item = add_info[add_i]
                annotation_file.write("\n{}: {}".format(type_item, info_item))
        # Termination symbol(s) for current protein accession.
        annotation_file.write("\n//{}\n\n\n".format(100*"-"))
    print("100%")
    # The new Excel and TXT files are saved in memory.
    annotation_file.close()


def excel_results(annotation_file_excel_name, dict_hm, dict_seqs, dict_top, dict_sp, dict_nr, dict_input_motifs, dict_genes, swiss_fams_len_comp_dict, contig_gene_dist_dict, pr_tax_info_dcit, add_type, add_info):
    print("\nWritting the results in an EXCEL File...")
    total_seqs = len(dict_seqs.keys())
    if total_seqs > 1000000:
        print("\nThe number of proteins is too large to be able to write their information in an Excel file. Annotation information will not be written in an Excel file.")
        return
    # If the file path to which the results are added is None then a new Excel file must be created. Otherwise, the results are saved to the already existing file.
    new_excel = Workbook()
    new_excel.save(annotation_file_excel_name)
    new_excel_edit = load_workbook(annotation_file_excel_name)
    # Creating, saving and loading the new or preexisting Excel file for editing by sving the results
    sheet_edit = new_excel_edit.active
    # If the new file already exists there is no need to rewrite the headers of the columns and set their widths.
    # The indexes of the rows and columns in the excel file begin at 1.
    row_i = 1
    col_i = 1
    # Protein ID.
    column_letter = get_column_letter(col_i)
    sheet_edit.column_dimensions[column_letter].width = 50
    sheet_edit.cell(row=row_i, column=col_i).value = "Putative Protein ID"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Putative Protein ID"
    col_i += 1
    # Protein sequence.
    column_letter = get_column_letter(col_i)
    sheet_edit.column_dimensions[column_letter].width = 20
    sheet_edit.cell(row=row_i, column=col_i).value = "Protein Sequence"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Protein Sequence"
    col_i += 1
    # Protein sequence length.
    for col_j in range(col_i, col_i+2):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Sequence Length"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Sequence Length"
    col_i += 1
    # Bin ID.
    for col_j in range(col_i, col_i+2):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Bin ID"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Bin ID"
    col_i += 1
    # Bin taxonomies.
    for col_j in range(col_i, col_i+2):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Bin Taxonomies"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Bin Taxonomies"
    col_i += 1
    # Domain information.
    for col_j in range(col_i, col_i+11):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+3).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+4).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+5).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+6).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+7).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+8).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+9).value = "Domains"
    sheet_edit.cell(row=row_i, column=col_i+10).value = "Domains"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Description"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Name"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "Pfam ID"
    sheet_edit.cell(row=row_i+1, column=col_i+3).value = "Sequence E-value"
    sheet_edit.cell(row=row_i+1, column=col_i+4).value = "Bitscore"
    sheet_edit.cell(row=row_i+1, column=col_i+5).value = "Domain c-E-value"
    sheet_edit.cell(row=row_i+1, column=col_i+6).value = "Domain i-E-value"
    sheet_edit.cell(row=row_i+1, column=col_i+7).value = "Domain from"
    sheet_edit.cell(row=row_i+1, column=col_i+8).value = "Domain to"
    sheet_edit.cell(row=row_i+1, column=col_i+9).value = "Sequence from"
    sheet_edit.cell(row=row_i+1, column=col_i+10).value = "Sequence to"
    col_i += 11
    # Input motifs.
    for col_j in range(col_i, col_i+3):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Input motifs"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Input motifs"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Input motifs"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Input motif"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Region(s) from"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "Region(s) to"
    col_i += 3
    # Protein family predictions.
    for col_j in range(col_i, col_i+2):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Protein family"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Protein family"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Protein family"
    sheet_edit.cell(row=row_i, column=col_i+3).value = "Protein family"
    sheet_edit.cell(row=row_i, column=col_i+4).value = "Protein family"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Name"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Mean length"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "Length difference"
    sheet_edit.cell(row=row_i+1, column=col_i+3).value = "Length relative change"
    sheet_edit.cell(row=row_i+1, column=col_i+4).value = "Family difference"
    col_i += 5
    # Signal peptide and transmembrane regions.
    for col_j in range(col_i, col_i+3):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Signal Peptide or Transmembrane Region"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Signal Peptide or Transmembrane Region"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Signal Peptide or Transmembrane Region"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Region"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Region from"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "Region to"
    col_i += 3
    # BLASTP against SwissProt information.
    for col_j in range(col_i, col_i+4):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Swiss-Prot"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Swiss-Prot"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Swiss-Prot"
    sheet_edit.cell(row=row_i, column=col_i+3).value = "Swiss-Prot"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Swiss-Prot Accession Number"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Percentage Identity"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "E-value"
    sheet_edit.cell(row=row_i+1, column=col_i+3).value = "Bitscore"
    col_i += 4
    # BLASTP against NR information.
    for col_j in range(col_i, col_i+4):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "FPD"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "FPD"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "FPD"
    sheet_edit.cell(row=row_i, column=col_i+3).value = "FPD"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Protein ID"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Percentage Identity"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "E-value"
    sheet_edit.cell(row=row_i+1, column=col_i+3).value = "Bitscore"
    col_i += 4
    # Gene information.
    for col_j in range(col_i, col_i+3):
        column_letter = get_column_letter(col_j)
        sheet_edit.column_dimensions[column_letter].width = 15
    sheet_edit.cell(row=row_i, column=col_i).value = "Gene Information"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Gene Information"
    sheet_edit.cell(row=row_i, column=col_i+2).value = "Gene Information"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Gene Sequence"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "Start Codon"
    sheet_edit.cell(row=row_i+1, column=col_i+2).value = "End Codon"
    col_i += 3
    # Gene information for start and end distances from contig's edges.
    sheet_edit.cell(row=row_i, column=col_i).value = "Gene Information"
    sheet_edit.cell(row=row_i, column=col_i+1).value = "Gene Information"
    sheet_edit.cell(row=row_i+1, column=col_i).value = "Start Distance"
    sheet_edit.cell(row=row_i+1, column=col_i+1).value = "End Distance"
    col_i += 2
    # Adding information set by the user
    if add_type:
        for add_i in range(0, len(add_type)):
            type_item = add_type[add_i]
            sheet_edit.cell(row=row_i, column=col_i + add_i).value = "Custom Information"
            sheet_edit.cell(row=row_i+1, column=col_i + add_i).value = type_item
        col_i += len(add_type)
    # If the results are going to be added in an EXCEL file then find the maximum row of that EXCEL file
    row_i = 3
    col_i = 1
    counter_sqs_1 = 0
    counter_sqs_2 = 0
    perc_sqs = total_seqs*0.1
    for key_pr_ac in dict_seqs:
        counter_sqs_1 += 1
        counter_sqs_2 += 1
        if counter_sqs_1 >= perc_sqs:
            print("{} / {}".format(counter_sqs_2, total_seqs))
            counter_sqs_1 = 0
        # Protein sequence and length
        seq = dict_seqs[key_pr_ac]
        seq_length = len(seq)
        # Protein ID
        sheet_edit.cell(row=row_i, column=col_i).value = key_pr_ac
        # Protein sequence
        col_i += 1
        sheet_edit.cell(row=row_i, column=col_i).value = seq
        # Sequence length
        col_i += 1
        sheet_edit.cell(row=row_i, column=col_i).value = seq_length
        # Bin information
        col_i += 1
        if pr_tax_info_dcit:
            if key_pr_ac in pr_tax_info_dcit.keys():
                pr_bin_id = pr_tax_info_dcit[key_pr_ac][0]
                pr_taxonomies = pr_tax_info_dcit[key_pr_ac][1]
                pr_cat_type = pr_tax_info_dcit[key_pr_ac][2]
                cell_phrase = "-"
                sheet_edit.cell(row=row_i, column=col_i).value = pr_bin_id
                col_i += 1
                pr_taxonomies_wr = ", ".join(pr_taxonomies)
                cell_phrase = "{}: {}".format(pr_cat_type, pr_taxonomies_wr)
                sheet_edit.cell(row=row_i, column=col_i).value = cell_phrase
            else:
                cell_phrase = "-"
                sheet_edit.cell(row=row_i, column=col_i).value = cell_phrase
                col_i += 1
                sheet_edit.cell(row=row_i, column=col_i).value = cell_phrase
        else:
            cell_phrase = "-"
            sheet_edit.cell(row=row_i, column=col_i).value = cell_phrase
            col_i += 1
            sheet_edit.cell(row=row_i, column=col_i).value = cell_phrase
        # Protein domains
        col_i += 1
        if key_pr_ac in dict_hm.keys():
            domain_info_seq_evalue = ""
            domain_info_bitscore = ""
            domain_info_name = ""
            domain_info_pfam_id = ""
            domain_info_description = ""
            domain_info_domain_cevalue = ""
            domain_info_domain_ievalue = ""
            domain_info_domain_from = ""
            domain_info_domain_to = ""
            domain_info_seq_from = ""
            domain_info_seq_to = ""
            for pr_i in range(0, len(dict_hm[key_pr_ac])):
                pr_domains = dict_hm[key_pr_ac][pr_i]
                domains_splited = pr_domains.split("\t")
                if pr_i == 0:
                    domain_info_seq_evalue = domains_splited[1]
                    domain_info_bitscore = domains_splited[2]
                    domain_info_name = domains_splited[3]
                    domain_info_pfam_id = domains_splited[4]
                    domain_info_description = domains_splited[5]
                    domain_info_domain_cevalue = domains_splited[6]
                    domain_info_domain_ievalue = domains_splited[7]
                    domain_info_domain_from = domains_splited[8]
                    domain_info_domain_to = domains_splited[9]
                    domain_info_seq_from = domains_splited[10]
                    domain_info_seq_to = domains_splited[11]
                else:
                    domain_info_seq_evalue = "{}\t{}".format(domain_info_seq_evalue, domains_splited[1])
                    domain_info_bitscore = "{}\t{}".format(domain_info_bitscore, domains_splited[2])
                    domain_info_name = "{}\t{}".format(domain_info_name, domains_splited[3])
                    domain_info_pfam_id = "{}\t{}".format(domain_info_pfam_id, domains_splited[4])
                    domain_info_description = "{}\t{}".format(domain_info_description, domains_splited[5])
                    domain_info_domain_cevalue = "{}\t{}".format(domain_info_domain_cevalue, domains_splited[6])
                    domain_info_domain_ievalue = "{}\t{}".format(domain_info_domain_ievalue, domains_splited[7])
                    domain_info_domain_from = "{}\t{}".format(domain_info_domain_from, domains_splited[8])
                    domain_info_domain_to = "{}\t{}".format(domain_info_domain_to, domains_splited[9])
                    domain_info_seq_from = "{}\t{}".format(domain_info_seq_from, domains_splited[10])
                    domain_info_seq_to = "{}\t{}".format(domain_info_seq_to, domains_splited[11])
            sheet_edit.cell(row=row_i, column=col_i).value = domain_info_description
            sheet_edit.cell(row=row_i, column=col_i+1).value = domain_info_name
            sheet_edit.cell(row=row_i, column=col_i+2).value = domain_info_pfam_id
            sheet_edit.cell(row=row_i, column=col_i+3).value = domain_info_seq_evalue
            sheet_edit.cell(row=row_i, column=col_i+4).value = domain_info_bitscore
            sheet_edit.cell(row=row_i, column=col_i+5).value = domain_info_domain_cevalue
            sheet_edit.cell(row=row_i, column=col_i+6).value = domain_info_domain_ievalue
            sheet_edit.cell(row=row_i, column=col_i+7).value = domain_info_domain_from
            sheet_edit.cell(row=row_i, column=col_i+8).value = domain_info_domain_to
            sheet_edit.cell(row=row_i, column=col_i+9).value = domain_info_seq_from
            sheet_edit.cell(row=row_i, column=col_i+10).value = domain_info_seq_to
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 2).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 3).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 4).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 5).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 6).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 7).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 8).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 9).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 10).value = "-"
        col_i += 11
        # Input motifs
        if key_pr_ac not in dict_input_motifs.keys():
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+2).value = "-"
        else:
            motif_sqs = None
            motifs_start = None
            motifs_end = None
            for pr_i in range(0, len(dict_input_motifs[key_pr_ac])):
                item = dict_input_motifs[key_pr_ac][pr_i]
                if pr_i == 0:
                    motif_pre_sqs = item[0]
                    motif_pre_sqs_splited = motif_pre_sqs.split("\\S")
                    motif_sqs = "X".join(motif_pre_sqs_splited)
                    motifs_start = item[1]
                    motifs_end = item[2]
                else:
                    motif_sqs_cur = item[0]
                    motif_sqs_cur_splited = motif_sqs_cur.split("\\S")
                    motif_sqs_cur = "X".join(motif_sqs_cur_splited)
                    motif_sqs = "{}\t{}".format(motif_sqs, motif_sqs_cur)
                    motifs_start = "{}\t{}".format(motifs_start, item[1])
                    motifs_end = "{}\t{}".format(motifs_end, item[2])
            sheet_edit.cell(row=row_i, column=col_i).value = motif_sqs
            sheet_edit.cell(row=row_i, column=col_i+1).value = motifs_start
            sheet_edit.cell(row=row_i, column=col_i+2).value = motifs_end
        col_i += 3
        # Protein family predictions.
        if (not swiss_fams_len_comp_dict) or (key_pr_ac not in swiss_fams_len_comp_dict.keys()):
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+2).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+3).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+4).value = "-"
        else:
            pf_name = None
            pf_mean_length = None
            len_diff = None
            len_rel_change = None
            for pr_i in range(0, len(swiss_fams_len_comp_dict[key_pr_ac])):
                item = swiss_fams_len_comp_dict[key_pr_ac][pr_i]
                if pr_i == 0:
                    pf_name = item[0]
                    pf_mean_length = item[1]
                    len_diff = item[2]
                    len_rel_change = item[3]
                    fam_dif = item[4]
                else:
                    pf_name_cur = item[0]
                    pf_mean_length_cur = item[1]
                    len_diff_cur = item[2]
                    len_rel_change_cur = item[3]
                    fam_dif_cur = item[4]
                    pf_name = "{}\t{}".format(pf_name, pf_name_cur)
                    pf_mean_length = "{}\t{}".format(pf_mean_length, pf_mean_length_cur)
                    len_diff = "{}\t{}".format(len_diff, len_diff_cur)
                    len_rel_change = "{}\t{}".format(len_rel_change, len_rel_change_cur)
                    fam_dif = "{}\t{}".format(fam_dif, fam_dif_cur)
            sheet_edit.cell(row=row_i, column=col_i).value = pf_name
            sheet_edit.cell(row=row_i, column=col_i+1).value = pf_mean_length
            sheet_edit.cell(row=row_i, column=col_i+2).value = len_diff
            sheet_edit.cell(row=row_i, column=col_i+3).value = len_rel_change
            sheet_edit.cell(row=row_i, column=col_i+4).value = fam_dif
        col_i += 5
        # Signal peptide and transmembrane regions
        if dict_top:
            if key_pr_ac not in dict_top.keys():
                sheet_edit.cell(row=row_i, column=col_i).value = "-"
                sheet_edit.cell(row=row_i, column=col_i+1).value = "-"
                sheet_edit.cell(row=row_i, column=col_i+2).value = "-"
            else:
                region_type = None
                region_from = None
                region_to = None
                for sp_i in range(0, len(dict_top[key_pr_ac])):
                    sp_tr_region = dict_top[key_pr_ac][sp_i]
                    sp_tr_region_splited = sp_tr_region.split("\t")
                    if sp_i == 0:
                        region_type = sp_tr_region_splited[2]
                        region_from = sp_tr_region_splited[0]
                        region_to = sp_tr_region_splited[1]
                    else:
                        region_type_cur = sp_tr_region_splited[2]
                        region_from_cur = sp_tr_region_splited[0]
                        region_to_cur = sp_tr_region_splited[1]
                        region_type = "{}\t{}".format(region_type, region_type_cur)
                        region_from = "{}\t{}".format(region_from, region_from_cur)
                        region_to = "{}\t{}".format(region_to, region_to_cur)
                    sheet_edit.cell(row=row_i, column=col_i).value = region_type
                    sheet_edit.cell(row=row_i, column=col_i+1).value = region_from
                    sheet_edit.cell(row=row_i, column=col_i+2).value = region_to
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+2).value = "-"
        col_i += 3
        # Best hit in the SwissProt protein database
        if key_pr_ac in dict_sp.keys():
            sp_hit = dict_sp[key_pr_ac]
            sp_hit = sp_hit.split("\t")
            sp_name = sp_hit[1]
            sp_evalue = sp_hit[2]
            sp_identity = sp_hit[3]
            sp_bitscore = sp_hit[4]
            sheet_edit.cell(row=row_i, column=col_i).value = sp_name
            sheet_edit.cell(row=row_i, column=col_i+1).value = sp_evalue
            sheet_edit.cell(row=row_i, column=col_i+2).value = sp_identity
            sheet_edit.cell(row=row_i, column=col_i+3).value = sp_bitscore
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+2).value = "-"
            sheet_edit.cell(row=row_i, column=col_i+3).value = "-"
        col_i += 4
        # Best hit in the NR protein database
        if key_pr_ac in dict_nr.keys():
            nr_hit = dict_nr[key_pr_ac]
            nr_hit_splited = nr_hit.split("\t")
            nr_protein_id = nr_hit_splited[1]
            nr_evalue = nr_hit_splited[2]
            nr_identity = nr_hit_splited[3]
            nr_bitscore = nr_hit_splited[4]
            sheet_edit.cell(row=row_i, column=col_i).value = nr_protein_id
            sheet_edit.cell(row=row_i, column=col_i + 1).value = nr_evalue
            sheet_edit.cell(row=row_i, column=col_i + 2).value = nr_identity
            sheet_edit.cell(row=row_i, column=col_i + 3).value = nr_bitscore
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 2).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 3).value = "-"
        col_i += 4
        # Information for gene sequences.
        if key_pr_ac in dict_genes.keys():
            gene_sequence = dict_genes[key_pr_ac][0]
            start_codon = dict_genes[key_pr_ac][1]
            end_codon = dict_genes[key_pr_ac][2]
            sheet_edit.cell(row=row_i, column=col_i).value = gene_sequence
            sheet_edit.cell(row=row_i, column=col_i + 1).value = start_codon
            sheet_edit.cell(row=row_i, column=col_i + 2).value = end_codon
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 1).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 2).value = "-"
        col_i += 3
        # Information for gene distance from contig's edges.
        if (contig_gene_dist_dict is not None) and (key_pr_ac in contig_gene_dist_dict.keys()):
            start_dist = contig_gene_dist_dict[key_pr_ac][0]
            end_dist = contig_gene_dist_dict[key_pr_ac][1]
            sheet_edit.cell(row=row_i, column=col_i).value = start_dist
            sheet_edit.cell(row=row_i, column=col_i + 1).value = end_dist
        else:
            sheet_edit.cell(row=row_i, column=col_i).value = "-"
            sheet_edit.cell(row=row_i, column=col_i + 1).value = "-"
        col_i += 2
        # Adding information set by the user
        if add_info:
            for add_i in range(0, len(add_info)):
                type_info = add_info[add_i]
                sheet_edit.cell(row=row_i, column=col_i + add_i).value = type_info
            col_i += len(add_info)
        # After all information for a protein have been written, the index of the row must be increased by the number of rows of the maximum number of rows between the information of domains, signal peptide - transmembrane domains and gene ontology.
        # The index of the column must be set again to 1.
        row_i += 1
        col_i = 1
    print("{} / {}".format(total_seqs, total_seqs))
    # The new Excel and TXT files are saved in memory.
    new_excel_edit.save(annotation_file_excel_name)


def enzannmtg(input_folder=None, sra_code=False, contigs=False, protein_input=False, adapters_path="adapters.fa", protein_db_path="", kraken_db_path="", profiles_path="", profiles_phylo_path="", profiles_broad_path="", swissprot_path="", motifs_path="motifs.txt", options_file_path="", output_path="", family_code=None, family_code_phylo=None, db_name="", db_name_phylo="", input_protein_names_status=False, input_protein_names=None, input_protein_names_phylo_status=False, input_protein_names_phylo=None, name_thr=0.5, seek_route=3, paired_end=True, compressed=True, create_nr_db_status=False, prefetch_size=20, adapters_status="pre", add_seek_info=True, add_taxonomy_info=True, skip_fastqc=False, bbduk_max_ram=4, clear_space=False, k_list=None, kraken_mode=True, kraken_threshold="0.1", kraken_memory_mapping=True, binning_tool=1, bin_ram_ammount=4, bin_num_contig_len=500, bin_num_kmer=4, comebin_batch_size=256, cd_hit_t=0.99, cd_hit_mem=4000, prs_source=1, genetic_code=11, val_type="--cut_ga ", second_dom_search=True, e_value_nodom_thr=1e-70, add_type=[], add_info=[], thread_num=4, pdf_threads=None, after_trimming=False, after_alignment=False, after_gene_pred=False, after_binning=False, after_db=False, after_tm=False, after_ap=False, up_to_sra=False, up_to_databases=False, up_to_trimming_com=False, up_to_trimming_uncom=False, up_to_alignment=False, sra_env="ps_sra_tools", fastqc_env="ps_fastqc", bbduk_env="ps_bbtools", megahit_env="ps_megahit", kraken_env="ps_kraken", metabinner_env="ps_metabinner", comebin_env="ps_comebin", cdhit_env="ps_cd_hit", genepred_env="", hmmer_env="ps_hmmer", diamond_env="ps_diamond", taxonkit_env="ps_taxonkit", phobius_env="ps_phobius", bowtie_env="ps_bowtie", ana_dir_path="", ana_sh_path="", prefetch_path="", vdb_validate_path="", fastq_dump_path="", fastqc_path="", gzip_path="", cat_path="", bbduk_path="", megahit_path="", kraken_path="", metabinner_bin_path="", comebin_bin_path="", fraggenescanrs_path="", hmmscan_path="", hmmpress_path="", hmmfetch_path="", diamond_path="", cd_hit_path="", taxonkit_path="", phobius_path="", bowtie_build_path="", bowtie_path="", input_command=None):
    # Start time
    start_time = time.time()
    
    # Print a new line in the output.
    print("\n----------------")
    print("Pipeline Start")
    print("----------------")
    print()

    if options_file_path:
        print("\nOptions file path: {}".format(options_file_path))
    else:
        print("\nOptions file path: None")
    
    # Read the values for the options from a file, if its was specified.
    if options_file_path:
        if not os.path.exists(options_file_path):
            print("\nThe path to the options file is wrong. Exiting.")
            exit()
        print("Option values set from the input file:")
        option_lines = read_file(options_file_path)
        for line in option_lines:
            if line:
                if line[0] == "#":
                    continue
                if "=" in line:
                    line_splited = line.split("=")
                    option_type = line_splited[0]
                    option_value = line_splited[1]
                    if option_value and option_value != "\"\"":
                        print("{}: {}".format(option_type, option_value))
                        # Procesing the option value.
                        if option_value[0] == "\"":
                            option_value = option_value[1:]
                        if option_value[-1] == "\"":
                            option_value = option_value[:-1]
                        # Passing the options value to the corresponding option types.
                        # Input and output options
                        if option_type == "input_folder":
                            input_folder = option_value
                        if option_type == "sra_code":
                            sra_code = option_value
                        if option_type == "contigs":
                            contigs = option_value
                            arg_str = "contigs"
                            contigs = check_i_value(contigs, arg_str)
                        if option_type == "protein_input":
                            protein_input = option_value
                            arg_str = "protein_input"
                            protein_input = check_i_value(protein_input, arg_str)
                        if option_type == "adapters_path":
                            adapters_path = option_value
                        if option_type == "protein_db_path":
                            protein_db_path = option_value
                        if option_type == "kraken_db_path":
                            kraken_db_path = option_value
                        if option_type == "profiles_path":
                            profiles_path = option_value
                        if option_type == "profiles_phylo_path":
                            profiles_phylo_path = option_value
                        if option_type == "profiles_broad_path":
                            profiles_broad_path = option_value
                        if option_type == "swissprot_path":
                            swissprot_path = option_value
                        if option_type == "motifs_path":
                            motifs_path = option_value
                        if option_type == "output_path":
                            output_path = option_value
                            if output_path[-1] == "/":
                                output_path = output_path[:-1]
                        # Protein family options
                        if option_type == "family_code":
                            family_code = option_value
                        if option_type == "family_code_phylo":
                            family_code_phylo = option_value
                        if option_type == "db_name":
                            db_name = option_value
                        if option_type == "db_name_phylo":
                            db_name_phylo = option_value
                        if option_type == "input_protein_names_status":
                            input_protein_names_status = option_value
                            arg_str = "input_protein_names_status"
                            input_protein_names_status = check_i_value(input_protein_names_status, arg_str)
                        if option_type == "input_protein_names":
                            input_protein_names = option_value
                            input_protein_names = input_protein_names.replace("_", " ")
                            input_protein_names = input_protein_names.split(",")
                        if option_type == "input_protein_names_phylo_status":
                            input_protein_names_phylo_status = option_value
                            arg_str = "input_protein_names_phylo_status"
                            input_protein_names_phylo_status = check_i_value(input_protein_names_phylo_status, arg_str)
                        if option_type == "input_protein_names_phylo":
                            input_protein_names_phylo = option_value
                            input_protein_names_phylo = input_protein_names_phylo.replace("_", " ")
                            input_protein_names_phylo = input_protein_names_phylo.split(",")
                        if option_type == "name_threshold":
                            name_thr = option_value
                            if "." in name_thr:
                                name_thr = float(name_thr)
                            else:
                                name_thr = int(name_thr)
                        # General options
                        # Pipeline
                        if option_type == "seek_route":
                            seek_route = int(option_value)
                        if option_type == "paired_end":
                            paired_end = option_value
                            arg_str = "paired_end"
                            paired_end = check_i_value(paired_end, arg_str)
                        if option_type == "compressed":
                            compressed = option_value
                            arg_str = "compressed"
                            compressed = check_i_value(compressed, arg_str)
                        if option_type == "create_nr_db_status":
                            create_nr_db_status = option_value
                            arg_str = "-create_nr_db_status"
                            create_nr_db_status = check_i_value(create_nr_db_status, arg_str)
                        if option_type == "prefetch_size":
                            prefetch_size = int(option_value)
                        if option_type == "adapters_status":
                            adapters_status = option_value
                        if option_type == "add_seek_info":
                            add_seek_info = option_value
                            arg_str = "-add_seek_info"
                            add_seek_info = check_i_value(add_seek_info, arg_str)
                        if option_type == "add_taxonomy_info":
                            add_taxonomy_info = option_value
                            arg_str = "-add_taxonomy_info"
                            add_taxonomy_info = check_i_value(add_taxonomy_info, arg_str)
                        # FastQC
                        if option_type == "skip_fastqc":
                            skip_fastqc = option_value
                            arg_str = "skip_fastqc"
                            skip_fastqc = check_i_value(skip_fastqc, arg_str)
                        # BBDuk
                        if option_type == "bbduk_max_ram":
                            bbduk_max_ram = int(option_value)
                        if option_type == "clear_space":
                            clear_space = option_value
                            arg_str = "clear_space"
                            clear_space = check_i_value(clear_space, arg_str)
                        # Megahit
                        if option_type == "k_list":
                            k_list = option_value
                        # Kraken
                        if option_type == "kraken_mode":
                            kraken_mode = option_value
                            arg_str = "kraken_mode"
                            kraken_mode = check_i_value(kraken_mode, arg_str)
                        if option_type == "kraken_threshold":
                            kraken_threshold = option_value
                        if option_type == "kraken_memory_mapping":
                            kraken_memory_mapping = option_value
                            arg_str = "kraken_memory_mapping"
                            kraken_memory_mapping = check_i_value(kraken_memory_mapping, arg_str)
                        # Binning
                        if option_type == "binning_tool":
                            binning_tool = int(option_value)                        
                        if option_type == "binning_max_ram":
                            bin_ram_ammount = int(option_value)
                        if option_type == "bin_contig_len":
                            bin_num_contig_len = int(option_value)
                        if option_type == "bin_kmer":
                            bin_num_kmer = int(option_value)
                        if option_type == "comebin_batch_size":
                            comebin_batch_size = int(option_value)
                        # CD-HIT
                        if option_type == "cd_hit_max_mem":
                            cd_hit_mem = int(option_value)
                        if option_type == "cd_hit_t":
                            cd_hit_t = float(option_value)
                        # Gene prediction
                        if option_type == "gene_encoding":
                            prs_source = int(option_value)
                        if option_type == "genetic_code":
                            genetic_code = int(option_value)
                        # HMMER
                        if option_type == "val_type":
                            val_type = option_value
                            if val_type == "cut_ga":
                                val_type = "--{} ".format(val_type)
                            elif val_type == "default":
                                val_type = ""
                        if option_type == "second_dom_search":
                            arg_second_dom_search = option_value
                            arg_str = "-second_dom_search"
                            arg_second_dom_search = check_i_value(arg_second_dom_search, arg_str)
                        if option_type == "e_value_nodom_thr":
                            e_value_nodom_thr = int(option_value)
                            e_value_nodom_thr = float(1*(10**(-e_value_nodom_thr)))                            
                        # Annotation
                        if option_type == "add_type":
                            add_type = option_value
                        if option_type == "add_info":
                            add_info = option_value
                        # Threads
                        if option_type == "thread_num":
                            thread_num = int(option_value)
                        if option_type == "filtering_threads":
                            pdf_threads = int(option_value)
                        # Processes performed after
                        if option_type == "after_preprocessing":
                            after_preprocessing = option_value
                            arg_str = "-after_preprocessing"
                            after_preprocessing = check_i_value(after_preprocessing, arg_str)
                        if option_type == "after_assembly":
                            after_alignment = option_value
                            arg_str = "after_assembly"
                            after_alignment = check_i_value(after_alignment, arg_str)
                        if option_type == "after_gene_pred":
                            after_gene_pred = option_value
                            arg_str = "after_gene_pred"
                            after_gene_pred = check_i_value(after_gene_pred, arg_str)
                        if option_type == "after_binning":
                            after_binning = option_value
                            arg_str = "after_binning"
                            after_binning = check_i_value(after_binning, arg_str)
                        if option_type == "after_db":
                            after_db = option_value
                            arg_str = "after_db"
                            after_db = check_i_value(after_db, arg_str)
                        if option_type == "after_tm":
                            after_tm = option_value
                            arg_str = "after_tm"
                            after_tm = check_i_value(after_tm, arg_str)
                        if option_type == "after_ap":
                            after_ap = option_value
                            arg_str = "after_ap"
                            after_ap = check_i_value(after_ap, arg_str)
                        # Processes performed up to
                        if option_type == "up_to_sra":
                            up_to_sra = option_value
                            arg_str = "up_to_sra"
                            up_to_sra = check_i_value(up_to_sra, arg_str)
                        if option_type == "up_to_databases":
                            up_to_databases = option_value
                            arg_str = "up_to_databases"
                            up_to_databases = check_i_value(up_to_databases, arg_str)
                        if option_type == "up_to_preprocessing_com":
                            up_to_preprocessing_com = option_value
                            arg_str = "up_to_preprocessing_com"
                            up_to_preprocessing_com = check_i_value(up_to_preprocessing_com, arg_str)
                        if option_type == "up_to_preprocessing_uncom":
                            up_to_preprocessing_uncom = option_value
                            arg_str = "up_to_preprocessing_uncom"
                            up_to_preprocessing_uncom = check_i_value(up_to_preprocessing_uncom, arg_str)
                        if option_type == "up_to_assembly":
                            up_to_alignment = option_value
                            arg_str = "up_to_assembly"
                            up_to_alignment = check_i_value(up_to_alignment, arg_str)
                        # Tool enviroments
                        if option_type == "sra_env":
                            sra_env = option_value
                            if sra_env in ["None", "none"]:
                                sra_env = ""
                        if option_type == "fastqc_env":
                            fastqc_env = option_value
                            if fastqc_env in ["None", "none"]:
                                fastqc_env = ""
                        if option_type == "bbduk_env":
                            bbduk_env = option_value
                            if bbduk_env in ["None", "none"]:
                                bbduk_env = ""
                        if option_type == "megahit_env":
                            megahit_env = option_value
                            if megahit_env in ["None", "none"]:
                                megahit_env = ""
                        if option_type == "kraken_env":
                            kraken_env = option_value
                            if kraken_env in ["None", "none"]:
                                kraken_env = ""
                        if option_type == "metabinner_env":
                            metabinner_env = option_value
                            if metabinner_env in ["None", "none"]:
                                metabinner_env = ""
                        if option_type == "comebin_env":
                            comebin_env = option_value
                            if comebin_env in ["None", "none"]:
                                comebin_env = ""
                        if option_type == "cdhit_env":
                            cdhit_env = option_value
                            if cdhit_env in ["None", "none"]:
                                cdhit_env = ""
                        if option_type == "genepred_env":
                            genepred_env = option_value
                            if genepred_env in ["None", "none"]:
                                genepred_env = ""
                        if option_type == "hmmer_env":
                            hmmer_env = option_value
                            if hmmer_env in ["None", "none"]:
                                hmmer_env = ""
                        if option_type == "diamond_env":
                            diamond_env = option_value
                            if diamond_env in ["None", "none"]:
                                diamond_env = ""
                        if option_type == "taxonkit_env":
                            taxonkit_env = option_value
                            if taxonkit_env in ["None", "none"]:
                                taxonkit_env = ""
                        if option_type == "phobius_env":
                            phobius_env = option_value
                            if phobius_env in ["None", "none"]:
                                phobius_env = ""
                        if option_type == "bowtie_env":
                            bowtie_env = option_value
                            if bowtie_env in ["None", "none"]:
                                bowtie_env = ""
                        # Tool paths
                        if option_type == "conda_bin":
                            ana_dir_path = option_value
                        if option_type == "conda_sh":
                            ana_sh_path = option_value
                        if option_type == "prefetch_path":
                            prefetch_path = option_value
                        if option_type == "vdb_validate_path":
                            vdb_validate_path = option_value
                        if option_type == "fastq_dump_path":
                            fastq_dump_path = option_value
                        if option_type == "fastqc_path":
                            fastqc_path = option_value
                        if option_type == "gzip_path":
                            gzip_path = option_value
                        if option_type == "cat_path":
                            cat_path = option_value
                        if option_type == "bbduk_path":
                            bbduk_path = option_value
                        if option_type == "megahit_path":
                            megahit_path = option_value
                        if option_type == "kraken_path":
                            kraken_path = option_value
                        if option_type == "metabinner_bin_path":
                            metabinner_bin_path = option_value
                        if option_type == "comebin_bin_path":
                            comebin_bin_path = option_value
                        if option_type == "cd_hit_path":
                            cd_hit_path = option_value
                        if option_type == "fraggenescanrs_path":
                            fraggenescanrs_path = option_value
                        if option_type == "hmmscan_path":
                            hmmscan_path = option_value
                        if option_type == "hmmpress_path":
                            hmmpress_path = option_value
                        if option_type == "hmmfetch_path":
                            hmmfetch_path = option_value
                        if option_type == "diamond_path":
                            diamond_path = option_value
                        if option_type == "taxonkit_path":
                            taxonkit_path = option_value
                        if option_type == "phobius_path":
                            phobius_path = option_value
                        if option_type == "bowtie_build_path":
                            bowtie_build_path = option_value
                        if option_type == "bowtie_path":
                            bowtie_path = option_value

    # Determine the filtering threads.
    if pdf_threads is None:
        pdf_threads = thread_num

    # The path to the anaconda bin folder is mandatory.
    if ((not ana_dir_path) or (ana_dir_path == "") or (not os.path.exists(ana_dir_path))) and ((not ana_sh_path) or (ana_sh_path == "") or (not os.path.exists(ana_sh_path))):
        print("\nThe path to the anaconda installation directory and the path to the conda.sh file were not found. At least one must be set. Exiting.")
        exit()
    if ana_sh_path:
        conda_sh_path = ana_sh_path
    else:
        conda_sh_path = "{}/etc/profile.d/conda.sh".format(ana_dir_path)

    # If an output path has not been provided then a random string of 10 characters in length is generated again and again, until
    # the output path does not correspond to an existing folder.
    output_path_suffix = None
    if (not output_path) or output_path == "":
        random_str_length = 10
        output_path_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=random_str_length))
        output_path = "results_{}".format(output_path_suffix)
        while(os.path.exists(output_path)):
            output_path_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=random_str_length))
            output_path = "results_{}".format(output_path_suffix)

    print("\nOutput path: {}".format(output_path))
    if (not os.path.exists(output_path)) and (after_ap or after_tm or after_trimming or after_alignment or after_gene_pred or after_binning or after_db):
        print("\nThe output folder does not exist and the processes have been set to be continued based on an already existing analysis.")
        exit()

    # Protein input
    if protein_input:
        paired_end = False
        compressed = False

    # Binning tool name
    if binning_tool == 1:
        binning_tool_name = "metabinner"
    else:
        binning_tool_name = "comebin"

    # Input paths and output paths.
    output_path_summaries_errors = "{}/summaries_errors".format(output_path)
    output_path_fastqc = "{}/fastqc_results".format(output_path)
    output_path_trimmed = "{}/trimmed_results".format(output_path)
    output_path_megahit = "{}/megahit_results".format(output_path)
    output_path_conitgs = "{}/contigs".format(output_path)
    if binning_tool == 1:
        output_path_bin = "{}/metabinner_results".format(output_path)
    elif binning_tool == 2:
        output_path_bin = "{}/comebin_results".format(output_path)
    else:
        print("\nWrong selection for the binning tool. Changing to the default selection (1) for the binning tool.")
        binning_tool == 1
        output_path_bin = "{}/metabinner_results".format(output_path)
    output_path_genepred = "{}/gene_results".format(output_path)
    output_path_cdhit = "{}/cd_hit".format(output_path)
    output_path_blastp = "{}/blastp_results".format(output_path)
    output_path_hmmer = "{}/hmmer_results".format(output_path)
    output_path_topology = "{}/topology_results".format(output_path)
    output_path_motifs = "{}/motifs_results".format(output_path)
    output_path_bowtie = "{}/bowtie_results".format(output_path)
    output_path_kraken = "{}/kraken_results".format(output_path)
    output_path_annotation = "{}/annotation_results".format(output_path)
    # SRA file paths
    sra_folder = "sra_files"
    sra_run_folder = "{}/{}_files".format(sra_folder, sra_code)
    sra_file = "{}/{}/{}.sra".format(sra_run_folder, sra_code, sra_code)
    fastq_folder = "{}/{}_fastq".format(sra_run_folder, sra_code)
    fastq_single_end = "{}/{}_fastq_single_end".format(sra_run_folder, sra_code)
    # Enzyme-Profile/Domain file paths
    enzyme_domains_folder = "profile_protein_dbs"
    fpr_db_gen_folder = "{}/filtered_protein_dbs".format(enzyme_domains_folder)
    fam_nums_file_name = "{}/prfamilies_numbered.tsv".format(enzyme_domains_folder)
    fam_pfam_file_name = "{}/prfamilies_pfamdomains.tsv".format(enzyme_domains_folder)
    pfam_domains_names = "{}/pfam_accs_names.tsv".format(enzyme_domains_folder)
    hmmer_profile_lengths_file = "{}/profiles_lengths.tsv".format(enzyme_domains_folder)
    # Adapters path
    adapters_ap_path = "{}/identified_adapters.fa".format(output_path)
    file_read_info_path = "{}/read_information.txt".format(output_path)
    # Contigs paths
    output_megahit_contigs = "{}/megahit_contigs/final.contigs.fa".format(output_path_megahit)
    output_final_contigs = "{}/contigs.fa".format(output_path_conitgs)
    output_final_contigs_formated = "{}/contigs_formated.fna".format(output_path_conitgs)
    output_final_contigs_formated_name = "contigs_formated.fna"
    bowtie_contigs_basename = "{}/contigs_formated_bwtb".format(output_path_conitgs)
    # Binning paths.
    # Get the path of the current working directory.
    enz_dir = os.getcwd()
    # Binning.
    relpath_contigs = "{}/{}".format(output_path_conitgs, output_final_contigs_formated_name)
    fullpath_contigs_formated = os.path.abspath(relpath_contigs)
    output_fullpath_trimmed = os.path.abspath(output_path_trimmed)
    fullpath_tr_fastq_files = "\"{}/\"*fastq".format(output_fullpath_trimmed)
    full_binning_folder = os.path.abspath(output_path_bin)
    relpath_bin_results_folder = "{}/bins".format(output_path_bin)
    full_bin_results_folder = os.path.abspath(relpath_bin_results_folder)
    full_coverage_folder = "{}/coverages".format(full_binning_folder)
    bin_bam_folder = "{}/bam_results".format(full_binning_folder)
    if binning_tool == 1:
        binning_results_path = "{}/metabinner_res/metabinner_result.tsv".format(full_bin_results_folder)
    else:
        binning_results_path = "{}/comebin_res/comebin_res.tsv".format(full_bin_results_folder)
    full_coverage_profile_path = "{}/coverage_profile_f{}.tsv".format(full_coverage_folder, bin_num_contig_len)
    if kraken_mode:
        binphylo_path_prsids = "{}/binphylo_prsids.tsv".format(output_path_kraken)
    else:
        binphylo_path_prsids = "{}/binphylo_prsids.tsv".format(output_path_bin)
    binphylo_path_prstax = "{}/binphylo_prstax.tsv".format(output_path_bin)
    binphylo_path_binstax = "{}/binphylo_binstax.tsv".format(output_path_bin)
    binphylo_path_binstax_max = "{}/binphylo_binstax_max.tsv".format(output_path_bin)
    binphylo_path_binstax_names = "{}/binphylo_binstax_names.txt".format(output_path_bin)
    binphylo_path_binstax_max_names = "{}/binphylo_binstax_max_names.txt".format(output_path_bin)
    binphylo_freq_taxids_path = "{}/binphylo_freq_taxids.tsv".format(output_path_bin)
    binphylo_maxfreq_taxids_path = "{}/binphylo_maxfreq_taxids.tsv".format(output_path_bin)
    freq_taxids_path = "{}/freq_taxids.txt".format(output_path_bin)
    maxfreq_taxids_path = "{}/maxfreq_taxids.txt".format(output_path_bin)
    freq_lineage_path = "{}/freq_lineages.tsv".format(output_path_bin)
    maxfreq_lineage_path = "{}/maxfreq_lineages.tsv".format(output_path_bin)
    freq_lineage_form_path = "{}/freq_lineages_form.txt".format(output_path_bin)
    maxfreq_lineage_form_path = "{}/maxfreq_lineages_form.txt".format(output_path_bin)
    family_profile_path = "{}/family_profile_frequencies.tsv".format(output_path_bin)
    bin_summary_info_path = "{}/b_summary_info_{}.tsv".format(output_path_bin, binning_tool_name)
    contig_read_summary_path = "{}/cr_summary_{}.tsv".format(output_path_bin, binning_tool_name)
    binner_bin_info_path = "{}/{}_bin_summary.tsv".format(output_path_bin, binning_tool_name)
    # FragGeneScan paths
    output_path_genepred_base = "{}/GP".format(output_path_genepred)
    output_path_genepred_faa = "{}.faa".format(output_path_genepred_base)
    output_path_genepred_ffn = "{}.ffn".format(output_path_genepred_base)
    output_fgs_protein_formatted_1 = "{}/gp_proteins_formatted.fasta".format(output_path_genepred)
    output_fgs_protein_formatted_2 = "{}/gc_proteins_formatted.fasta".format(output_path_genepred)
    gene_contig_dist_path = "{}/gene_contig_distance.txt".format(output_path_genepred)
    # CD_HIT paths
    cd_hit_results_path = "{}/cd_hit_results".format(output_path_cdhit)
    cd_hit_results_fasta_path = "{}/cd_hit_results_formatted.fasta".format(output_path_cdhit)
    # Hmmer paths
    hmmer_dmbl_results = "{}/hmmer_domtblout.txt".format(output_path_hmmer)
    hmmer_dmbl_results_phylo = "{}/hmmer_domtblout_phylo.txt".format(output_path_hmmer)
    hmmer_simple_results = "{}/hmmer_simple.txt".format(output_path_hmmer)
    hmmer_simple_results_phylo = "{}/hmmer_simple_phylo.txt".format(output_path_hmmer)
    hmmer_enz_domains_all_proteins = "{}/hmmer_enz_domains_all_proteins_info.tsv".format(output_path_hmmer)
    hmmer_enz_domains_all_proteins_phylo = "{}/hmmer_enz_domains_all_proteins_info_phylo.tsv".format(output_path_hmmer)
    file_prs_seq_enz_domains_name = "{}/protein_enz_domains.fasta".format(output_path_hmmer)
    file_prs_seq_enz_domains_phylo_name = "{}/protein_enz_domains_phylo.fasta".format(output_path_hmmer)
    file_prs_seq_no_enzs_name = "{}/proteins_no_enz_domains.fasta".format(output_path_hmmer)
    comb_dmbl_results = "{}/comb_domtblout.txt".format(output_path_hmmer)
    comb_simple_results = "{}/comb_simple.txt".format(output_path_hmmer)
    # File with the information for all the domains found in the combined proteins.
    comb_all_domains_proteins = "{}/comb_all_domains_info.tsv".format(output_path_hmmer)
    # Hmmer Profiles paths
    hmmer_common_lengths_file = "{}/profile_similarity_perc.tsv".format(output_path_hmmer)
    # Motifs paths
    input_motifs_results = "{}/input_motifs_output.tsv".format(output_path_motifs)
    # Blastp paths
    # Blastp - SwissProt
    blastp_results_swissprot_file = "{}/blastp_results_swissprot.tsv".format(output_path_blastp)
    blastp_results_swissprot_phylo_file = "{}/blastp_results_phylo_swissprot.tsv".format(output_path_blastp)
    blastp_info_swissprot_file = "{}/blastp_swissprot_info.tsv".format(output_path_blastp)
    blastp_info_swissprot_phylo_file = "{}/blastp_phylo_swissprot_info.tsv".format(output_path_blastp)
    # Blastp - Domains - Nr
    blastp_doms_nr_file = "{}/blastp_doms_nr.tsv".format(output_path_blastp)
    blastp_doms_nr_phylo_file = "{}/blastp_doms_nr_phylo.tsv".format(output_path_blastp)
    blastp_doms_info_nr_file = "{}/blastp_doms_nr_info.tsv".format(output_path_blastp)
    blastp_doms_info_nr_phylo_file = "{}/blastp_doms_nr_phylo_info.tsv".format(output_path_blastp)
    # Blastp - No domains - Nr
    blastp_results_no_doms_nr_file = "{}/blastp_results_no_doms_nr.tsv".format(output_path_blastp)
    blastp_info_no_doms_nr_file = "{}/blastp_no_doms_nr_info.tsv".format(output_path_blastp)
    blastp_no_doms_below_threshold = "{}/blastp_no_doms_nr_below_threshold.fasta".format(output_path_blastp)
    blastp_info_no_doms_below_threshold = "{}/blastp_no_doms_nr_below_threshold_info.tsv".format(output_path_blastp)
    # Blastp - Combined Info - Nr
    blastp_info_comb_nr_file = "{}/blastp_comb_nr_info.tsv".format(output_path_blastp)
    # Proteins found from the two methods combined in a file. This file is located in the HMMER folder, because the
    # next analysis is based on findind all domains of those proteins.
    proteins_combined_file = "{}/proteins_combined.fasta".format(output_path_hmmer)
    # Gene paths
    gene_info_file = "{}/gene_info.tsv".format(output_path_genepred)
    # The path for the file below must change to "PrFamilies_Length.txt". The information about the family of the Swiss-Prot protein which is the first hit of a putative protein
    # against the Swiss-Prot database with the mean and median length of that family should be drawn for this new file instead of the one below. Then a comparison should be made between
    # this fmaily and the one selected at the beginning. A score should be set for the match or mismatch of these familes.
    pr_fams_path = "{}/prfamilies_length.tsv".format(enzyme_domains_folder)
    family_info_path = "{}/family_pred_info.tsv".format(output_path_blastp)
    # Bowtie2 analysis
    bowtie_single_unaligned_path = "{}/unaligned_singe_end.fastq".format(output_path_bowtie)
    bowtie_paired_unaligned_con_path = "{}/unaligned_con_paired_end.fastq".format(output_path_bowtie)
    bowtie_stats_path = "{}/bowtie_stats.txt".format(output_path_bowtie)
    mapped_reads_path = "{}/mapped_reads.sam".format(output_path_bowtie)
    # kraken2
    kraken_results_path = "{}/taxon_results.tsv".format(output_path_kraken)
    kraken_report_path = "{}/report_info.tsv".format(output_path_kraken)
    kraken_species_path = "{}/kraken_species.tsv".format(output_path_kraken)
    kraken_species_thr_path = "{}/kraken_species_thr".format(output_path_kraken)
    kraken_reads_path = "{}/kraken_reads.tsv".format(output_path_kraken)
    kraken_taxname_path = "{}/kraken_taxa_names.tsv".format(output_path_kraken)
    binned_ctb_path = "{}/contigs_to_bins.tsv".format(output_path_kraken)
    binned_btc_path = "{}/bins_to_contigs.tsv".format(output_path_kraken)
    binned_taxa_path = "{}/binned_taxa.txt".format(output_path_kraken)
    kraken_bin_info_path = "{}/kraken_bin_summary.txt".format(output_path_kraken)
    kraken_filters_path = "{}/kraken_filters.txt".format(output_path_kraken)
    # Phobius paths
    phobius_input_file_name = "{}/phobius_input.txt".format(output_path_topology)
    phobius_output_file_name = "{}/phobius_results.txt".format(output_path_topology)
    # Phobius information file path
    topology_info_path = "{}/topology_info.tsv".format(output_path_topology)
    # Annotation paths
    if output_path_suffix is not None:
        annotation_file_txt_name = ("{}/annotation_info_{}.txt".format(output_path_annotation, output_path_suffix))
        annotation_file_excel_name = ("{}/annotation_info_{}.xlsx".format(output_path_annotation, output_path_suffix))
    else:
        annotation_file_txt_name = ("{}/annotation_info.txt".format(output_path_annotation))
        annotation_file_excel_name = ("{}/annotation_info.xlsx".format(output_path_annotation))
    # Bash scripts
    diamond_db_bash_name = "diamond_db.sh"
    sra_bash_script = "{}/sra_run.sh".format(sra_run_folder)
    fastqc_1_bash_script = "{}/fastqc_initial.sh".format(output_path_fastqc)
    fastqc_2_bash_script = "{}/fastqc_trimmed.sh".format(output_path_fastqc)
    bbduk_bash_script = "{}/bbduk.sh".format(output_path_trimmed)
    megahit_bash_script = "{}/megahit.sh".format(output_path_megahit)
    metabinner_bash_script = "{}/metabinner.sh".format(output_path_bin)
    comebin_bash_script = "{}/comebin.sh".format(full_binning_folder, output_path_bin)
    gene_bash_script = "{}/genepred.sh".format(output_path_genepred)
    cdhit_bash_script = "{}/cdhit.sh".format(output_path_cdhit)
    hmmer_spec_bash_script = "{}/hmmer_spec.sh".format(output_path_hmmer)
    hmmer_spec_phylo_bash_script = "{}/hmmer_spec_phylo.sh".format(output_path_hmmer)
    blastp_nodoms_script = "{}/blast_nr_nodoms.sh".format(output_path_blastp)
    hmmer_broad_bash_script = "{}/hmmer_broad.sh".format(output_path_hmmer)
    blast_swissprot_bash_script = "{}/dimaond_blast_swissprot.sh".format(output_path_blastp)
    blast_swissprot_phylo_bash_script = "{}/diamond_blast_swissprot_phylo.sh".format(output_path_blastp)
    blast_nr_bash_script = "{}/diamond_blast_nr_doms.sh".format(output_path_blastp)
    blast_nr_phylo_bash_script = "{}/diamond_blast_nr_phylo.sh".format(output_path_blastp)
    tax_freq_bash_script = "{}/taxonkit_freq_names.sh".format(output_path_bin)
    tax_maxfreq_bash_script = "{}/taxonkit_maxfreq_names.sh".format(output_path_bin)
    taxonkit_freq_line_bash_script = "{}/taxonkit_freq_lineages.sh".format(output_path_bin)
    taxonkit_maxfreq_line_bash_script = "{}/taxonkit_maxfreq_lineages.sh".format(output_path_bin)
    kraken_bash_script = "{}/kraken.sh".format(output_path_kraken)
    phobius_bash_script = "{}/phobius.sh".format(output_path_topology)
    bowtie_bash_script = "{}/bowtie.sh".format(output_path_bowtie)
    # TXT stdout and stderr
    fastqc_stdoe_bt_path = "{}/fastqc_bt_stdoe.txt".format(output_path_fastqc)
    fastqc_stdoe_at_path = "{}/fastqc_at_stdoe.txt".format(output_path_fastqc)
    bbduk_stdoe_path = "{}/bbduk_stdoe.txt".format(output_path_trimmed)
    megahit_stdoe_path = "{}/megahit_stdoe.txt".format(output_path_megahit)
    fraggenescanrs_stdoe_path = "{}/fraggenescanrs_stdoe.txt".format(output_path_genepred)
    cd_hit_stdoe_path = "{}/cd_hit_stdoe.txt".format(output_path_cdhit)
    hmmscan_spec_stdoe_path = "{}/hmmscan_spec_stdoe.txt".format(output_path_hmmer)
    hmmscan_spec_phylo_stdoe_path = "{}/hmmscan_spec_phylo_stdoe.txt".format(output_path_hmmer)
    diamond_blastp_nr1_stdoe_path = "{}/blastp_nr1_stdoe.txt".format(output_path_blastp)
    hmmscan_broad_stdoe_path = "{}/hmmscan_broad_stdoe.txt".format(output_path_hmmer)
    diamond_blastp_swissprot_stdoe_path = "{}/blastp_swissprot_stdoe.txt".format(output_path_blastp)
    diamond_blastp_nr2_stdoe_path = "{}/blastp_nr2_stdoe.txt".format(output_path_blastp)
    diamond_blastp_swissprot_phylo_stdoe_path = "{}/blastp_swissprot_phylo_stdoe.txt".format(output_path_blastp)
    diamond_blastp_nr2_phylo_stdoe_path = "{}/blastp_nr2_phylo_stdoe.txt".format(output_path_blastp)
    bin_gen_coverage_stdoe_path = "{}/gen_coverage_stdoe.txt".format(full_binning_folder)
    bin_gen_kmer_stdoe_path = "{}/gen_kmer_stdoe.txt".format(full_binning_folder)
    bin_filter_tooshort_stdoe_path = "{}/filter_tooshort_stdoe.txt".format(full_binning_folder)
    binning_stdoe_path = "{}/binning_stdoe.txt".format(full_binning_folder)
    taxoknit_freq_stdoe_path = "{}/taxonkit_freq_taxids_stdoe.txt".format(output_path_bin)
    taxoknit_maxfreq_stdoe_path = "{}/taxonkit_maxfreq_taxids_stdoe.txt".format(output_path_bin)
    taxonkit_freq_line_stdoe_path = "{}/taxonkit_freq_lineage_stdoe.txt".format(output_path_bin)
    taxonkit_maxfreq_line_stdoe_path = "{}/taxonkit_maxfreq_lineage_stdoe.txt".format(output_path_bin)
    csvtk_freq_stdoe_path = "{}/csvtk_freq_stdoe.txt".format(output_path_bin)
    csvtk_maxfreq_stdoe_path = "{}/csvtk_maxfreq_stdoe.txt".format(output_path_bin)
    phobius_stde_path = "{}/phobius_stdoe.txt".format(output_path_topology)
    bowtie_build_stdoe_path = "{}/bowtie_build_stdoe.txt".format(output_path_bowtie)
    bowtie_stdoe_path = "{}/bowtie_stdoe.txt".format(output_path_bowtie)
    kraken_stde_path = "{}/kraken_stde.txt".format(output_path_kraken)
    # TXT files for versions
    fastqc_version_bt_path = "{}/fastqc_bt_version.txt".format(output_path_fastqc)
    fastqc_version_at_path = "{}/fastqc_at_version.txt".format(output_path_fastqc)
    bbduk_version_path = "{}/bbduk_version.txt".format(output_path_trimmed)
    megahit_version_path = "{}/megahit_version.txt".format(output_path_megahit)
    fraggenescanrs_version_path = "{}/fraggenescanrs_version.txt".format(output_path_genepred)
    cd_hit_version_path = "{}/cd_hit_version.txt".format(output_path_cdhit)
    hmmscan_spec_version_path = "{}/hmmscan_spec_version.txt".format(output_path_hmmer)
    hmmscan_spec_phylo_version_path = "{}/hmmscan_spec_phylo_version.txt".format(output_path_hmmer)
    hmmscan_broad_version_path = "{}/hmmscan_broad_version.txt".format(output_path_hmmer)
    taxoknit_freq_version_path = "{}/taxonkit_freq_taxids_version.txt".format(output_path_bin)
    taxoknit_maxfreq_version_path = "{}/taxonkit_maxfreq_taxids_version.txt".format(output_path_bin)
    taxonkit_freq_line_version_path = "{}/taxonkit_freq_lineage_version.txt".format(output_path_bin)
    taxonkit_maxfreq_line_version_path = "{}/taxonkit_maxfreq_lineage_version.txt".format(output_path_bin)
    csvtk_freq_version_path = "{}/csvtk_freq_version.txt".format(output_path_bin)
    csvtk_maxfreq_version_path = "{}/csvtk_maxfreq_version.txt".format(output_path_bin)
    phobius_version_path = "{}/phobius_version.txt".format(output_path_topology)
    bowtie_build_version_path = "{}/bowtie_build_version.txt".format(output_path_bowtie)
    bowtie_version_path = "{}/bowtie_version.txt".format(output_path_bowtie)
    kraken_version_path = "{}/kraken_version.txt".format(output_path_kraken)
    # Time
    time_analyis_path = "{}/time_analysis.tsv".format(output_path)

    # Initialization of variables.
    enzyme_names_dict = {}
    dict_contigs_bins = {}
    dict_prs_bins = {}
    prs_with_enz_domains = []
    prs_with_enz_domains_phylo = []
    prs_blast_thr = []
    protein_ids_below_thr = []
    analysis_fam_names = []
    dict_hm = {}
    dict_sp = {}
    dict_nr = {}
    dict_top = {}
    bins_prs_tax_dict = {}
    bin_tax_dict = {}
    bin_tax_max_dict = {}
    bin_group_tax_dict = {}
    pr_tax_info_dcit = {}
    dict_bins_prs_sorted = {}
    dict_input_motifs = {}
    dict_genes = {}
    contig_gene_dist_dict = {}
    swiss_fams_len_comp_dict = {}
    reads_to_contigs_dict = {}
    contigs_to_reads_dict = {}
    kraken_species_dict = {}
    kraken_species_thr_dict = {}
    read_to_species_dict = {}
    taxid_to_species_dict = {}
    tr_ex_file_paths = []
    tr_ex_file_paths_p = {}
    dict_seqs = {}
    time_dict = {
        "tool_time": "None",
        "sra_time": "None",
        "dbs_time": "None",
        "fastqc_initial_time": "None",
        "preprocessing_time": "None",
        "fastqc_final_time": "None",
        "assembly_time": "None",
        "gene_prediction_time": "None",
        "gene_annotation_time": "None",
        "cd_hit_time": "None",
        "kraken_time": "None",
        "kraken_specific_time": "None",
        "binning_time": "None",
        "bowtie_time": "None",
        "hmmer_spec_time": "None",
        "hmmer_spec_taxonomy_time": "None",
        "blastp_fpd_no_doms_time": "None",
        "blastp_fpd_swiss_doms_time": "None",
        "blastp_fpd_swiss_taxonomy_time": "None",
        "bin_analysis_cm_time": "None",
        "bin_taxonomy_cm_time": "None",
        "kraken_binning_time": "None",
        "hmmer_broad": "None",
        "topology_time": "None",
        "motifs_time": "None",
        "family_prediction_time": "None",
        "info_collection_time": "None",
        "results_time": "None"
    }
    if input_protein_names is None:
        input_protein_names = []
    if input_protein_names_phylo is None:
        input_protein_names_phylo = []

    # Initialize dictionaries for the different kinds of information.
    # Determine whether the output folder will be overwritten.
    if (not after_trimming) and (not after_alignment) and (not after_gene_pred) and (not protein_input) and (not after_binning) and (not after_db) and (not after_tm) and (not after_ap):
        # The output folder is created. If it already exists then it is deleted and recreated.
        if os.path.exists(output_path):
            shutil.rmtree(output_path)
        os.mkdir(output_path)
        # The output folder of the summary and error files.
        if os.path.exists(output_path_summaries_errors):
            shutil.rmtree(output_path_summaries_errors)
        os.mkdir(output_path_summaries_errors)
    elif protein_input:
        # Create the folder for the results and a folder for the CD-HIT results
        # Copy the input file in the CD-HIT results as the formatted file that would be generated by CD-HIT
        # The output folder is created. If it already exists then it is deleted and recreated.
        if os.path.exists(output_path):
            shutil.rmtree(output_path)
        os.mkdir(output_path)
        # The output_path_summaries_errors is only needed for trimming, thus not for steps after CD-HIT.
        # Create the CD-HIT folder
        if os.path.exists(output_path_cdhit):
            shutil.rmtree(output_path_cdhit)
        os.mkdir(output_path_cdhit)
    
    # Create the input and output log files in the output folder.
    input_log_file = open("{}/log_input.txt".format(output_path), "a+")
    output_log_file = open("{}/log_output.txt".format(output_path), "a+")
    input_log_file.write("Input command:\n")
    input_log_file.write("{}\n\n".format(input_command))

    # Convert the input kraken thresholds to a list of kraken thresholds.
    if kraken_mode:
        if "," in kraken_threshold:
            kraken_threshold = kraken_threshold.split(",")
        else:
            kraken_threshold = [kraken_threshold]

    # If no family code and no database name have been provided then only phylogenetic analysi will be performed.
    # If no family code phylo and no database phylo have been provided then only search for a specific protein famlily is performed.
    if (family_code is None) and (not db_name):
        only_phylo_mode = True
        family_group_name = None
    elif (family_code is not None) and db_name:
        only_phylo_mode = False
        if "," in family_code:
            family_code = family_code.split(",")
        else:
            family_code = [family_code]
        family_group_name = db_name
    elif (family_code is not None):
        only_phylo_mode = False
        if "," in family_code:
            family_code = family_code.split(",")
        else:
            family_code = [family_code]
        family_group_name = "_".join(family_code)
    elif db_name:
        only_phylo_mode = False
        family_group_name = db_name
    # Databases for phylogenetic analysis.
    if (family_code_phylo is None) and (not db_name_phylo):
        only_seek_mode = True
        family_group_name_phylo = None
    elif (family_code_phylo is not None) and db_name_phylo:
        only_seek_mode = False
        if "," in family_code_phylo:
            family_code_phylo = family_code_phylo.split(",")
        else:
            family_code_phylo = [family_code_phylo]
        family_group_name_phylo = db_name_phylo
    elif (family_code_phylo is not None):
        only_seek_mode = False
        if "," in family_code_phylo:
            family_code_phylo = family_code_phylo.split(",")
        else:
            family_code_phylo = [family_code_phylo]
        family_group_name_phylo = "_".join(family_code_phylo)
    elif db_name_phylo:
        only_seek_mode = False
        family_group_name_phylo = db_name_phylo
    
    # Determine which route is used for a taxonomic analysis, if any.
    # 1: Protein families + Binning from MetaBinner -> Species for each bin, Bowtie2 -> Abundance of each bin/species group
    # 2: Protein families + Binning from COMEBin -> Species for each bin, Bowtie2 -> Abundance of each bin/species group
    # 3: Kraken analysis -> A species to each read, Abundance of each species found -> Removing species of low read cound -> Binning the contigs based on their reads' species assignment
    if (not only_seek_mode) and kraken_mode:
        print("\nTaxonomy analysis must be performed by either MetaBinner/COMEBin or kraken2. Protein family selection for taxonomy analysis and kraken2 analysis can not be combined. Exiting.")
        exit()
    if kraken_mode:
        only_seek_mode = False

    # SRA file collection, if an SRA code was selected.
    if sra_code:
        start_time_sra = time.time()
        paired_end, input_folder = collect_sra(sra_env, sra_folder, prefetch_size, sra_run_folder, sra_file, fastq_folder, fastq_single_end, sra_code, input_log_file, output_log_file, prefetch_path, vdb_validate_path, fastq_dump_path, sra_bash_script, conda_sh_path)
        label = "Process of SRA processing:"
        elpased_time = end_time_analysis(label, start_time_sra, output_log_file)
        time_dict["sra_time"] = elpased_time
        compressed = True
        contigs = False
    
    # If chosen, the analysis stops here.
    if up_to_sra:
        label = "Process of SRA processing:"
        elpased_time = end_time_analysis(label, start_time, output_log_file)
        time_dict["tool_time"] = elpased_time
        write_time_dict(time_dict, time_analyis_path)
        exit()

    # Type and info provided to add in the annotation results.
    if add_type:
        if "\t" in add_type:
            add_type = add_type.split("\t")
        else:
            add_type = [add_type]
    if add_info:
        if "\t" in add_info:
            add_info = add_info.split("\t")
        else:
            add_info = [add_info]
    if add_type and add_info:
        if len(add_type) != len(add_info):
            print("\nThe types and information provided to add in the annotation results differ. Exiting.")
            exit()
    elif add_type and (not add_info):
        print("\nType(s) was(were) provided to be added in the annotation results but no information was provided. Exiting.")
        exit()
    elif add_info and (not add_type):
        print("\nInformation was provided to be added in the annotation results but no type(s) was(were) provided. Exiting.")
        exit()

    # Time for databases.
    # The pHMM database.
    # Determine the protein names.
    profiles_path, analysis_fam_names, family_to_profile_seek_dict = craete_phmm_db(family_code, family_group_name, enzyme_domains_folder, fam_nums_file_name, fam_pfam_file_name, pfam_domains_names, hmmer_env, hmmfetch_path, hmmpress_path, profiles_broad_path, input_log_file, output_log_file, conda_sh_path)
    fpr_db_fasta, fpr_db_name, fpr_db_folder, enzyme_names_dict = find_fam_names(family_code, family_group_name, create_nr_db_status, protein_db_path, fpr_db_gen_folder, fam_nums_file_name, name_thr, input_protein_names_status, input_protein_names, enzyme_names_dict, conda_sh_path)
    # The phylo pHMM database.
    # Determine the phylo protein names.
    profiles_phylo_path, analysis_fam_names_phylo, family_to_profile_phylo_dict = craete_phmm_db(family_code_phylo, family_group_name_phylo, enzyme_domains_folder, fam_nums_file_name, fam_pfam_file_name, pfam_domains_names, hmmer_env, hmmfetch_path, hmmpress_path, profiles_broad_path, input_log_file, output_log_file, conda_sh_path)
    fpr_db_fasta_phylo, fpr_db_name_phylo, pfpr_db_gen_folder_phylo_name, enzyme_names_dict = find_fam_names(family_code_phylo, family_group_name_phylo, create_nr_db_status, protein_db_path, fpr_db_gen_folder, fam_nums_file_name, name_thr, input_protein_names_phylo_status, input_protein_names_phylo, enzyme_names_dict, conda_sh_path)
    # Create nr phylo db
    start_time_dbs = time.time()
    create_nr_db(enzyme_names_dict, protein_db_path, diamond_db_bash_name, diamond_env, diamond_path, thread_num, pdf_threads, input_log_file, output_log_file, conda_sh_path)    
    # Time for databases
    label = "Process of creating the pHMM and fnr databases (including phylo):"
    elpased_time = end_time_analysis(label, start_time_dbs, output_log_file)
    time_dict["dbs_time"] = elpased_time

    # If chosen, the analysis stops here.
    if up_to_databases:
        label = "Databases creation:"
        elpased_time = end_time_analysis(label, start_time, output_log_file)
        time_dict["tool_time"] = elpased_time
        write_time_dict(time_dict, time_analyis_path)
        exit()

    # Determine the input file paths based on whether the input files are paired-end or not, only if no variable has
    # been set to indicate omitting early steps of the pipeline.
    if (not after_gene_pred) and (not after_binning) and (not after_db) and (not after_tm) and (not after_ap):
        file_paths, file_paths_p = file_paths_creation(input_folder, paired_end)
        # If file paths should be determined but no input files were found, then stop the pipeline.
        if (not file_paths) and (not file_paths_p):
            print("No input files were found.")
            exit()
        # If proteins have been given as input, then copy the input file as the file for CD-HIT.
        if protein_input:
            input_protein_file_path = file_paths[0]
            shutil.copyfile(input_protein_file_path, cd_hit_results_fasta_path)
    else:
        print("\nInput file check skipped.")

    # Folder for the contigs.
    if (not after_alignment) and (not after_gene_pred) and (not after_binning) and (not after_db) and (not after_tm) and (not after_ap):
        # The folder that contains the contigs right untill the point before the process of binning, in the case where binning is implemented. Otherwise, this folder contais the final contigs.
        if os.path.exists(output_path_conitgs):
            shutil.rmtree(output_path_conitgs)
        os.mkdir(output_path_conitgs)
    
    # Quality control
    if (not contigs) and (not after_alignment) and (not after_gene_pred) and (not after_binning) and (not after_db) and (not protein_input) and (not after_tm) and (not after_ap):
        if not after_trimming:
            # FastQC has no specific option for analyzing paired-end reads together. It analyzes them seperatly as it does with each single-end read.
            start_time_fastqc_initial = time.time()
            clear = True
            fastqc(fastqc_env, output_path_fastqc, file_paths, fastqc_path, clear, thread_num, input_log_file, output_log_file, fastqc_1_bash_script, fastqc_version_bt_path, fastqc_stdoe_bt_path, conda_sh_path)
            label = "Process of initial fastqc:"
            elpased_time = end_time_analysis(label, start_time_fastqc_initial, output_log_file)
            time_dict["fastqc_initial_time"] = elpased_time

            # Process the FastQC results.
            file_reads(output_path_fastqc, paired_end, file_paths, file_paths_p, file_read_info_path)

            # Fix the path to the proper adapters file based on the choice of the user.
            adapter_file(output_path_fastqc, adapters_status, adapters_path, adapters_ap_path)
            
            # BBDuk.
            start_time_trimming = time.time()
            bbduk(bbduk_env, output_path_trimmed, file_paths, adapters_ap_path, paired_end, file_paths_p, bbduk_path, thread_num, bbduk_max_ram, output_path_summaries_errors, input_log_file, output_log_file, bbduk_bash_script, bbduk_version_path, bbduk_stdoe_path, conda_sh_path)
            label = "Process of trimming:"
            elpased_time = end_time_analysis(label, start_time_trimming, output_log_file)
            time_dict["preprocessing_time"] = elpased_time

            # Delete input files.
            reduce_volume(clear_space, input_folder, output_path_trimmed, input_log_file, [1])
            # If chosen, the analysis stops here.
            if up_to_trimming_com:
                label = "Up to trimming (1):"
                elpased_time = end_time_analysis(label, start_time, output_log_file)
                time_dict["tool_time"] = elpased_time
                write_time_dict(time_dict, time_analyis_path)
                exit()
            
            # FastQC has no specific option for analyzing paired-end reads together. It analyzes them seperatly as it does with each single-end read.
            ca_file_paths = []
            pre_enz_file_paths = os.listdir(output_path_trimmed)
            for i in pre_enz_file_paths:
                if "fastq" in i:
                    local_file_path = "{}/{}".format(output_path_trimmed, i)
                    ca_file_paths.append(local_file_path)
            if not skip_fastqc:
                start_time_fastqc_final = time.time()
                clear = False
                fastqc(fastqc_env, output_path_fastqc, ca_file_paths, fastqc_path, clear, thread_num, input_log_file, output_log_file, fastqc_2_bash_script, fastqc_version_at_path, fastqc_stdoe_at_path, conda_sh_path)
                label = "Process of final fastqc:"
                elpased_time = end_time_analysis(label, start_time_fastqc_final, output_log_file)
                time_dict["fastqc_final_time"] = elpased_time

            # If the original files were compressed then the files which are generated as results by the trimming tool are also compressed.
            # Megahit accepts only non-compressed files. Therefore, if the files were originally compressed they must be uncompressed in order
            # to be used by Megahit.
            if compressed:
                unzip_files(ca_file_paths, None, False, input_log_file, output_log_file, gzip_path)
        
        # Non-compressed trimmed files are located.
        tr_ex_file_paths_p, tr_ex_file_paths = locate_nc_files(paired_end, output_path_trimmed)

        file_reads(output_path_fastqc, paired_end, tr_ex_file_paths, tr_ex_file_paths_p, file_read_info_path)

        # Delete compressed trimmed files (if any) before the alignment.
        reduce_volume(clear_space, input_folder, output_path_trimmed, input_log_file, [2])
        
        # If chosen, the analysis stops here.
        if up_to_trimming_uncom:
            label = "Up to trimming (2):"
            elpased_time = end_time_analysis(label, start_time, output_log_file)
            time_dict["tool_time"] = elpased_time
            write_time_dict(time_dict, time_analyis_path)
            exit()

        start_time_assembly = time.time()
        megahit(megahit_env, output_path_megahit, tr_ex_file_paths, paired_end, tr_ex_file_paths_p, megahit_path, k_list, thread_num, input_log_file, output_log_file, megahit_bash_script, megahit_version_path, megahit_stdoe_path, conda_sh_path)
        label = "Process of read assembly:"
        elpased_time = end_time_analysis(label, start_time_assembly, output_log_file)
        time_dict["assembly_time"] = elpased_time

        # The function below leads to the final contigs being saved in the proper folder (regardless of whether contigs were the initial input files or theyr were formed by reads).
        contig_formation(contigs, file_paths, output_final_contigs, output_megahit_contigs, output_final_contigs_formated, input_log_file, output_log_file, cat_path)
        
        # If chosen, the analysis stops here.
        if up_to_alignment:
            label = "Up to alignment:"
            elpased_time = end_time_analysis(label, start_time, output_log_file)
            time_dict["tool_time"] = elpased_time
            write_time_dict(time_dict, time_analyis_path)
            exit()

    if (not after_db) and (not protein_input) and (not after_binning) and (not after_tm) and (not after_ap):
        if not after_gene_pred:
            # Gene scanning.
            start_time_genepred = time.time()
            gene_prediction(genepred_env, output_path_genepred, output_path_genepred_base, output_path_genepred_faa, output_fgs_protein_formatted_1, fullpath_contigs_formated, fraggenescanrs_path, thread_num, input_log_file, output_log_file, gene_bash_script, fraggenescanrs_version_path, fraggenescanrs_stdoe_path, conda_sh_path)
            label = "Process of gene prediction:"
            elpased_time = end_time_analysis(label, start_time_genepred, output_log_file)
            time_dict["gene_prediction_time"] = elpased_time
            
        # Characterize the gene sequences of the putative proteins.
        if (not only_phylo_mode):
            start_time_geneann = time.time()
            gene_sequences_dict, dict_genes, contig_gene_dist_dict = gene_annotation(output_path_genepred, output_path_genepred_ffn, output_fgs_protein_formatted_2, output_final_contigs_formated, gene_contig_dist_path, gene_info_file, genetic_code)
            label = "Process of gene annotation:"
            elpased_time = end_time_analysis(label, start_time_geneann, output_log_file)
            time_dict["gene_annotation_time"] = elpased_time

        # Protein redundancy.
        start_time_cd_hit = time.time()
        cd_hit(cdhit_env, output_path_cdhit, prs_source, output_fgs_protein_formatted_1, output_fgs_protein_formatted_2, cd_hit_results_path, cd_hit_results_fasta_path, cd_hit_t, cd_hit_path, cd_hit_mem, thread_num, input_log_file, output_log_file, cdhit_bash_script, cd_hit_version_path, cd_hit_stdoe_path, conda_sh_path)
        label = "Process of cd-hit:"
        elpased_time = end_time_analysis(label, start_time_cd_hit, output_log_file)
        time_dict["cd_hit_time"] = elpased_time

    if (not after_db) and (not protein_input) and (not after_tm) and (not after_ap):
        if not after_binning:
            # Non-compressed trimmed files are located.
            tr_ex_file_paths_p, tr_ex_file_paths = locate_nc_files(paired_end, output_path_trimmed)

            # Taxonmic analysis from kraken2.
            if (not only_seek_mode) and kraken_mode:
                start_time_kraken = time.time()
                kraken_status = True
                kraken_species_dict, kraken_species_thr_dict, read_to_species_dict, taxid_to_species_dict, time_dict = kraken(paired_end, tr_ex_file_paths_p, tr_ex_file_paths, output_path_kraken, kraken_db_path, kraken_results_path, kraken_report_path, kraken_threshold, kraken_species_path, kraken_species_thr_path, kraken_reads_path, kraken_taxname_path, conda_sh_path, kraken_env, kraken_path, kraken_bash_script, kraken_stde_path, kraken_version_path, kraken_status, kraken_memory_mapping, thread_num, time_dict, kraken_filters_path, input_log_file, output_log_file)
                label = "Process of kraken2:"
                elpased_time = end_time_analysis(label, start_time_kraken, output_log_file)
                time_dict["kraken_time"] = elpased_time

            # If selected and if FastQC were given as input then run MetaBinner2
            if (not only_seek_mode) and (not kraken_mode):
                start_time_binning = time.time()
                binning(binning_tool, metabinner_env, comebin_env, contigs, protein_input, output_path_bin, metabinner_bin_path, comebin_bin_path, fullpath_contigs_formated, bin_bam_folder, enz_dir, output_path_conitgs, full_coverage_folder, fullpath_tr_fastq_files, full_bin_results_folder, full_coverage_profile_path, thread_num, input_log_file, output_log_file, metabinner_bash_script, comebin_bash_script, bin_ram_ammount, bin_num_contig_len, bin_num_kmer, bin_gen_coverage_stdoe_path, bin_gen_kmer_stdoe_path, bin_filter_tooshort_stdoe_path, binning_stdoe_path, comebin_batch_size, conda_sh_path)
                label = "Process of binning:"
                elpased_time = end_time_analysis(label, start_time_binning, output_log_file)
                time_dict["binning_time"] = elpased_time
        
        # Bowtie
        if not only_seek_mode:
            # Non-compressed trimmed files are located.
            tr_ex_file_paths_p, tr_ex_file_paths = locate_nc_files(paired_end, output_path_trimmed)
                    
            # Mapping reads to contigs.
            start_time_bowtie = time.time()
            bowtie_status = True
            reads_to_contigs_dict, contigs_to_reads_dict = bowtie(output_path_bowtie, paired_end, tr_ex_file_paths, tr_ex_file_paths_p, conda_sh_path, bowtie_build_path, bowtie_build_version_path, bowtie_build_stdoe_path, bowtie_env, bowtie_bash_script, bowtie_path, bowtie_version_path, bowtie_stdoe_path, bowtie_single_unaligned_path, bowtie_paired_unaligned_con_path, bowtie_stats_path, mapped_reads_path, thread_num, output_final_contigs_formated, bowtie_contigs_basename, bowtie_status, input_log_file, output_log_file)
            label = "Process of read mapping:"
            elpased_time = end_time_analysis(label, start_time_bowtie, output_log_file)
            time_dict["bowtie_time"] = elpased_time

    # Analysis modes.
    # 1: Proteins identified by HMMER with any of the specified domains.
    # 2: Proteins with e-score value below the set (or default) threshold from the results of Blastp against the
    # filtered nr adtabase.
    # 3: Proteins combined from both modes of analyis.
    if (not after_tm) and (not after_ap):
        if not after_db:
            if (seek_route == 1 or seek_route == 3):
                if not only_phylo_mode:
                    # HMMER Specified for selected protein family.
                    start_time_hmmer_spec = time.time()
                    phylo_analysis = False
                    prs_with_enz_domains = hmmer_spec(hmmer_env, output_path_hmmer, hmmer_dmbl_results, hmmer_simple_results, hmmer_enz_domains_all_proteins, cd_hit_results_path, hmmscan_path, profiles_path, val_type, file_prs_seq_enz_domains_name, thread_num, input_log_file, output_log_file, hmmer_spec_bash_script, phylo_analysis, hmmscan_spec_version_path, hmmscan_spec_stdoe_path, conda_sh_path)
                    label = "Process of specified HMMER:"
                    elpased_time = end_time_analysis(label, start_time_hmmer_spec, output_log_file)
                    time_dict["hmmer_spec_time"] = elpased_time

                # HMMER specified for phylogenetic analysis.
                if profiles_phylo_path:
                    start_time_hmmer_spec_phylo = time.time()
                    phylo_analysis = True
                    prs_with_enz_domains_phylo = hmmer_spec(hmmer_env, output_path_hmmer, hmmer_dmbl_results_phylo, hmmer_simple_results_phylo, hmmer_enz_domains_all_proteins_phylo, cd_hit_results_path, hmmscan_path, profiles_phylo_path, val_type, file_prs_seq_enz_domains_phylo_name, thread_num, input_log_file, output_log_file, hmmer_spec_phylo_bash_script, phylo_analysis, hmmscan_spec_phylo_version_path, hmmscan_spec_phylo_stdoe_path, conda_sh_path)
                    label = "Process of specified taxonomy HMMER:"
                    elpased_time = end_time_analysis(label, start_time_hmmer_spec_phylo, output_log_file)
                    time_dict["hmmer_spec_taxonomy_time"] = elpased_time

            if (seek_route == 2 or seek_route == 3) and (not only_phylo_mode):
                # Blast proteins with no domains from the selected protein family.
                start_time_nr_no_doms = time.time()
                prs_blast_thr, protein_ids_below_thr = analyze_no_enzs(diamond_env, output_path_hmmer, output_path_blastp, prs_with_enz_domains, blastp_results_no_doms_nr_file, blastp_info_no_doms_nr_file, blastp_no_doms_below_threshold, blastp_info_no_doms_below_threshold, cd_hit_results_path, file_prs_seq_no_enzs_name, fpr_db_fasta, fpr_db_name, e_value_nodom_thr, thread_num, input_log_file, output_log_file, diamond_path, blastp_nodoms_script, diamond_blastp_nr1_stdoe_path, conda_sh_path)
                label = "Process of BLASTP against fnr for proteins without domains:"
                elpased_time = end_time_analysis(label, start_time_nr_no_doms, output_log_file)
                time_dict["blastp_fpd_no_doms_time"] = elpased_time

            if (not only_phylo_mode):
                # Combine both the proteins from both lists in a file and process to their annotation.
                combine_predictions(file_prs_seq_enz_domains_name, blastp_no_doms_below_threshold, proteins_combined_file)

                # Counters for the number of proteins with hits against the seek profiles and for the proteins with not hits but with low e-value scores against the fpd.
                pr_doms_num = 0
                prs_nr_num = 0
                prs_nr_thr_num = 0
                if prs_with_enz_domains:
                    pr_doms_num = len(prs_with_enz_domains)
                if prs_blast_thr:
                    prs_nr_num = len(prs_blast_thr)
                if protein_ids_below_thr:
                    prs_nr_thr_num = len(protein_ids_below_thr)
                all_prs_num = pr_doms_num + prs_nr_thr_num
                print("\nNumber of proteins with specified domains: {}.".format(pr_doms_num))
                print("Number of proteins without specified domains: {}.".format(prs_nr_num))
                print("Number of proteins without specified domains and below threshold: {}.".format(prs_nr_thr_num))
                print("Number of all proteins for further procesing: {}.".format(all_prs_num))
                if all_prs_num == 0:
                    print("\nNo proteins were found for annotation.")
                    label = "Up to combining predictions (after the analysis of proteins without domains of interest):"
                    elpased_time = end_time_analysis(label, start_time, output_log_file)
                    time_dict["tool_time"] = elpased_time
                    write_time_dict(time_dict, time_analyis_path)
                    exit()

                # Domain coverages percentages.
                dom_percs(comb_all_domains_proteins, hmmer_profile_lengths_file, hmmer_common_lengths_file)
                
                # Blastp. The proteins coming from nr-threshold detection should not be blasted again against the nr database.
                # Combined proteins: SwissProt
                start_time_nr_swiss_doms = time.time()
                phylo_analysis = False
                dict_sp = blastp(diamond_env, output_path_blastp, blastp_results_swissprot_file, blastp_doms_nr_file, blastp_info_swissprot_file, blastp_doms_info_nr_file, file_prs_seq_enz_domains_name, proteins_combined_file, swissprot_path, fpr_db_fasta, fpr_db_name, thread_num, input_log_file, output_log_file, diamond_path, blast_swissprot_bash_script, blast_nr_bash_script, phylo_analysis, conda_sh_path, diamond_blastp_swissprot_stdoe_path, diamond_blastp_nr2_stdoe_path)
                label = "Process of BLASTP against fnr and Swiss-Prot for proteins with domains:"
                elpased_time = end_time_analysis(label, start_time_nr_swiss_doms, output_log_file)
                time_dict["blastp_fpd_swiss_doms_time"] = elpased_time

            # Blastp. The proteins of specifies HMMER for the phylogenetic analysis are blasted against the pnr protein database for the phylogenetic analysis.
            if profiles_phylo_path:
                start_time_nr_doms_phyo = time.time()
                phylo_analysis = True
                dict_sp_phylo = blastp(diamond_env, output_path_blastp, blastp_results_swissprot_phylo_file, blastp_doms_nr_phylo_file, blastp_info_swissprot_phylo_file, blastp_doms_info_nr_phylo_file, file_prs_seq_enz_domains_phylo_name, proteins_combined_file, swissprot_path, fpr_db_fasta_phylo, fpr_db_name_phylo, thread_num, input_log_file, output_log_file, diamond_path, blast_swissprot_phylo_bash_script, blast_nr_phylo_bash_script, phylo_analysis, conda_sh_path, diamond_blastp_swissprot_phylo_stdoe_path, diamond_blastp_nr2_phylo_stdoe_path)
                label = "Process of BLASTP against fnr and Swiss-Prot for proteins with taxonomy domains:"
                elpased_time = end_time_analysis(label, start_time_nr_doms_phyo, output_log_file)
                time_dict["blastp_fpd_swiss_taxonomy_time"] = elpased_time

            if (not only_phylo_mode):
                # Combine the results of blastp against the nr database from analysis modes 1 or/and 2.
                dict_nr = combine_nr(blastp_info_no_doms_below_threshold, blastp_doms_info_nr_file, blastp_info_comb_nr_file)

    if not after_ap:
        if not after_tm:
            if not only_seek_mode:
                if not kraken_mode:
                    # Taxonomy of bins
                    start_time_binpals = time.time()
                    dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict = bin_phylo(fpr_db_fasta_phylo, blastp_doms_info_nr_phylo_file, dict_contigs_bins, binning_results_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prstax, binphylo_path_binstax, binphylo_path_binstax_max, binphylo_path_prsids, binphylo_path_binstax_names, binphylo_path_binstax_max_names, binphylo_freq_taxids_path, binphylo_maxfreq_taxids_path, cd_hit_results_fasta_path, taxonkit_env, taxonkit_path, tax_freq_bash_script, tax_maxfreq_bash_script, taxoknit_freq_version_path, taxoknit_freq_stdoe_path, taxoknit_maxfreq_version_path, taxoknit_maxfreq_stdoe_path, thread_num, conda_sh_path, freq_taxids_path, maxfreq_taxids_path, taxonkit_freq_line_bash_script, taxonkit_maxfreq_line_bash_script, taxonkit_freq_line_version_path, taxonkit_freq_line_stdoe_path, taxonkit_maxfreq_line_version_path, taxonkit_maxfreq_line_stdoe_path, freq_lineage_path, maxfreq_lineage_path, freq_lineage_form_path, maxfreq_lineage_form_path, csvtk_freq_version_path, csvtk_freq_stdoe_path, csvtk_maxfreq_version_path, csvtk_maxfreq_stdoe_path, family_to_profile_phylo_dict, hmmer_enz_domains_all_proteins_phylo, family_profile_path, input_log_file, output_log_file)
                    label = "Process of bin analysis:"
                    elpased_time = end_time_analysis(label, start_time_binpals, output_log_file)
                    time_dict["bin_analysis_cm_time"] = elpased_time

                    # Analysis of the bins
                    start_time_metacome = time.time()
                    binning_analysis(mapped_reads_path, binning_results_path, file_read_info_path, reads_to_contigs_dict, contigs_to_reads_dict, output_final_contigs_formated, dict_contigs_bins, contig_read_summary_path, binphylo_path_binstax_max, binphylo_maxfreq_taxids_path, maxfreq_lineage_path, bin_summary_info_path, bin_group_tax_dict, binner_bin_info_path)
                    label = "Process of bin taxonomy:"
                    elpased_time = end_time_analysis(label, start_time_metacome, output_log_file)
                    time_dict["bin_taxonomy_cm_time"] = elpased_time
                    
                if kraken_mode:
                    start_time_kraken_binning = time.time()
                    bin_group_tax_dict = kraken_binning(kraken_mode, contigs_to_reads_dict, read_to_species_dict, taxid_to_species_dict, kraken_taxname_path, kraken_reads_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prsids, binned_ctb_path, binned_btc_path, binned_taxa_path, kraken_bin_info_path)
                    label = "Process of binning based on kraken:"
                    elpased_time = end_time_analysis(label, start_time_kraken_binning, output_log_file)
                    time_dict["kraken_binning_time"] = elpased_time

    if not only_phylo_mode:
        if not after_ap:
            if not after_tm:
                # HMMER Broad.
                start_time_hmmer_broad = time.time()
                prs_with_enz_domains, dict_hm = hmmer_broad(hmmer_env, prs_with_enz_domains, hmmer_dmbl_results, hmmer_simple_results, hmmer_enz_domains_all_proteins, proteins_combined_file, comb_dmbl_results, comb_simple_results, comb_all_domains_proteins, profiles_broad_path, hmmscan_path, val_type, second_dom_search, thread_num, input_log_file, output_log_file, hmmer_broad_bash_script, hmmscan_broad_version_path, hmmscan_broad_stdoe_path, conda_sh_path)
                label = "Process of broad HMMER:"
                elpased_time = end_time_analysis(label, start_time_hmmer_broad, output_log_file)
                time_dict["hmmer_broad"] = elpased_time
                
                # Topology predictions
                start_time_tm = time.time()
                dict_top = phobius(phobius_env, proteins_combined_file, output_path_topology, phobius_input_file_name, phobius_output_file_name, topology_info_path, phobius_path, input_log_file, output_log_file, phobius_bash_script, phobius_version_path, phobius_stde_path, conda_sh_path)
                label = "Process of topology prediction:"
                elpased_time = end_time_analysis(label, start_time_tm, output_log_file)
                time_dict["topology_time"] = elpased_time

            # Input motifs search.
            start_time_motifs = time.time()
            dict_input_motifs = input_motif_search(motifs_path, output_path_motifs, input_motifs_results, proteins_combined_file, input_log_file)
            label = "Process of motif search:"
            elpased_time = end_time_analysis(label, start_time_motifs, output_log_file)
            time_dict["motifs_time"] = elpased_time
            
            # Protein family type prediction and mean length comparison.
            start_time_swiss_fam_pred = time.time()
            swiss_fams_len_comp_dict, dict_seqs = pr_fam_pred(analysis_fam_names, blastp_info_swissprot_file, proteins_combined_file, pr_fams_path, family_info_path)
            label = "Process of Swiss-Prot family prediction:"
            elpased_time = end_time_analysis(label, start_time_swiss_fam_pred, output_log_file)
            time_dict["family_prediction_time"] = elpased_time

    # Collect information. Information is collected and not used directly from the functions above because if processes have been ommited, then it would be possible to gather
    # the information provided by already processes already run in previous analysis from the files generated and present in the folder which will hold the results of an analysis.
    start_time_info_collection = time.time()
    dict_seqs, dict_hm, dict_top, dict_sp, dict_nr, dict_genes, contig_gene_dist_dict, dict_input_motifs, swiss_fams_len_comp_dict, dict_contigs_bins, dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict = info_collection(comb_all_domains_proteins, proteins_combined_file, blastp_info_swissprot_file, blastp_info_comb_nr_file, topology_info_path, gene_info_file, gene_contig_dist_path, input_motifs_results, family_info_path, binning_results_path, output_path_genepred, output_path_genepred_faa, binphylo_path_prstax, binphylo_path_binstax, binphylo_path_binstax_max, binphylo_path_prsids, dict_hm, dict_sp, dict_nr, dict_top, swiss_fams_len_comp_dict, dict_input_motifs, dict_genes, contig_gene_dist_dict, dict_contigs_bins, dict_prs_bins, dict_bins_prs_sorted, bins_prs_tax_dict, bin_tax_dict, bin_tax_max_dict, bin_group_tax_dict, kraken_mode, binned_taxa_path, dict_seqs, only_seek_mode, only_phylo_mode, cd_hit_results_fasta_path, add_seek_info, add_taxonomy_info)
    label = "Process of information collection:"
    elpased_time = end_time_analysis(label, start_time_info_collection, output_log_file)
    time_dict["info_collection_time"] = elpased_time

    # Convert the taxonomy information to be protein based as the key of a dictionary.
    pr_tax_info_dcit = tax_info_convert(bin_group_tax_dict)

    # Writing results
    start_time_results = time.time()
    # TXT results.
    txt_results(annotation_file_txt_name, output_path_annotation, dict_hm, dict_seqs, dict_top, dict_sp, dict_nr, dict_input_motifs, dict_genes, swiss_fams_len_comp_dict, contig_gene_dist_dict, pr_tax_info_dcit, add_type, add_info)
    # EXCEL results.
    excel_results(annotation_file_excel_name, dict_hm, dict_seqs, dict_top, dict_sp, dict_nr, dict_input_motifs, dict_genes, swiss_fams_len_comp_dict, contig_gene_dist_dict, pr_tax_info_dcit, add_type, add_info)
    label = "Process of writing results:"
    elpased_time = end_time_analysis(label, start_time_results, output_log_file)
    time_dict["results_time"] = elpased_time

    # End time
    label = "End of analysis:"
    elpased_time = end_time_analysis(label, start_time, output_log_file)
    time_dict["tool_time"] = elpased_time
    write_time_dict(time_dict, time_analyis_path)

    # Close log files.
    input_log_file.close()
    output_log_file.close()
    print()
    print("----------------")
    print("Pipeline End")
    print("----------------")
    print()


if __name__ == "__main__":
    # Input and output options
    arg_input_folder = None
    arg_sra_code = False
    arg_prs_source = 1
    arg_adapters_path = "adapters.fa"
    arg_protein_db_path = ""
    arg_kraken_db_path = ""
    arg_profiles_path = ""
    arg_profiles_phylo_path = ""
    arg_profiles_broad_path = ""
    arg_swissprot_path = ""
    arg_motifs_path = ""
    arg_output_path = ""
    # Protein family options
    arg_family_code = None
    arg_family_code_phylo = None
    arg_db_name = ""
    arg_db_name_phylo = ""
    arg_input_protein_names_status = False
    arg_input_protein_names = None
    arg_input_protein_names_phylo_status = False
    arg_input_protein_names_phylo = None
    arg_name_thr = 0.5
    # General options
    # Pipeline
    arg_seek_route = 3
    arg_paired_end = True
    arg_compressed = True
    arg_contigs = False
    arg_protein_input = False
    arg_create_nr_db_status = False
    arg_prefetch_size = 20
    arg_adapters_status = "pre"
    arg_add_seek_info = True
    arg_add_taxonomy_info = True
    # FastQC
    arg_skip_fastqc = False
    # BBDuk
    arg_bbduk_max_ram = 4
    arg_clear_space = False
    # Megahit
    arg_k_list = None
    # Kraken
    arg_kraken_mode = True
    arg_kraken_threshold = "0.1"
    arg_kraken_memory_mapping = True
    # Binning
    arg_binning_tool = 1
    arg_bin_ram_ammount = 4
    arg_bin_num_contig_len = 500
    arg_bin_num_kmer = 4
    arg_comebin_batch_size = 256
    # CD-HIT
    arg_cd_hit_t = 0.99
    arg_cd_hit_mem = 4000
    # Gene prediction
    arg_genetic_code = 11
    # HMMER
    arg_val_type = "--cut_ga "
    arg_second_dom_search = True
    arg_e_value_nodom_thr = 1e-70
    # Annotation
    arg_add_type = []
    arg_add_info = []
    # Threads
    arg_thread_num = 4
    arg_pdf_threads = None
    # Processes performed after
    arg_after_trimming = False
    arg_after_alignment = False
    arg_after_binning = False
    arg_after_gene_pred = False
    arg_after_db = False
    arg_after_tm = False
    arg_after_ap = False
    # Processes performed up to
    arg_up_to_sra = False
    arg_up_to_databases = False
    arg_up_to_trimming_com = False
    arg_up_to_trimming_uncom = False
    arg_up_to_alignment = False
    # Tool enviroments
    arg_sra_env = "ps_sra_tools"
    arg_fastqc_env = "ps_fastqc"
    arg_bbduk_env = "ps_bbtools"
    arg_megahit_env = "ps_megahit"
    arg_kraken_env = "ps_kraken"
    arg_metabinner_env = "ps_metabinner"
    arg_comebin_env = "ps_comebin"
    arg_cdhit_env = "ps_cd_hit"
    arg_genepred_env = ""
    arg_hmmer_env = "ps_hmmer"
    arg_diamond_env = "ps_diamond"
    arg_taxonkit_env = "ps_taxonkit"
    arg_phobius_env = "ps_phobius"
    arg_bowtie_env = "ps_bowtie"
    # Tool paths
    arg_ana_dir_path = ""
    arg_ana_sh_path = ""
    arg_prefetch_path = ""
    arg_vdb_validate_path = ""
    arg_fastq_dump_path = ""
    arg_fastqc_path = ""
    arg_gzip_path = ""
    arg_cat_path = ""
    arg_bbduk_path = ""
    arg_megahit_path = ""
    arg_kraken_path = ""
    arg_metabinner_bin_path = ""
    arg_comebin_bin_path = ""
    arg_cd_hit_path = ""
    arg_fraggenescanrs_path = ""
    arg_hmmscan_path = ""
    arg_hmmpress_path = ""
    arg_hmmfetch_path = ""
    arg_diamond_path = ""
    arg_taxonkit_path = ""
    arg_phobius_path = ""
    arg_bowtie_build_path = ""
    arg_bowtie_path = ""
    # Options file path
    arg_options_file_path = ""
    # Variable to store the input command
    arg_input_command = ""
    # The input file is the last argument, therefore all following items in the list are parts of the filename.
    # If more than one item is in the list then join them with spaces.
    if len(sys.argv) > 1:
        arg_input_command = "{}".format(sys.argv[0])
        for i in range(1, len(sys.argv), 2):
            if sys.argv[i] == "-i" or sys.argv[i] == "--input":
                arg_input_folder = sys.argv[i+1]
                if len(sys.argv[i:]) > 1:
                    arg_input_folder = " ".join(sys.argv[i+1:])
                elif len(sys.argv[i:]) == 0:
                    print("Error. No input file was found. The input file name should be the last argument in the command. Exiting.")
                    exit()
                else:
                    arg_input_folder = sys.argv[i + 1]
            elif sys.argv[i] == "-sc" or sys.argv[i] == "--sra-code":
                arg_sra_code = sys.argv[i+1]
            elif sys.argv[i] == "-c" or sys.argv[i] == "--contigs":
                arg_contigs = sys.argv[i+1]
                arg_str = "-c"
                arg_contigs = check_i_value(arg_contigs, arg_str)
            elif sys.argv[i] == "-pi" or sys.argv[i] == "--protein-input":
                arg_protein_input = sys.argv[i+1]
                arg_str = "-pi"
                arg_protein_input = check_i_value(arg_protein_input, arg_str)
            elif sys.argv[i] == "-a" or sys.argv[i] == "--adapters":
                arg_adapters_path = sys.argv[i+1]
            elif sys.argv[i] == "-pdp" or sys.argv[i] == "--protein-database-path":
                arg_protein_db_path = sys.argv[i+1]
            elif sys.argv[i] == "-kdp" or sys.argv[i] == "--kraken-database-path":
                arg_kraken_db_path = sys.argv[i+1]
            elif sys.argv[i] == "-psp" or sys.argv[i] == "--profiles-seek-path":
                arg_profiles_path = sys.argv[i+1]
            elif sys.argv[i] == "-pyp" or sys.argv[i] == "--profiles-phylo-path":
                arg_profiles_phylo_path = sys.argv[i+1]
            elif sys.argv[i] == "-pbp" or sys.argv[i] == "--profiles-broad-path":
                arg_profiles_broad_path = sys.argv[i+1]
            elif sys.argv[i] == "-sp" or sys.argv[i] == "--swissprot-path":
                arg_swissprot_path = sys.argv[i+1]
            elif sys.argv[i] == "-mop" or sys.argv[i] == "--motifs-path":
                arg_motifs_path = sys.argv[i+1]
            elif sys.argv[i] == "-pfp" or sys.argv[i] == "--parameters-file-path":
                arg_options_file_path = sys.argv[i+1]
            elif sys.argv[i] == "-o" or sys.argv[i] == "--output":
                arg_output_path = sys.argv[i+1]
                if arg_output_path[-1] == "/":
                    arg_output_path = arg_output_path[:-1]
            elif sys.argv[i] == "-fc" or sys.argv[i] == "--family-code":
                arg_family_code = sys.argv[i+1]
            elif sys.argv[i] == "-fct" or sys.argv[i] == "--family-code-taxonomy":
                arg_family_code_phylo = sys.argv[i+1]
            elif sys.argv[i] == "-dn" or sys.argv[i] == "--database-name":
                arg_db_name = sys.argv[i+1]
            elif sys.argv[i] == "-dnt" or sys.argv[i] == "--database-name-taxonomy":
                arg_db_name_phylo = sys.argv[i+1]
            elif sys.argv[i] == "-sns" or sys.argv[i] == "--seek-names-status":
                arg_input_protein_names_status = sys.argv[i+1]
                if arg_input_protein_names_status == "False":
                    arg_input_protein_names_status = False
            elif sys.argv[i] == "-spn" or sys.argv[i] == "--seek-protein-names":
                arg_input_protein_names = sys.argv[i+1]
                arg_input_protein_names = arg_input_protein_names.replace("_", " ")
                arg_input_protein_names = arg_input_protein_names.split(",")
            elif sys.argv[i] == "-tns" or sys.argv[i] == "--taxonomy-names-status":
                arg_input_protein_names_phylo_status = sys.argv[i+1]
                if arg_input_protein_names_phylo_status == "False":
                    arg_input_protein_names_phylo_status = False
            elif sys.argv[i] == "-tpn" or sys.argv[i] == "--taxonomy-protein-names":
                arg_input_protein_names_phylo = sys.argv[i+1]
                arg_input_protein_names_phylo = arg_input_protein_names_phylo.replace("_", " ")
                arg_input_protein_names_phylo = arg_input_protein_names_phylo.split(",")
            elif sys.argv[i] == "-nt" or sys.argv[i] == "--name-threshold":
                arg_name_thr = sys.argv[i+1]
                if "." in arg_name_thr:
                    arg_name_thr = float(arg_name_thr)
                else:
                    arg_name_thr = int(arg_name_thr)
            elif sys.argv[i] == "-sr" or sys.argv[i] == "--seek-route":
                arg_seek_route = int(sys.argv[i+1])
            elif sys.argv[i] == "-p" or sys.argv[i] == "--paired-end":
                arg_paired_end = sys.argv[i+1]
                arg_str = "-p"
                arg_paired_end = check_i_value(arg_paired_end, arg_str)
            elif sys.argv[i] == "-k" or sys.argv[i] == "--compressed":
                arg_compressed = sys.argv[i+1]
                arg_str = "-k"
                arg_compressed = check_i_value(arg_compressed, arg_str)
            elif sys.argv[i] == "-fpd" or sys.argv[i] == "--filter-protein-database":
                arg_create_nr_db_status = sys.argv[i+1]
                arg_str = "-fpd"
                arg_create_nr_db_status = check_i_value(arg_create_nr_db_status, arg_str)
            elif sys.argv[i] == "-ps" or sys.argv[i] == "--prefetch-size":
                arg_prefetch_size = int(sys.argv[i+1])
            elif sys.argv[i] == "-as" or sys.argv[i] == "--adapters-status":
                arg_adapters_status = sys.argv[i+1]                
            elif sys.argv[i] == "-asi" or sys.argv[i] == "--add-seek-info":
                arg_add_seek_info = sys.argv[i+1]
                arg_str = "-osi"
                arg_add_seek_info = check_i_value(arg_add_seek_info, arg_str)
            elif sys.argv[i] == "-ati" or sys.argv[i] == "--add-taxonomy-info":
                arg_add_taxonomy_info = sys.argv[i+1]
                arg_str = "-ati"
                arg_add_taxonomy_info = check_i_value(arg_add_taxonomy_info, arg_str)
            elif sys.argv[i] == "-sf" or sys.argv[i] == "--skip-fastqc":
                arg_skip_fastqc = sys.argv[i+1]
                arg_str = "-sf"
                arg_skip_fastqc = check_i_value(arg_skip_fastqc, arg_str)
            elif sys.argv[i] == "-umr" or sys.argv[i] == "--bbduk-max-ram":
                arg_bbduk_max_ram = int(sys.argv[i+1])
            elif sys.argv[i] == "-cs" or sys.argv[i] == "--clear-space":
                arg_clear_space = sys.argv[i+1]
                arg_str = "-cs"
                arg_clear_space = check_i_value(arg_clear_space, arg_str)
            elif sys.argv[i] == "-kl" or sys.argv[i] == "--k-list":
                arg_k_list = sys.argv[i+1]
            elif sys.argv[i] == "-km" or sys.argv[i] == "--kraken-mode":
                arg_kraken_mode = sys.argv[i+1]
                arg_str = "-km"
                arg_kraken_mode = check_i_value(arg_kraken_mode, arg_str)
            elif sys.argv[i] == "-kt" or sys.argv[i] == "--kraken-threshold":
                arg_kraken_threshold = sys.argv[i+1]
            elif sys.argv[i] == "-kmm" or sys.argv[i] == "--kraken-memory-mapping":
                arg_kraken_memory_mapping = sys.argv[i+1]
                arg_str = "-kmm"
                arg_kraken_memory_mapping = check_i_value(arg_kraken_memory_mapping, arg_str)
            elif sys.argv[i] == "-bt" or sys.argv[i] == "--binning-tool":
                arg_binning_tool = int(sys.argv[i+1])
            elif sys.argv[i] == "-bmr" or sys.argv[i] == "--binning-max-ram":
                arg_bin_ram_ammount = int(sys.argv[i+1])
            elif sys.argv[i] == "-bc" or sys.argv[i] == "--bin-contig-len":
                arg_bin_num_contig_len = int(sys.argv[i+1])
            elif sys.argv[i] == "-bk" or sys.argv[i] == "--bin-kmer":
                arg_bin_num_kmer = int(sys.argv[i+1])
            elif sys.argv[i] == "-cbs" or sys.argv[i] == "--comebin-batch-size":
                arg_comebin_batch_size = int(sys.argv[i+1])
            elif sys.argv[i] == "-ct" or sys.argv[i] == "--cdhit-threshold":
                arg_cd_hit_t = float(sys.argv[i+1])
            elif sys.argv[i] == "-cmr" or sys.argv[i] == "--cd-hit-max-ram":
                arg_cd_hit_mem = int(sys.argv[i+1])
            elif sys.argv[i] == "-ge" or sys.argv[i] == "--gene-encoding":
                arg_prs_source = int(sys.argv[i+1])
            elif sys.argv[i] == "-gc" or sys.argv[i] == "--genetic-code":
                arg_genetic_code = int(sys.argv[i+1])
            elif sys.argv[i] == "-st" or sys.argv[i] == "--score-type":
                arg_val_type = sys.argv[i+1]
                if arg_val_type == "cut_ga":
                    arg_val_type = "--{} ".format(arg_val_type)
                elif arg_val_type == "default":
                    arg_val_type = ""
            elif sys.argv[i] == "-sds" or sys.argv[i] == "--second-domain-search":
                arg_second_dom_search = sys.argv[i+1]
                arg_str = "-sds"
                arg_second_dom_search = check_i_value(arg_second_dom_search, arg_str)
            elif sys.argv[i] == "-ndt" or sys.argv[i] == "--no-domains-thr":
                arg_e_value_nodom_thr = int(sys.argv[i+1])
                arg_e_value_nodom_thr = float(1*(10**(-arg_e_value_nodom_thr)))
            elif sys.argv[i] == "-at" or sys.argv[i] == "--add-type":
                arg_add_type = sys.argv[i+1]
            elif sys.argv[i] == "-ai" or sys.argv[i] == "--add-info":
                arg_add_info = sys.argv[i+1]
            elif sys.argv[i] == "-t" or sys.argv[i] == "--threads":
                arg_thread_num = int(sys.argv[i+1])
            elif sys.argv[i] == "-ft" or sys.argv[i] == "--filtering-threads":
                arg_pdf_threads = int(sys.argv[i+1])
            elif sys.argv[i] == "-afp" or sys.argv[i] == "--after-preprocessing":
                arg_after_trimming = sys.argv[i+1]
                arg_str = "-afp"
                arg_after_trimming = check_i_value(arg_after_trimming, arg_str)
            elif sys.argv[i] == "-afa" or sys.argv[i] == "--after-assembly":
                arg_after_alignment = sys.argv[i + 1]
                arg_str = "-afa"
                arg_after_alignment = check_i_value(arg_after_alignment, arg_str)
            elif sys.argv[i] == "-afg" or sys.argv[i] == "--after-gene-pred":
                arg_after_gene_pred = sys.argv[i+1]
                arg_str = "-afg"
                arg_after_gene_pred = check_i_value(arg_after_gene_pred, arg_str)
            elif sys.argv[i] == "-afb" or sys.argv[i] == "--after-binning":
                arg_after_binning = sys.argv[i+1]
                arg_str = "-afb"
                arg_after_binning = check_i_value(arg_after_binning, arg_str)
            elif sys.argv[i] == "-adb" or sys.argv[i] == "--after-db":
                arg_after_db = sys.argv[i+1]
                arg_str = "-adb"
                arg_after_db = check_i_value(arg_after_db, arg_str)
            elif sys.argv[i] == "-atp" or sys.argv[i] == "--after-topology-prediction":
                arg_after_tm = sys.argv[i+1]
                arg_str = "-atp"
                arg_after_tm = check_i_value(arg_after_tm, arg_str)
            elif sys.argv[i] == "-afr" or sys.argv[i] == "--after-analysis-processes":
                arg_after_ap = sys.argv[i+1]
                arg_str = "-afr"
                arg_after_ap = check_i_value(arg_after_ap, arg_str)
            elif sys.argv[i] == "-uts" or sys.argv[i] == "--up-to-sra":
                arg_up_to_sra = sys.argv[i+1]
                arg_str = "-uts"
                arg_up_to_sra = check_i_value(arg_up_to_sra, arg_str)
            elif sys.argv[i] == "-utd" or sys.argv[i] == "--up-to-databases":
                arg_up_to_databases = sys.argv[i+1]
                arg_str = "-utd"
                arg_up_to_databases = check_i_value(arg_up_to_databases, arg_str)
            elif sys.argv[i] == "-utpc" or sys.argv[i] == "--up-to-preprocessing-com":
                arg_up_to_trimming_com = sys.argv[i+1]
                arg_str = "-uttc"
                arg_up_to_trimming_com = check_i_value(arg_up_to_trimming_com, arg_str)
            elif sys.argv[i] == "-utpu" or sys.argv[i] == "--up-to-preprocessing-uncom":
                arg_up_to_trimming_uncom = sys.argv[i+1]
                arg_str = "-uttu"
                arg_up_to_trimming_uncom = check_i_value(arg_up_to_trimming_uncom, arg_str)
            elif sys.argv[i] == "-uta" or sys.argv[i] == "--up-to-assembly":
                arg_up_to_alignment = sys.argv[i+1]
                arg_str = "-uta"
                arg_up_to_alignment = check_i_value(arg_up_to_alignment, arg_str)
            elif sys.argv[i] == "-sen" or sys.argv[i] == "--sra-env":
                arg_sra_env = sys.argv[i+1]
                if arg_sra_env in ["None", "none"]:
                    arg_sra_env = ""
            elif sys.argv[i] == "-fen" or sys.argv[i] == "--fastqc-env":
                arg_fastqc_env = sys.argv[i+1]
                if arg_fastqc_env in ["None", "none"]:
                    arg_fastqc_env = ""
            elif sys.argv[i] == "-uen" or sys.argv[i] == "--bbtools-env":
                arg_bbduk_env = sys.argv[i+1]
                if arg_bbduk_env in ["None", "none"]:
                    arg_bbduk_env = ""
            elif sys.argv[i] == "-men" or sys.argv[i] == "--megahit-env":
                arg_megahit_env = sys.argv[i+1]
                if arg_megahit_env in ["None", "none"]:
                    arg_megahit_env = ""
            elif sys.argv[i] == "-ken" or sys.argv[i] == "--kraken-env":
                arg_kraken_env = sys.argv[i+1]
                if arg_kraken_env in ["None", "none"]:
                    arg_kraken_env = ""
            elif sys.argv[i] == "-nen" or sys.argv[i] == "--metabinner-env":
                arg_metabinner_env = sys.argv[i+1]
                if arg_metabinner_env in ["None", "none"]:
                    arg_metabinner_env = ""
            elif sys.argv[i] == "-cen" or sys.argv[i] == "--comebin-env":
                arg_comebin_env = sys.argv[i+1]
                if arg_comebin_env in ["None", "none"]:
                    arg_comebin_env = ""
            elif sys.argv[i] == "-ien" or sys.argv[i] == "--cdhit-env":
                arg_cdhit_env = sys.argv[i+1]
                if arg_cdhit_env in ["None", "none"]:
                    arg_cdhit_env = ""
            elif sys.argv[i] == "-gen" or sys.argv[i] == "--genepred-env":
                arg_genepred_env = sys.argv[i+1]
                if arg_genepred_env in ["None", "none"]:
                    arg_genepred_env = ""
            elif sys.argv[i] == "-hen" or sys.argv[i] == "--hmmer-env":
                arg_hmmer_env = sys.argv[i+1]
                if arg_hmmer_env in ["None", "none"]:
                    arg_hmmer_env = ""
            elif sys.argv[i] == "-den" or sys.argv[i] == "--dimaond-env":
                arg_diamond_env = sys.argv[i+1]
                if arg_diamond_env in ["None", "none"]:
                    arg_diamond_env = ""
            elif sys.argv[i] == "-ten" or sys.argv[i] == "--taxonkit-env":
                arg_taxonkit_env = sys.argv[i+1]
                if arg_taxonkit_env in ["None", "none"]:
                    arg_taxonkit_env = ""
            elif sys.argv[i] == "-pen" or sys.argv[i] == "--phobius-env":
                arg_phobius_env = sys.argv[i+1]
                if arg_phobius_env in ["None", "none"]:
                    arg_phobius_env = ""
            elif sys.argv[i] == "-ben" or sys.argv[i] == "--bowtie-env":
                arg_bowtie_env = sys.argv[i+1]
                if arg_bowtie_env in ["None", "none"]:
                    arg_bowtie_env = ""
            elif sys.argv[i] == "-adp" or sys.argv[i] == "--anaconda-dir-path":
                arg_ana_dir_path = sys.argv[i+1]
            elif sys.argv[i] == "-asp" or sys.argv[i] == "--anaconda-sh-path":
                arg_ana_sh_path = sys.argv[i+1]
            elif sys.argv[i] == "-rfp" or sys.argv[i] == "--prefetch-path":
                arg_prefetch_path = sys.argv[i+1]
            elif sys.argv[i] == "-vvp" or sys.argv[i] == "--vdb_validate-path":
                arg_vdb_validate_path = sys.argv[i+1]
            elif sys.argv[i] == "-fdp" or sys.argv[i] == "--fastq-dump-path":
                arg_fastq_dump_path = sys.argv[i+1]
            elif sys.argv[i] == "-fp" or sys.argv[i] == "--fastqc-path":
                arg_fastqc_path = sys.argv[i+1]
            elif sys.argv[i] == "-gzp" or sys.argv[i] == "--gzip-path":
                arg_gzip_path = sys.argv[i+1]
            elif sys.argv[i] == "-ctp" or sys.argv[i] == "--cat-path":
                arg_cat_path = sys.argv[i+1]
            elif sys.argv[i] == "-bdp" or sys.argv[i] == "--bbduk-path":
                arg_bbduk_path = sys.argv[i+1]
            elif sys.argv[i] == "-mp" or sys.argv[i] == "--megahit-path":
                arg_megahit_path = sys.argv[i+1]
            elif sys.argv[i] == "-kp" or sys.argv[i] == "--kraken-path":
                arg_kraken_path = sys.argv[i+1]
            elif sys.argv[i] == "-bfp" or sys.argv[i] == "--binner-folder-path":
                arg_metabinner_bin_path = sys.argv[i+1]
            elif sys.argv[i] == "-cfp" or sys.argv[i] == "--comebin-folder-path":
                arg_comebin_bin_path = sys.argv[i+1]
            elif sys.argv[i] == "-chp" or sys.argv[i] == "--cd-hit-path":
                arg_cd_hit_path = sys.argv[i+1]
            elif sys.argv[i] == "-fgp" or sys.argv[i] == "--fraggenescars-path":
                arg_fraggenescanrs_path = sys.argv[i+1]
            elif sys.argv[i] == "-hp" or sys.argv[i] == "--hmmscan-path":
                arg_hmmscan_path = sys.argv[i+1]
            elif sys.argv[i] == "-hpp" or sys.argv[i] == "--hmmpress-path":
                arg_hmmpress_path = sys.argv[i+1]
            elif sys.argv[i] == "-hfp" or sys.argv[i] == "--hmmfetch-path":
                arg_hmmfetch_path = sys.argv[i+1]
            elif sys.argv[i] == "-dp" or sys.argv[i] == "--diamond-path":
                arg_diamond_path = sys.argv[i+1]
            elif sys.argv[i] == "-tkp" or sys.argv[i] == "--taxonkit-path":
                arg_taxonkit_path = sys.argv[i+1]
            elif sys.argv[i] == "-php" or sys.argv[i] == "--phobius-folder-path":
                arg_phobius_path = sys.argv[i+1]
            elif sys.argv[i] == "-bbp" or sys.argv[i] == "--bowtie-build-path":
                arg_bowtie_build_path = sys.argv[i+1]
            elif sys.argv[i] == "-bwp" or sys.argv[i] == "--bowtie-path":
                arg_bowtie_path = sys.argv[i+1]
            elif sys.argv[i] == "-h" or sys.argv[i] == "--help":
                help_message()
                exit()
            arg_input_command = "{} {} {}".format(arg_input_command, sys.argv[i], sys.argv[i+1])
    enzannmtg(arg_input_folder, arg_sra_code, arg_contigs, arg_protein_input, arg_adapters_path, arg_protein_db_path, arg_kraken_db_path, arg_profiles_path, arg_profiles_phylo_path, arg_profiles_broad_path, arg_swissprot_path, arg_motifs_path, arg_options_file_path, arg_output_path, arg_family_code, arg_family_code_phylo, arg_db_name, arg_db_name_phylo, arg_input_protein_names_status, arg_input_protein_names, arg_input_protein_names_phylo_status, arg_input_protein_names_phylo, arg_name_thr, arg_seek_route, arg_paired_end, arg_compressed, arg_create_nr_db_status, arg_prefetch_size, arg_adapters_status, arg_add_seek_info, arg_add_taxonomy_info, arg_skip_fastqc, arg_bbduk_max_ram, arg_clear_space, arg_k_list, arg_kraken_mode, arg_kraken_threshold, arg_kraken_memory_mapping, arg_binning_tool, arg_bin_ram_ammount, arg_bin_num_contig_len, arg_bin_num_kmer, arg_comebin_batch_size, arg_cd_hit_t, arg_cd_hit_mem, arg_prs_source, arg_genetic_code, arg_val_type, arg_second_dom_search, arg_e_value_nodom_thr, arg_add_type, arg_add_info, arg_thread_num, arg_pdf_threads, arg_after_trimming, arg_after_alignment, arg_after_gene_pred, arg_after_binning, arg_after_db, arg_after_tm, arg_after_ap, arg_up_to_sra, arg_up_to_databases, arg_up_to_trimming_com, arg_up_to_trimming_uncom, arg_up_to_alignment, arg_sra_env, arg_fastqc_env, arg_bbduk_env, arg_megahit_env, arg_kraken_env, arg_metabinner_env, arg_comebin_env, arg_cdhit_env, arg_genepred_env, arg_hmmer_env, arg_diamond_env, arg_taxonkit_env, arg_phobius_env, arg_bowtie_env, arg_ana_dir_path, arg_ana_sh_path, arg_prefetch_path, arg_vdb_validate_path, arg_fastq_dump_path, arg_fastqc_path, arg_gzip_path, arg_cat_path, arg_bbduk_path, arg_megahit_path, arg_kraken_path, arg_metabinner_bin_path, arg_comebin_bin_path, arg_fraggenescanrs_path, arg_hmmscan_path, arg_hmmpress_path, arg_hmmfetch_path, arg_diamond_path, arg_cd_hit_path, arg_taxonkit_path, arg_phobius_path, arg_bowtie_build_path, arg_bowtie_path, arg_input_command)
